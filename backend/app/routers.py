from fastapi import APIRouter, Depends, HTTPException, status, Response, Request, UploadFile, File
from pydantic import BaseModel, validator
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
import re
import logging
from datetime import datetime, timedelta, date

# Configure logging
logger = logging.getLogger(__name__)

from .auth import authenticate_user, create_access_token, get_current_user, require_role
from .db import get_db
from .models import Product, User, Party, CompanySettings, Invoice, InvoiceItem, StockLedgerEntry, Purchase, PurchaseItem, Payment, PurchasePayment, Expense, AuditTrail, RecurringInvoiceTemplate, RecurringInvoiceTemplateItem, RecurringInvoice, PurchaseOrder, PurchaseOrderItem, InvoiceTemplate
from .audit import AuditService
from .gst import money, split_gst
from .gst_reports import generate_gstr1_report, generate_gstr3b_report
from .currency import get_exchange_rate, get_supported_currencies, format_currency, format_currency_for_pdf
from .recurring_invoices import RecurringInvoiceService, generate_recurring_invoices
from .purchase_orders import PurchaseOrderService, convert_po_to_purchase
from .profitpath_service import ProfitPathService
from .payment_scheduler import PaymentScheduler, PaymentStatus, PaymentReminderType
from .inventory_manager import InventoryManager, StockValuationMethod
from .financial_reports import FinancialReports, ReportType
from .branding import router as branding_router
from .dental import router as dental_router
from .manufacturing import router as manufacturing_router
from decimal import Decimal
from .emailer import send_email, create_invoice_email_template, create_purchase_email_template
from fastapi import Query
import json
import calendar
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.colors import black, white, grey, darkblue
import os
import base64


# Indian States for GST Compliance
INDIAN_STATES = {
    "Andhra Pradesh": "37",
    "Arunachal Pradesh": "12",
    "Assam": "18",
    "Bihar": "10",
    "Chhattisgarh": "22",
    "Goa": "30",
    "Gujarat": "24",
    "Haryana": "06",
    "Himachal Pradesh": "02",
    "Jharkhand": "20",
    "Karnataka": "29",
    "Kerala": "32",
    "Madhya Pradesh": "23",
    "Maharashtra": "27",
    "Manipur": "14",
    "Meghalaya": "17",
    "Mizoram": "15",
    "Nagaland": "13",
    "Odisha": "21",
    "Punjab": "03",
    "Rajasthan": "08",
    "Sikkim": "11",
    "Tamil Nadu": "33",
    "Telangana": "36",
    "Tripura": "16",
    "Uttar Pradesh": "09",
    "Uttarakhand": "05",
    "West Bengal": "19",
    "Delhi": "07",
    "Jammu and Kashmir": "01",
    "Ladakh": "38",
    "Chandigarh": "04",
    "Dadra and Nagar Haveli": "26",
    "Daman and Diu": "25",
    "Lakshadweep": "31",
    "Puducherry": "34",
    "Andaman and Nicobar Islands": "35"
}

# Utility functions for invoice calculations
def calculate_due_date(invoice_date: str, terms: str) -> datetime:
    """Calculate due date based on terms"""
    date_obj = datetime.fromisoformat(invoice_date.replace('Z', '+00:00'))
    
    if terms == "Due on Receipt":
        return date_obj
    elif terms == "15 days":
        return date_obj + timedelta(days=15)
    elif terms == "30 days":
        return date_obj + timedelta(days=30)
    elif terms == "45 days":
        return date_obj + timedelta(days=45)
    elif terms == "60 days":
        return date_obj + timedelta(days=60)
    elif terms == "90 days":
        return date_obj + timedelta(days=90)
    else:
        return date_obj


def number_to_words(amount: float) -> str:
    """Convert number to words (simplified version)"""
    if amount == 0:
        return "Zero Rupees Only"
    
    # This is a simplified version - in production, you'd want a more comprehensive implementation
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def convert_less_than_one_thousand(n):
        if n == 0:
            return ""
        elif n < 10:
            return units[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        else:
            return units[n // 100] + " Hundred" + (" " + convert_less_than_one_thousand(n % 100) if n % 100 != 0 else "")
    
    rupees = int(amount)
    paise = int((amount - rupees) * 100)
    
    if rupees == 0:
        return f"{paise} Paise Only"
    elif paise == 0:
        return f"{convert_less_than_one_thousand(rupees)} Rupees Only"
    else:
        return f"{convert_less_than_one_thousand(rupees)} Rupees and {paise} Paise Only"


api = APIRouter()


class LoginRequest(BaseModel):
    username: str
    password: str


@api.post("/auth/login")
def login(payload: LoginRequest, request: Request, db: Session = Depends(get_db)):
    user = authenticate_user(db, payload.username, payload.password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Log successful login
    AuditService.log_login(db, user, request)
    
    token = create_access_token(user.username)
    return {"access_token": token, "token_type": "bearer"}


class ProductOut(BaseModel):
    id: int
    name: str
    description: str | None
    item_type: str
    sales_price: float
    purchase_price: float | None
    stock: int
    sku: str | None
    unit: str
    supplier: str | None
    category: str | None
    notes: str | None
    hsn: str | None
    gst_rate: float | None
    is_active: bool

    class Config:
        from_attributes = True


@api.get("/products", response_model=list[ProductOut])
def list_products(
    search: str | None = None,
    category: str | None = None,
    item_type: str | None = None,
    gst_rate: float | None = None,
    supplier: str | None = None,
    stock_level: str | None = None,
    price_min: float | None = None,
    price_max: float | None = None,
    status: str | None = None,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    query = db.query(Product)
    
    if search:
        search_filter = (
            Product.name.ilike(f"%{search}%") |
            Product.description.ilike(f"%{search}%") |
            Product.sku.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    if category:
        query = query.filter(Product.category == category)
    
    if item_type:
        query = query.filter(Product.item_type == item_type)
    
    if gst_rate is not None:
        query = query.filter(Product.gst_rate == gst_rate)
    
    if supplier:
        query = query.filter(Product.supplier.ilike(f"%{supplier}%"))
    
    if stock_level:
        if stock_level == 'low_stock':
            query = query.filter(Product.stock < 10)
        elif stock_level == 'out_of_stock':
            query = query.filter(Product.stock == 0)
        elif stock_level == 'in_stock':
            query = query.filter(Product.stock > 0)
    
    if price_min is not None:
        query = query.filter(Product.sales_price >= price_min)
    
    if price_max is not None:
        query = query.filter(Product.sales_price <= price_max)
    
    if status:
        if status == 'active':
            query = query.filter(Product.is_active == True)
        elif status == 'inactive':
            query = query.filter(Product.is_active == False)
    else:
        query = query.filter(Product.is_active == True)
    
    return query.order_by(Product.id).all()


class ProductCreate(BaseModel):
    name: str
    description: str | None = None
    item_type: str = "tradable"
    sales_price: float
    purchase_price: float | None = None
    stock: int = 0
    sku: str | None = None
    unit: str = "Pcs"
    supplier: str | None = None
    category: str | None = None
    notes: str | None = None
    hsn: str | None = None
    gst_rate: float | None = None


class ProductUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    item_type: str | None = None
    sales_price: float | None = None
    purchase_price: float | None = None
    stock: int | None = None
    sku: str | None = None
    unit: str | None = None
    supplier: str | None = None
    category: str | None = None
    notes: str | None = None
    hsn: str | None = None
    gst_rate: float | None = None
    is_active: bool | None = None


@api.post("/products", response_model=ProductOut, status_code=status.HTTP_201_CREATED)
def create_product(payload: ProductCreate, _: User = Depends(require_role("Admin")), db: Session = Depends(get_db)):
    try:
        # Validation
        if not payload.name or len(payload.name.strip()) == 0:
            raise HTTPException(status_code=400, detail="Name is required")
        if len(payload.name) > 100:
            raise HTTPException(status_code=400, detail="Name must be 100 characters or less")
        if not re.match(r'^[a-zA-Z0-9\s]+$', payload.name):
            raise HTTPException(status_code=400, detail="Name must be alphanumeric with spaces only")
        
        if payload.description and len(payload.description) > 200:
            raise HTTPException(status_code=400, detail="Description must be 200 characters or less")
        
        if payload.sales_price < 0 or payload.sales_price > 999999.99:
            raise HTTPException(status_code=400, detail="Sales price must be between 0 and 999999.99")
        
        if payload.purchase_price is not None:
            if payload.purchase_price < 0 or payload.purchase_price > 999999.99:
                raise HTTPException(status_code=400, detail="Purchase price must be between 0 and 999999.99")
        
        if payload.stock < 0 or payload.stock > 999999:
            raise HTTPException(status_code=400, detail="Stock must be between 0 and 999999")
        
        if payload.sku and len(payload.sku) > 50:
            raise HTTPException(status_code=400, detail="SKU must be 50 characters or less")
        if payload.sku and not re.match(r'^[a-zA-Z0-9\s]+$', payload.sku):
            raise HTTPException(status_code=400, detail="SKU must be alphanumeric with spaces only")
        
        if payload.supplier and len(payload.supplier) > 100:
            raise HTTPException(status_code=400, detail="Supplier must be 100 characters or less")
        
        if payload.category and len(payload.category) > 100:
            raise HTTPException(status_code=400, detail="Category must be 100 characters or less")
        
        # Item type validation
        if payload.item_type not in ["tradable", "consumable", "manufactured"]:
            raise HTTPException(status_code=400, detail="Item type must be tradable, consumable, or manufactured")
        
        # Check if SKU already exists (only if SKU is provided)
        if payload.sku:
            exists = db.query(Product).filter(Product.sku == payload.sku).first()
            if exists:
                raise HTTPException(status_code=400, detail="SKU already exists")
        
        product = Product(**payload.model_dump())
        db.add(product)
        db.commit()
        db.refresh(product)
        
        # Log product creation
        AuditService.log_create(
            db=db,
            user=_,  # Current user from dependency
            table_name="products",
            record_id=product.id,
            new_values=payload.model_dump(),
            request=None  # We don't have request context here
        )
        
        return product
    except Exception as e:
        db.rollback()
        # Check for specific database constraint violations
        error_str = str(e).lower()
        if "unique constraint" in error_str and "sku" in error_str:
            raise HTTPException(status_code=400, detail="SKU already exists")
        elif "duplicate key value" in error_str and "sku" in error_str:
            raise HTTPException(status_code=400, detail="SKU already exists")
        elif "not null constraint" in error_str:
            raise HTTPException(status_code=400, detail="Required fields cannot be empty")
        elif "check constraint" in error_str:
            raise HTTPException(status_code=400, detail="Invalid data provided")
        else:
            # Log the actual error for debugging
            print(f"Database error: {e}")
            raise HTTPException(status_code=500, detail="Database error occurred")


@api.put("/products/{product_id}", response_model=ProductOut)
def update_product(product_id: int, payload: ProductUpdate, _: User = Depends(require_role("Admin")), db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(product, field, value)
    db.commit()
    db.refresh(product)
    return product


@api.patch("/products/{product_id}/toggle", response_model=ProductOut)
def toggle_product(product_id: int, _: User = Depends(require_role("Admin")), db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Not found")
    product.is_active = not product.is_active
    db.commit()
    db.refresh(product)
    return product


class StockAdjustmentIn(BaseModel):
    quantity: float
    purchase_price: float
    sales_price: float
    date_of_receipt: str  # ISO date string
    reference_bill_number: str | None = None
    notes: str | None = None


@api.post("/products/{product_id}/stock", response_model=ProductOut, status_code=status.HTTP_201_CREATED)
def add_stock_to_product(product_id: int, payload: StockAdjustmentIn, _: User = Depends(require_role("Admin")), db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Add stock to product
    product.stock += payload.quantity
    
    # Update prices if provided
    if payload.purchase_price > 0:
        product.purchase_price = payload.purchase_price
    if payload.sales_price > 0:
        product.sales_price = payload.sales_price
    
    # Create stock ledger entry
    from datetime import datetime
    stock_entry = StockLedgerEntry(
        product_id=product_id,
        qty=payload.quantity,
        entry_type="in",
        ref_type="stock_adjustment",
        ref_id=0
    )
    db.add(stock_entry)
    
    db.commit()
    db.refresh(product)
    return product
class InvoiceItemIn(BaseModel):
    product_id: int
    qty: float
    rate: float
    discount: float = 0
    discount_type: str = "Percentage"  # Percentage, Fixed
    description: str | None = None
    hsn_code: str | None = None


class PaymentIn(BaseModel):
    payment_date: str  # ISO date string
    payment_amount: float
    payment_method: str
    account_head: str
    reference_number: str | None = None
    notes: str | None = None


class PaymentOut(BaseModel):
    id: int
    invoice_id: int
    payment_date: datetime
    payment_amount: float
    payment_method: str
    reference_number: str | None
    notes: str | None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class PurchasePaymentOut(BaseModel):
    id: int
    purchase_id: int
    payment_amount: float
    payment_method: str
    account_head: str
    reference_number: str | None
    payment_date: str  # ISO date string
    notes: str | None
    vendor_name: str
    purchase_number: str

    class Config:
        from_attributes = True


class StockLedgerEntryOut(BaseModel):
    id: int
    product_id: int
    entry_type: str
    qty: float
    reference_bill_number: str | None
    notes: str | None
    created_at: str  # ISO date string
    product_name: str

    class Config:
        from_attributes = True


class InvoiceOut(BaseModel):
    id: int
    customer_id: int
    supplier_id: int
    invoice_no: str
    date: datetime
    due_date: datetime
    terms: str
    
    # Invoice Details
    invoice_type: str
    currency: str
    exchange_rate: Decimal
    status: str
    
    # GST Compliance Fields
    place_of_supply: str
    place_of_supply_state_code: str
    eway_bill_number: str | None
    reverse_charge: bool
    export_supply: bool
    
    # Address Details
    bill_to_address: str
    ship_to_address: str
    
    # Amount Details
    taxable_value: float
    total_discount: float
    cgst: float
    sgst: float
    igst: float
    utgst: float
    cess: float
    round_off: float
    grand_total: float
    
    # Additional Fields
    notes: str | None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class InvoiceCreate(BaseModel):
    customer_id: int
    supplier_id: int
    invoice_no: str | None = None  # Optional, will auto-generate if not provided
    date: str  # ISO date string
    due_date: str | None = None  # Optional, will auto-calculate if not provided
    terms: str = "Due on Receipt"
    
    # Invoice Details
    invoice_type: str = "Invoice"
    currency: str = "INR"
    exchange_rate: Decimal = Decimal('1.0')
    
    # GST Compliance Fields
    place_of_supply: str
    place_of_supply_state_code: str
    eway_bill_number: str | None = None
    reverse_charge: bool = False
    export_supply: bool = False
    
    # Address Details
    bill_to_address: str
    ship_to_address: str
    
    # Items and Notes
    items: list[InvoiceItemIn]
    notes: str | None = None


def _next_invoice_no(db: Session) -> str:
    """Generate next invoice number with FY prefix and auto-generated sequence"""
    from datetime import datetime
    
    # Get current financial year (April to March)
    current_date = datetime.now()
    if current_date.month >= 4:
        fy_year = current_date.year
    else:
        fy_year = current_date.year - 1
    
    fy_prefix = f"FY{fy_year}"
    
    # Find the last invoice number for this financial year
    last_invoice = db.query(Invoice).filter(
        Invoice.invoice_no.like(f"{fy_prefix}/INV-%")
    ).order_by(Invoice.invoice_no.desc()).first()
    
    if last_invoice:
        # Extract sequence number from last invoice
        try:
            last_seq = int(last_invoice.invoice_no.split('-')[-1])
            seq = last_seq + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    
    # Format as FY<year>/INV-<4 digit sequence>
    invoice_no = f"{fy_prefix}/INV-{seq:04d}"
    
    # Ensure total length doesn't exceed 16 characters
    if len(invoice_no) > 16:
        # Truncate if necessary
        invoice_no = invoice_no[:16]
    
    return invoice_no
    
    # Ensure final length doesn't exceed 16
    result = f"{prefix}{seq_str}"
    if len(result) > 16:
        result = result[:16]
    
    return result


@api.post('/invoices', response_model=InvoiceOut, status_code=status.HTTP_201_CREATED)
def create_invoice(payload: InvoiceCreate, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Validation
    if payload.invoice_no and len(payload.invoice_no) > 16:
        raise HTTPException(status_code=400, detail="Invoice number must be 16 characters or less as per GST law")
    if payload.invoice_no and not re.match(r'^[a-zA-Z0-9\s\-/]+$', payload.invoice_no):
        raise HTTPException(status_code=400, detail="Invoice number must be alphanumeric with spaces, hyphens, and forward slashes only")
    
    # Check for duplicate invoice number (AC1)
    if payload.invoice_no:
        existing_invoice = db.query(Invoice).filter(Invoice.invoice_no == payload.invoice_no).first()
        if existing_invoice:
            raise HTTPException(status_code=400, detail="Invoice number already exists")
    if len(payload.bill_to_address) > 200:
        raise HTTPException(status_code=400, detail="Bill to address must be 200 characters or less")
    if len(payload.ship_to_address) > 200:
        raise HTTPException(status_code=400, detail="Ship to address must be 200 characters or less")
    if payload.notes and len(payload.notes) > 200:
        raise HTTPException(status_code=400, detail="Notes must be 200 characters or less")
    if payload.eway_bill_number and len(payload.eway_bill_number) > 50:
        raise HTTPException(status_code=400, detail="E-way bill number must be 50 characters or less")
    if payload.eway_bill_number and not re.match(r'^[0-9]+$', payload.eway_bill_number):
        raise HTTPException(status_code=400, detail="E-way bill number must contain only numbers")
    if not payload.place_of_supply:
        raise HTTPException(status_code=400, detail="Place of supply is mandatory as per GST law")
    if not payload.place_of_supply_state_code:
        raise HTTPException(status_code=400, detail="Place of supply state code is mandatory as per GST law")
    
    # Enhanced required fields validation (AC10)
    if not payload.bill_to_address:
        raise HTTPException(status_code=400, detail="Bill to address is required")
    if not payload.ship_to_address:
        raise HTTPException(status_code=400, detail="Ship to address is required")
    if not payload.date:
        raise HTTPException(status_code=400, detail="Invoice date is required")
    
    # Validate date format and future date (AC14)
    try:
        invoice_date = datetime.fromisoformat(payload.date.replace('Z', '+00:00'))
        if invoice_date > datetime.now():
            raise HTTPException(status_code=400, detail="Invoice date cannot be in the future")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid invoice date format")
    
    company = db.query(CompanySettings).first()
    customer = db.query(Party).filter(Party.id == payload.customer_id).first()
    supplier = db.query(Party).filter(Party.id == payload.supplier_id).first()
    
    if not customer:
        raise HTTPException(status_code=400, detail='Customer not found')
    if not supplier:
        raise HTTPException(status_code=400, detail='Supplier not found')
    
    # Validate company profile exists (AC8)
    if not company:
        raise HTTPException(status_code=400, detail='Company profile not configured')
    
    # Validate GSTIN format if provided (AC9)
    if hasattr(customer, 'gstin') and customer.gstin:
        if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', customer.gstin):
            raise HTTPException(status_code=400, detail='Invalid customer GSTIN format')
    
    if hasattr(company, 'gstin') and company.gstin:
        if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', company.gstin):
            raise HTTPException(status_code=400, detail='Invalid company GSTIN format')
    
    # Generate invoice number if not provided
    invoice_no = payload.invoice_no if payload.invoice_no else _next_invoice_no(db)
    
    # Calculate due date
    if payload.due_date:
        due_date = datetime.fromisoformat(payload.due_date.replace('Z', '+00:00'))
    else:
        due_date = calculate_due_date(payload.date, payload.terms)
    
    intra = company and payload.place_of_supply == company.state

    # Check if GST is enabled for the customer
    gst_enabled = customer.gst_enabled if hasattr(customer, 'gst_enabled') else True

    taxable_total = money(0)
    discount_total = money(0)
    cgst_total = money(0)
    sgst_total = money(0)
    igst_total = money(0)
    utgst_total = money(0)
    cess_total = money(0)

    inv = Invoice(
        customer_id=customer.id,
        supplier_id=supplier.id,
        invoice_no=invoice_no,
        date=datetime.fromisoformat(payload.date.replace('Z', '+00:00')),
        due_date=due_date,
        terms=payload.terms,
        invoice_type=payload.invoice_type,
        currency=payload.currency,
        exchange_rate=payload.exchange_rate if hasattr(payload, 'exchange_rate') else Decimal('1.0'),
        place_of_supply=payload.place_of_supply,
        place_of_supply_state_code=payload.place_of_supply_state_code,
        eway_bill_number=payload.eway_bill_number,
        reverse_charge=payload.reverse_charge,
        export_supply=payload.export_supply,
        bill_to_address=payload.bill_to_address,
        ship_to_address=payload.ship_to_address,
        taxable_value=money(0),
        total_discount=money(0),
        cgst=money(0), sgst=money(0), igst=money(0),
        utgst=money(0), cess=money(0), round_off=money(0),
        grand_total=money(0),
        paid_amount=money(0),
        balance_amount=money(0),
        notes=payload.notes,
        status="Draft"
    )
    db.add(inv)
    db.flush()

    for it in payload.items:
        prod = db.query(Product).filter(Product.id == it.product_id).first()
        if not prod:
            raise HTTPException(status_code=400, detail='Invalid product')
        
        # Calculate line item amounts
        line_total = money(Decimal(it.qty) * Decimal(it.rate))
        
        # Apply discount
        if it.discount > 0:
            if it.discount_type == "Percentage":
                discount_amount = money(line_total * Decimal(it.discount) / Decimal(100))
            else:  # Fixed
                discount_amount = money(Decimal(it.discount))
            line_total -= discount_amount
            discount_total += discount_amount
        
        # Calculate GST only if enabled for customer
        if gst_enabled and prod.gst_rate:
            cgst, sgst, igst = split_gst(line_total, prod.gst_rate, bool(intra), gst_enabled)
        else:
            cgst, sgst, igst = money(0), money(0), money(0)
        
        # Use provided description and HSN code or fall back to product defaults
        description = it.description if it.description else prod.name
        hsn_code = it.hsn_code if it.hsn_code else prod.hsn
        
        item = InvoiceItem(
            invoice_id=inv.id,
            product_id=prod.id,
            description=description,
            hsn_code=hsn_code,
            qty=it.qty,
            rate=money(it.rate),
            discount=money(it.discount),
            discount_type=it.discount_type,
            taxable_value=line_total,
            gst_rate=prod.gst_rate,
            cgst=cgst, sgst=sgst, igst=igst,
            utgst=money(0), cess=money(0),  # These can be calculated based on specific requirements
            amount=line_total + cgst + sgst + igst
        )
        db.add(item)
        # stock out for sale
        db.add(StockLedgerEntry(product_id=prod.id, qty=it.qty, entry_type='out', ref_type='invoice', ref_id=inv.id))
        taxable_total += line_total
        cgst_total += cgst
        sgst_total += sgst
        igst_total += igst

    # Calculate round off
    subtotal = taxable_total + cgst_total + sgst_total + igst_total + utgst_total + cess_total
    round_off = money(round(subtotal) - subtotal)
    
    inv.taxable_value = taxable_total
    inv.total_discount = discount_total
    inv.cgst = cgst_total
    inv.sgst = sgst_total
    inv.igst = igst_total
    inv.utgst = utgst_total
    inv.cess = cess_total
    inv.round_off = round_off
    inv.grand_total = money(subtotal + round_off)
    inv.balance_amount = money(subtotal + round_off)
    db.commit()
    db.refresh(inv)
    return inv


@api.get('/invoices/{invoice_id}/pdf')
def invoice_pdf(invoice_id: int, template_id: int | None = None, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    # Get template
    if template_id:
        template = db.query(InvoiceTemplate).filter(InvoiceTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail='Invoice template not found')
    else:
        # Get default template
        template = db.query(InvoiceTemplate).filter(
            InvoiceTemplate.is_default == True,
            InvoiceTemplate.is_active == True
        ).first()
        if not template:
            # Create default template if none exists
            template = InvoiceTemplate(
                name="Default Professional",
                description="Default professional invoice template",
                template_type="professional",
                is_default=True
            )
            db.add(template)
            db.commit()
            db.refresh(template)
    
    # Get related data
    company = db.query(CompanySettings).first()
    customer = db.query(Party).filter(Party.id == inv.customer_id).first()
    supplier = db.query(Party).filter(Party.id == inv.supplier_id).first()
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == inv.id).all()
    
    # Create PDF buffer
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Register fonts that support Unicode characters (including ₹ symbol)
    try:
        # Try to register a Unicode-supporting font
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        unicode_font = 'DejaVuSans'
    except:
        try:
            # Fallback to Arial Unicode MS if available
            pdfmetrics.registerFont(TTFont('ArialUnicode', '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf'))
            unicode_font = 'ArialUnicode'
        except:
            # Ultimate fallback to default fonts
            unicode_font = 'Helvetica'
    
    # Define styles based on template
    styles = getSampleStyleSheet()
    
    # Convert hex colors to reportlab colors
    def hex_to_color(hex_color):
        hex_color = hex_color.lstrip('#')
        return colors.HexColor(f'#{hex_color}')
    
    primary_color = hex_to_color(template.primary_color)
    secondary_color = hex_to_color(template.secondary_color)
    accent_color = hex_to_color(template.accent_color)
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=template.header_font_size,
        textColor=primary_color,
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName=template.header_font
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=template.body_font_size + 2,
        textColor=primary_color,
        spaceAfter=6,
        fontName=template.header_font
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=template.body_font_size,
        spaceAfter=3,
        fontName=template.body_font
    )
    
    # Build PDF content
    story = []
    
    # Header - Company Details
    if template.show_company_details and company:
        story.append(Paragraph(f"<b>{company.name}</b>", title_style))
        story.append(Paragraph(f"GSTIN: {company.gstin}", normal_style))
        story.append(Paragraph(f"State: {company.state} - {company.state_code}", normal_style))
    elif template.show_company_details:
        story.append(Paragraph("<b>ProfitPath</b>", title_style))
        story.append(Paragraph("Track Your Success, Step by Step", normal_style))
    
    story.append(Spacer(1, 20))
    
    # Invoice Header
    story.append(Paragraph(f"<b>{template.header_text}</b>", heading_style))
    
    # Invoice Details Table
    invoice_data = [
        ['Invoice No:', inv.invoice_no, 'Date:', inv.date.strftime('%d/%m/%Y')],
        ['Due Date:', inv.due_date.strftime('%d/%m/%Y'), 'Terms:', inv.terms],
        ['Place of Supply:', inv.place_of_supply, 'State Code:', inv.place_of_supply_state_code]
    ]
    
    if inv.eway_bill_number:
        invoice_data.append(['E-way Bill No:', inv.eway_bill_number, '', ''])
    
    invoice_table = Table(invoice_data, colWidths=[2*cm, 6*cm, 2*cm, 6*cm])
    invoice_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), template.body_font),
        ('FONTSIZE', (0, 0), (-1, -1), template.body_font_size - 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, 0), (0, -1), secondary_color),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('FONTNAME', (0, 0), (0, -1), template.header_font),
    ]))
    story.append(invoice_table)
    story.append(Spacer(1, 15))
    
    # Customer and Supplier Details
    details_data = []
    
    if template.show_customer_details and customer:
        details_data.append(['Bill To:', customer.name])
        details_data.append(['', f"GSTIN: {customer.gstin}" if customer.gstin else "GSTIN: Not Available"])
        customer_address = f"{customer.billing_address_line1}"
        if customer.billing_address_line2:
            customer_address += f", {customer.billing_address_line2}"
        customer_address += f", {customer.billing_city}, {customer.billing_state} - {customer.billing_pincode or ''}"
        details_data.append(['', customer_address])
        if customer.email:
            details_data.append(['', f"Email: {customer.email}"])
        if customer.contact_number:
            details_data.append(['', f"Phone: {customer.contact_number}"])
    
    if template.show_supplier_details and supplier:
        details_data.append(['Ship From:', supplier.name])
        details_data.append(['', f"GSTIN: {supplier.gstin}" if supplier.gstin else "GSTIN: Not Available"])
        supplier_address = f"{supplier.billing_address_line1}"
        if supplier.billing_address_line2:
            supplier_address += f", {supplier.billing_address_line2}"
        supplier_address += f", {supplier.billing_city}, {supplier.billing_state} - {supplier.billing_pincode or ''}"
        details_data.append(['', supplier_address])
    
    if details_data:
        details_table = Table(details_data, colWidths=[2*cm, 14*cm])
        details_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), template.body_font),
            ('FONTSIZE', (0, 0), (-1, -1), template.body_font_size - 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('FONTNAME', (0, 0), (0, -1), template.header_font),
            ('BACKGROUND', (0, 0), (0, -1), secondary_color),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ]))
        story.append(details_table)
        story.append(Spacer(1, 15))
    
    # Items Table
    story.append(Paragraph("<b>Item Details</b>", heading_style))
    
    # Table headers
    headers = ['S.No', 'Description', 'HSN', 'Qty', 'Rate', 'Amount', 'GST %', 'CGST', 'SGST', 'Total']
    table_data = [headers]
    
    # Add items
    for i, item in enumerate(items, 1):
        product = db.query(Product).filter(Product.id == item.product_id).first()
        description = product.name if product else item.description
        
        row = [
            str(i),
            description,
            item.hsn_code or '',
            str(item.qty),
            format_currency_for_pdf(float(item.rate), inv.currency),
            format_currency_for_pdf(float(item.taxable_value), inv.currency),
            f"{item.gst_rate}%",
            format_currency_for_pdf(float(item.cgst), inv.currency),
            format_currency_for_pdf(float(item.sgst), inv.currency),
            format_currency_for_pdf(float(item.amount), inv.currency)
        ]
        table_data.append(row)
    
    # Create items table
    items_table = Table(table_data, colWidths=[0.8*cm, 4*cm, 1.5*cm, 1*cm, 1.5*cm, 1.5*cm, 1*cm, 1.2*cm, 1.2*cm, 1.5*cm])
    items_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Description left-aligned
        ('FONTNAME', (0, 0), (-1, 0), template.header_font),  # Header row
        ('FONTSIZE', (0, 0), (-1, -1), template.body_font_size - 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 1, primary_color),
        ('BACKGROUND', (0, 0), (-1, 0), secondary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 15))
    
    # Totals Table
    totals_data = [
        ['Subtotal:', format_currency_for_pdf(float(inv.taxable_value), inv.currency)],
        ['CGST:', format_currency_for_pdf(float(inv.cgst), inv.currency)],
        ['SGST:', format_currency_for_pdf(float(inv.sgst), inv.currency)],
        ['IGST:', format_currency_for_pdf(float(inv.igst), inv.currency)],
        ['Total:', format_currency_for_pdf(float(inv.grand_total), inv.currency)]
    ]
    
    totals_table = Table(totals_data, colWidths=[4*cm, 2*cm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), template.body_font),
        ('FONTSIZE', (0, 0), (-1, -1), template.body_font_size),
        ('FONTNAME', (0, -1), (-1, -1), template.header_font),  # Total row bold
        ('FONTSIZE', (0, -1), (-1, -1), template.body_font_size + 2),  # Total row larger
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BACKGROUND', (0, -1), (-1, -1), accent_color),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 20))
    
    # Terms
    if template.show_terms:
        story.append(Paragraph(f"<b>Terms:</b> {template.terms_text}", normal_style))
        story.append(Spacer(1, 10))
    
    # Notes
    if template.show_notes and inv.notes:
        story.append(Paragraph("<b>Notes:</b>", heading_style))
        story.append(Paragraph(inv.notes, normal_style))
        story.append(Spacer(1, 15))
    
    # Footer
    if template.show_footer:
        story.append(Paragraph(template.footer_text, normal_style))
        story.append(Paragraph("This is a computer generated invoice", normal_style))
    
    # Build PDF
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    
    return Response(content=pdf, media_type='application/pdf')


class EmailRequest(BaseModel):
    to: str


@api.get('/invoices/{invoice_id}', response_model=InvoiceOut)
def get_invoice(invoice_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    return inv


@api.put('/invoices/{invoice_id}', response_model=InvoiceOut)
def update_invoice(invoice_id: int, payload: InvoiceCreate, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    # Validation (same as create_invoice)
    if payload.invoice_no and len(payload.invoice_no) > 16:
        raise HTTPException(status_code=400, detail="Invoice number must be 16 characters or less as per GST law")
    if payload.invoice_no and not re.match(r'^[a-zA-Z0-9\s\-/]+$', payload.invoice_no):
        raise HTTPException(status_code=400, detail="Invoice number must be alphanumeric with spaces, hyphens, and forward slashes only")
    if len(payload.bill_to_address) > 200:
        raise HTTPException(status_code=400, detail="Bill to address must be 200 characters or less")
    if len(payload.ship_to_address) > 200:
        raise HTTPException(status_code=400, detail="Ship to address must be 200 characters or less")
    if payload.notes and len(payload.notes) > 200:
        raise HTTPException(status_code=400, detail="Notes must be 200 characters or less")
    if payload.eway_bill_number and len(payload.eway_bill_number) > 50:
        raise HTTPException(status_code=400, detail="E-way bill number must be 50 characters or less")
    if payload.eway_bill_number and not re.match(r'^[0-9]+$', payload.eway_bill_number):
        raise HTTPException(status_code=400, detail="E-way bill number must contain only numbers")
    if not payload.place_of_supply:
        raise HTTPException(status_code=400, detail="Place of supply is mandatory as per GST law")
    if not payload.place_of_supply_state_code:
        raise HTTPException(status_code=400, detail="Place of supply state code is mandatory as per GST law")
    
    # Check if invoice number is being changed and if it's unique
    if payload.invoice_no and payload.invoice_no != inv.invoice_no:
        existing = db.query(Invoice).filter(Invoice.invoice_no == payload.invoice_no).first()
        if existing:
            raise HTTPException(status_code=400, detail="Invoice number already exists")
    
    # Update invoice details
    inv.invoice_no = payload.invoice_no or inv.invoice_no
    inv.date = datetime.fromisoformat(payload.date.replace('Z', '+00:00'))
    inv.due_date = calculate_due_date(payload.date, payload.terms)
    inv.terms = payload.terms
    inv.place_of_supply = payload.place_of_supply
    inv.place_of_supply_state_code = payload.place_of_supply_state_code
    inv.eway_bill_number = payload.eway_bill_number
    inv.reverse_charge = payload.reverse_charge
    inv.export_supply = payload.export_supply
    inv.bill_to_address = payload.bill_to_address
    inv.ship_to_address = payload.ship_to_address
    inv.notes = payload.notes
    
    # Delete existing items and recreate them
    db.query(InvoiceItem).filter(InvoiceItem.invoice_id == inv.id).delete()
    
    # Recalculate totals
    taxable_total = money(0)
    discount_total = money(0)
    cgst_total = money(0)
    sgst_total = money(0)
    igst_total = money(0)
    
    company = db.query(CompanySettings).first()
    customer = db.query(Party).filter(Party.id == inv.customer_id).first()
    intra = company and inv.place_of_supply == company.state
    
    for it in payload.items:
        prod = db.query(Product).filter(Product.id == it.product_id).first()
        if not prod:
            raise HTTPException(status_code=400, detail='Invalid product')
        
        # Calculate line item amounts
        line_total = money(Decimal(it.qty) * Decimal(it.rate))
        
        # Apply discount
        if it.discount > 0:
            if it.discount_type == "Percentage":
                discount_amount = money(line_total * Decimal(it.discount) / Decimal(100))
            else:  # Fixed
                discount_amount = money(Decimal(it.discount))
            line_total -= discount_amount
            discount_total += discount_amount
        
        # Calculate GST
        cgst, sgst, igst = split_gst(line_total, prod.gst_rate, bool(intra))
        
        item = InvoiceItem(
            invoice_id=inv.id,
            product_id=prod.id,
            description=prod.name,
            hsn_code=prod.hsn,
            qty=it.qty,
            rate=money(it.rate),
            discount=money(it.discount),
            discount_type=it.discount_type,
            taxable_value=line_total,
            gst_rate=prod.gst_rate,
            cgst=cgst, sgst=sgst, igst=igst,
            amount=line_total + cgst + sgst + igst
        )
        db.add(item)
        taxable_total += line_total
        cgst_total += cgst
        sgst_total += sgst
        igst_total += igst
    
    # Update invoice totals
    inv.taxable_value = taxable_total
    inv.total_discount = discount_total
    inv.cgst = cgst_total
    inv.sgst = sgst_total
    inv.igst = igst_total
    inv.grand_total = money(taxable_total + cgst_total + sgst_total + igst_total)
    
    db.commit()
    db.refresh(inv)
    return inv


@api.patch('/invoices/{invoice_id}/status', response_model=InvoiceOut)
def update_invoice_status(invoice_id: int, status: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    if status not in ["Draft", "Sent", "Paid", "Overdue"]:
        raise HTTPException(status_code=400, detail='Invalid status')
    
    inv.status = status
    db.commit()
    db.refresh(inv)
    return inv


@api.delete('/invoices/{invoice_id}')
def delete_invoice(invoice_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not inv:
            raise HTTPException(status_code=404, detail='Invoice not found')
        
        # Check if invoice has payments
        payments = db.query(Payment).filter(Payment.invoice_id == invoice_id).first()
        if payments:
            raise HTTPException(status_code=400, detail='Cannot delete invoice with existing payments. Please delete payments first.')
        
        # Delete related items first (cascade delete)
        db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_id).delete()
        
        # Delete the invoice
        db.delete(inv)
        db.commit()
        
        return {"message": "Invoice deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete invoice: {str(e)}")


# Payment Management Endpoints
@api.post('/invoices/{invoice_id}/payments', status_code=201)
def add_payment(invoice_id: int, payload: PaymentIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    # Validate payment amount
    if payload.payment_amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be greater than 0")
    
    outstanding_amount = float(inv.grand_total - inv.paid_amount)
    if payload.payment_amount > outstanding_amount:
        raise HTTPException(status_code=400, detail=f"Payment amount cannot exceed outstanding amount of ₹{outstanding_amount:.2f}")
    
    try:
        pay = Payment(
            invoice_id=invoice_id, 
            payment_amount=money(payload.payment_amount), 
            payment_method=payload.payment_method, 
            account_head=payload.account_head,
            reference_number=payload.reference_number,
            notes=payload.notes
        )
        db.add(pay)
        
        # Update invoice paid amount
        inv.paid_amount += money(payload.payment_amount)
        inv.balance_amount = inv.grand_total - inv.paid_amount
        
        # Update invoice status
        if inv.balance_amount == 0:
            inv.status = "Paid"
        elif inv.paid_amount > 0:
            inv.status = "Partially Paid"
        
        # Cashflow transaction is now handled by the source table (Payment) directly
        
        db.commit()
        return {"id": pay.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process payment: {str(e)}")


@api.get('/invoice-payments', response_model=list[PaymentOut])
def list_invoice_payments(
    search: str | None = None,
    payment_status: str | None = None,
    payment_method: str | None = None,
    customer_id: int | None = None,
    amount_min: float | None = None,
    amount_max: float | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Get all invoice payments with filtering"""
    query = db.query(Payment).join(Invoice, Payment.invoice_id == Invoice.id).join(Party, Invoice.customer_id == Party.id)
    
    if search:
        search_filter = (
            Payment.reference_number.ilike(f"%{search}%") |
            Payment.notes.ilike(f"%{search}%") |
            Party.name.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    if payment_status:
        if payment_status == 'paid':
            query = query.filter(Payment.payment_amount > 0)
        elif payment_status == 'pending':
            query = query.filter(Payment.payment_amount == 0)
    
    if payment_method:
        query = query.filter(Payment.payment_method == payment_method)
    
    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)
    
    if amount_min is not None:
        query = query.filter(Payment.payment_amount >= amount_min)
    
    if amount_max is not None:
        query = query.filter(Payment.payment_amount <= amount_max)
    
    if date_from:
        query = query.filter(Payment.payment_date >= datetime.fromisoformat(date_from))
    
    if date_to:
        query = query.filter(Payment.payment_date <= datetime.fromisoformat(date_to))
    
    payments = query.order_by(Payment.payment_date.desc()).all()
    
    result = []
    for payment in payments:
        invoice = db.query(Invoice).filter(Invoice.id == payment.invoice_id).first()
        customer = db.query(Party).filter(Party.id == invoice.customer_id).first() if invoice else None
        
        result.append(PaymentOut(
            id=payment.id,
            invoice_id=payment.invoice_id,
            payment_amount=float(payment.payment_amount),
            payment_method=payment.payment_method,
            account_head=payment.account_head,
            reference_number=payment.reference_number,
            payment_date=payment.payment_date,
            notes=payment.notes,
            created_at=payment.created_at,
            updated_at=payment.updated_at
        ))
    
    return result

@api.get('/invoices/{invoice_id}/payments', response_model=list[PaymentOut])
def get_invoice_payments(invoice_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    pays = db.query(Payment).filter(Payment.invoice_id == invoice_id).order_by(Payment.payment_date.desc()).all()
    return pays


@api.delete('/payments/{payment_id}')
def delete_payment(payment_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=400, detail='Payment not found')
    
    # Update invoice payment status
    inv = db.query(Invoice).filter(Invoice.id == payment.invoice_id).first()
    if inv:
        inv.paid_amount -= payment.payment_amount
        inv.balance_amount = inv.grand_total - inv.paid_amount
        
        # Update invoice status based on payment
        if inv.balance_amount == inv.grand_total:
            inv.status = "Sent"
        elif inv.balance_amount == 0:
            inv.status = "Paid"
        elif inv.paid_amount > 0:
            inv.status = "Partially Paid"
    
    db.delete(payment)
    db.commit()
    return {"message": "Payment deleted successfully"}


@api.post('/invoices/{invoice_id}/email', status_code=202)
def email_invoice(invoice_id: int, payload: EmailRequest, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    # Update status to Sent when emailing
    inv.status = "Sent"
    db.commit()
    
    # Get customer details
    customer = db.query(Party).filter(Party.id == inv.customer_id).first()
    customer_name = customer.name if customer else "Valued Customer"
    
    # Get company details
    company = db.query(CompanySettings).first()
    company_name = company.name if company else "ProfitPath"
    
    # Generate PDF
    try:
        # Create PDF buffer
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=darkblue,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=darkblue,
            spaceAfter=6
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=3
        )
        
        # Build PDF content (same as invoice_pdf function)
        story = []
        
        # Header - Company Details
        if company:
            story.append(Paragraph(f"<b>{company.name}</b>", title_style))
            story.append(Paragraph(f"GSTIN: {company.gstin}", normal_style))
            story.append(Paragraph(f"State: {company.state} - {company.state_code}", normal_style))
        else:
            story.append(Paragraph("<b>CASHFLOW</b>", title_style))
            story.append(Paragraph("Financial Management System", normal_style))
        
        story.append(Spacer(1, 20))
        
        # Invoice Header
        story.append(Paragraph(f"<b>TAX INVOICE</b>", heading_style))
        
        # Invoice Details Table
        invoice_data = [
            ['Invoice No:', inv.invoice_no, 'Date:', inv.date.strftime('%d/%m/%Y')],
            ['Due Date:', inv.due_date.strftime('%d/%m/%Y'), 'Terms:', inv.terms],
            ['Place of Supply:', inv.place_of_supply, 'State Code:', inv.place_of_supply_state_code]
        ]
        
        if inv.eway_bill_number:
            invoice_data.append(['E-way Bill No:', inv.eway_bill_number, '', ''])
        
        invoice_table = Table(invoice_data, colWidths=[2*cm, 6*cm, 2*cm, 6*cm])
        invoice_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(invoice_table)
        story.append(Spacer(1, 15))
        
        # Customer and Supplier Details
        details_data = []
        
        if customer:
            details_data.append(['Bill To:', customer.name])
            details_data.append(['', f"GSTIN: {customer.gstin}" if customer.gstin else "GSTIN: Not Available"])
            customer_address = f"{customer.billing_address_line1}"
            if customer.billing_address_line2:
                customer_address += f", {customer.billing_address_line2}"
            customer_address += f", {customer.billing_city}, {customer.billing_state} - {customer.billing_pincode or ''}"
            details_data.append(['', customer_address])
            if customer.email:
                details_data.append(['', f"Email: {customer.email}"])
            if customer.contact_number:
                details_data.append(['', f"Phone: {customer.contact_number}"])
        
        supplier = db.query(Party).filter(Party.id == inv.supplier_id).first()
        if supplier:
            details_data.append(['Ship From:', supplier.name])
            details_data.append(['', f"GSTIN: {supplier.gstin}" if supplier.gstin else "GSTIN: Not Available"])
            supplier_address = f"{supplier.billing_address_line1}"
            if supplier.billing_address_line2:
                supplier_address += f", {supplier.billing_address_line2}"
            supplier_address += f", {supplier.billing_city}, {supplier.billing_state} - {supplier.billing_pincode or ''}"
            details_data.append(['', supplier_address])
        
        if details_data:
            details_table = Table(details_data, colWidths=[2*cm, 14*cm])
            details_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            story.append(details_table)
            story.append(Spacer(1, 15))
        
        # Items Table
        story.append(Paragraph("<b>Item Details</b>", heading_style))
        
        # Table headers
        headers = ['S.No', 'Description', 'HSN', 'Qty', 'Rate', 'Amount', 'GST %', 'CGST', 'SGST', 'Total']
        table_data = [headers]
        
        # Add items
        items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == inv.id).all()
        for i, item in enumerate(items, 1):
            product = db.query(Product).filter(Product.id == item.product_id).first()
            description = product.name if product else item.description
            
            row = [
                str(i),
                description,
                item.hsn_code or '',
                str(item.qty),
                f"₹{float(item.rate):.2f}",
                f"₹{float(item.taxable_value):.2f}",
                f"{item.gst_rate}%",
                f"₹{float(item.cgst):.2f}",
                f"₹{float(item.sgst):.2f}",
                f"₹{float(item.amount):.2f}"
            ]
            table_data.append(row)
        
        # Create items table
        items_table = Table(table_data, colWidths=[0.8*cm, 4*cm, 1.5*cm, 1*cm, 1.5*cm, 1.5*cm, 1*cm, 1.2*cm, 1.2*cm, 1.5*cm])
        items_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Description left-aligned
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header row
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 15))
        
        # Totals Table
        totals_data = [
            ['Subtotal:', f"₹{float(inv.taxable_value):.2f}"],
            ['CGST:', f"₹{float(inv.cgst):.2f}"],
            ['SGST:', f"₹{float(inv.sgst):.2f}"],
            ['IGST:', f"₹{float(inv.igst):.2f}"],
            ['Total:', f"₹{float(inv.grand_total):.2f}"]
        ]
        
        totals_table = Table(totals_data, colWidths=[4*cm, 2*cm])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Total row bold
            ('FONTSIZE', (0, -1), (-1, -1), 12),  # Total row larger
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 20))
        
        # Notes
        if inv.notes:
            story.append(Paragraph("<b>Notes:</b>", heading_style))
            story.append(Paragraph(inv.notes, normal_style))
            story.append(Spacer(1, 15))
        
        # Footer
        story.append(Paragraph("Thank you for your business!", normal_style))
        story.append(Paragraph("This is a computer generated invoice", normal_style))
        
        # Build PDF
        doc.build(story)
        pdf_content = buf.getvalue()
        buf.close()
        
        # Create email template
        text_body, html_body = create_invoice_email_template(
            invoice_no=inv.invoice_no,
            customer_name=customer_name,
            amount=float(inv.grand_total),
            due_date=inv.due_date.strftime('%d/%m/%Y'),
            company_name=company_name
        )
        
        # Send email with PDF attachment
        subject = f"Invoice {inv.invoice_no} - {company_name}"
        filename = f"Invoice_{inv.invoice_no}.pdf"
        
        success = send_email(
            to=payload.to,
            subject=subject,
            body=html_body,  # Use HTML version
            pdf_attachment=pdf_content,
            filename=filename
        )
        
        return {"status": "sent" if success else "failed"}
        
    except Exception as e:
        return {"status": "failed", "error": str(e)}


# Audit Trail Endpoints
class AuditTrailOut(BaseModel):
    id: int
    user_id: int
    action: str
    table_name: str
    record_id: int | None
    old_values: str | None
    new_values: str | None
    ip_address: str | None
    user_agent: str | None
    created_at: datetime

    class Config:
        from_attributes = True


@api.get('/audit-trail', response_model=list[AuditTrailOut])
def get_audit_trail(
    table_name: str | None = None,
    user_id: int | None = None,
    action: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    page: int = 1,
    limit: int = 50,
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Get audit trail entries with filtering and pagination"""
    query = db.query(AuditTrail)
    
    # Apply filters
    if table_name:
        query = query.filter(AuditTrail.table_name == table_name)
    if user_id:
        query = query.filter(AuditTrail.user_id == user_id)
    if action:
        query = query.filter(AuditTrail.action == action)
    if from_date:
        query = query.filter(AuditTrail.created_at >= datetime.fromisoformat(from_date))
    if to_date:
        query = query.filter(AuditTrail.created_at <= datetime.fromisoformat(to_date))
    
    # Apply pagination
    offset = (page - 1) * limit
    total = query.count()
    entries = query.order_by(AuditTrail.created_at.desc()).offset(offset).limit(limit).all()
    
    return entries


@api.get('/audit-trail/export')
def export_audit_trail(
    table_name: str | None = None,
    user_id: int | None = None,
    action: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Export audit trail to CSV"""
    import csv
    from io import StringIO
    
    query = db.query(AuditTrail)
    
    # Apply filters
    if table_name:
        query = query.filter(AuditTrail.table_name == table_name)
    if user_id:
        query = query.filter(AuditTrail.user_id == user_id)
    if action:
        query = query.filter(AuditTrail.action == action)
    if from_date:
        query = query.filter(AuditTrail.created_at >= datetime.fromisoformat(from_date))
    if to_date:
        query = query.filter(AuditTrail.created_at <= datetime.fromisoformat(to_date))
    
    entries = query.order_by(AuditTrail.created_at.desc()).all()
    
    # Create CSV
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'User ID', 'Action', 'Table', 'Record ID', 'Old Values', 'New Values', 'IP Address', 'User Agent', 'Created At'])
    
    for entry in entries:
        writer.writerow([
            entry.id,
            entry.user_id,
            entry.action,
            entry.table_name,
            entry.record_id,
            entry.old_values,
            entry.new_values,
            entry.ip_address,
            entry.user_agent,
            entry.created_at.isoformat()
        ])
    
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=audit_trail.csv'}
    )


class InvoiceListOut(BaseModel):
    id: int
    invoice_no: str
    customer_id: int
    customer_name: str
    date: datetime
    due_date: datetime
    grand_total: float
    status: str

    class Config:
        from_attributes = True


@api.get('/invoices', response_model=dict)
def list_invoices(
    search: str | None = None,
    status: str | None = None,
    customer_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    amount_min: float | None = None,
    amount_max: float | None = None,
    gst_type: str | None = None,
    payment_status: str | None = None,
    page: int = 1,
    limit: int = 10,
    sort_field: str = 'date',
    sort_direction: str = 'desc',
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    # Validation
    if page < 1:
        page = 1
    if limit < 1 or limit > 100:
        limit = 10
    
    query = db.query(Invoice).join(Party, Invoice.customer_id == Party.id)
    
    if search:
        search_filter = (
            Invoice.invoice_no.ilike(f"%{search}%") |
            Party.name.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    if status:
        query = query.filter(Invoice.status == status)
    
    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)
    
    if date_from:
        try:
            date_from_dt = datetime.fromisoformat(date_from)
            query = query.filter(Invoice.date >= date_from_dt)
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid date_from format. Use ISO format (YYYY-MM-DD)")
    
    if date_to:
        try:
            date_to_dt = datetime.fromisoformat(date_to)
            query = query.filter(Invoice.date <= date_to_dt)
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid date_to format. Use ISO format (YYYY-MM-DD)")
    
    if amount_min is not None:
        query = query.filter(Invoice.grand_total >= amount_min)
    
    if amount_max is not None:
        query = query.filter(Invoice.grand_total <= amount_max)
    
    if gst_type:
        if gst_type == 'cgst_sgst':
            query = query.filter(Invoice.igst == 0)
        elif gst_type == 'igst':
            query = query.filter(Invoice.igst > 0)
    
    if payment_status:
        if payment_status == 'paid':
            query = query.filter(Invoice.paid_amount >= Invoice.grand_total)
        elif payment_status == 'partially_paid':
            query = query.filter(
                Invoice.paid_amount > 0,
                Invoice.paid_amount < Invoice.grand_total
            )
        elif payment_status == 'unpaid':
            query = query.filter(Invoice.paid_amount == 0)
        elif payment_status == 'overdue':
            query = query.filter(
                Invoice.due_date < func.date(func.now()),
                Invoice.paid_amount < Invoice.grand_total
            )
    
    # Get total count for pagination
    total_count = query.count()
    
    # Apply sorting
    sort_column_map = {
        'invoice_no': Invoice.invoice_no,
        'customer_name': Party.name,
        'date': Invoice.date,
        'due_date': Invoice.due_date,
        'grand_total': Invoice.grand_total,
        'status': Invoice.status
    }
    
    sort_column = sort_column_map.get(sort_field, Invoice.date)
    if sort_direction.lower() == 'asc':
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())
    
    # Apply pagination
    offset = (page - 1) * limit
    invoices = query.offset(offset).limit(limit).all()
    
    # Convert to response format with customer name
    result = []
    for inv in invoices:
        customer = db.query(Party).filter(Party.id == inv.customer_id).first()
        result.append(InvoiceListOut(
            id=inv.id,
            invoice_no=inv.invoice_no,
            customer_id=inv.customer_id,
            customer_name=customer.name if customer else "Unknown",
            date=inv.date,
            due_date=inv.due_date,
            grand_total=float(inv.grand_total),
            status=inv.status
        ))
    
    # Calculate pagination info
    total_pages = (total_count + limit - 1) // limit
    has_next = page < total_pages
    has_prev = page > 1
    
    return {
        "invoices": result,
        "pagination": {
            "page": page,
            "limit": limit,
            "total_count": total_count,
            "total_pages": total_pages,
            "has_next": has_next,
            "has_prev": has_prev
        }
    }


@api.get('/reports/gst-summary')
def gst_summary(from_: str = Query(alias='from'), to: str = Query(alias='to'), _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # naive impl: aggregate all invoices between dates
    from sqlalchemy import func, cast, Date
    from datetime import datetime
    
    # Convert string dates to proper date objects
    from_date = datetime.strptime(from_, '%Y-%m-%d').date()
    to_date = datetime.strptime(to, '%Y-%m-%d').date()
    
    q = db.query(Invoice).filter(
        cast(Invoice.date, Date) >= from_date,
        cast(Invoice.date, Date) <= to_date
    )
    invoices = q.all()
    taxable = sum([float(i.taxable_value) for i in invoices], 0.0)
    cgst = sum([float(i.cgst) for i in invoices], 0.0)
    sgst = sum([float(i.sgst) for i in invoices], 0.0)
    igst = sum([float(i.igst) for i in invoices], 0.0)
    grand = sum([float(i.grand_total) for i in invoices], 0.0)
    # rate-wise breakup using items
    from sqlalchemy import select
    from .models import InvoiceItem, Product
    rows = db.execute(
        select(Product.gst_rate, func.sum(InvoiceItem.taxable_value))
        .join(Product, Product.id == InvoiceItem.product_id)
        .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        .filter(
            cast(Invoice.date, Date) >= from_date,
            cast(Invoice.date, Date) <= to_date
        )
        .group_by(Product.gst_rate)
    ).all()
    rate_breakup = [{"rate": float(r[0]), "taxable_value": float(r[1])} for r in rows]
    return {
        "taxable_value": taxable,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "grand_total": grand,
        "rate_breakup": rate_breakup,
    }


@api.get('/reports/gst-summary.csv')
def gst_summary_csv(from_: str = Query(alias='from'), to: str = Query(alias='to'), _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    data = gst_summary(from_, to, _, db)
    import csv
    from io import StringIO
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(["taxable_value", "cgst", "sgst", "igst", "grand_total"])
    w.writerow([data['taxable_value'], data['cgst'], data['sgst'], data['igst'], data['grand_total']])
    w.writerow([])
    w.writerow(["rate", "taxable_value"]) 
    for row in data['rate_breakup']:
        w.writerow([row['rate'], row['taxable_value']])
    return Response(content=buf.getvalue(), media_type='text/csv')


@api.get('/reports/gst-filing')
def gst_filing_report(
    period_type: str = Query(..., description="month/quarter/year"),
    period_value: str = Query(..., description="YYYY-MM for month, YYYY-Q1/Q2/Q3/Q4 for quarter, YYYY for year"),
    report_type: str = Query(..., description="gstr1/gstr2/gstr3b"),
    format: str = Query("json", description="json/csv/excel"),
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Generate GST filing reports compliant with Indian GST portal requirements"""
    
    # Calculate date range based on period
    start_date, end_date = _calculate_period_dates(period_type, period_value)
    
    if report_type == "gstr1":
        return _generate_gstr1_report(start_date, end_date, format, db)
    elif report_type == "gstr2":
        return _generate_gstr2_report(start_date, end_date, format, db)
    elif report_type == "gstr3b":
        return _generate_gstr3b_report(start_date, end_date, format, db)
    else:
        raise HTTPException(status_code=400, detail="Invalid report type. Use gstr1, gstr2, or gstr3b")


@api.get('/reports/gstr1')
def generate_gstr1_report_api(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    format: str = Query("json", description="json/csv"),
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Generate GSTR-1 report in GST portal format"""
    result = generate_gstr1_report(db, start_date, end_date)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    
    if format == "csv":
        return Response(
            content=result["csv_content"],
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=gstr1_{start_date}_to_{end_date}.csv"}
        )
    
    return result


@api.get('/reports/gstr3b')
def generate_gstr3b_report_api(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    format: str = Query("json", description="json/csv"),
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Generate GSTR-3B report in GST portal format"""
    result = generate_gstr3b_report(db, start_date, end_date)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result["message"])
    
    if format == "csv":
        return Response(
            content=result["csv_content"],
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=gstr3b_{start_date}_to_{end_date}.csv"}
        )
    
    return result


@api.get('/reports/gst-validation')
def validate_gst_data(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Validate GST data for report generation"""
    from .gst_reports import GSTReportGenerator
    from datetime import datetime
    
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        generator = GSTReportGenerator(db)
        
        gstr1_errors = generator.validate_data_for_gstr1(start_dt, end_dt)
        gstr3b_errors = generator.validate_data_for_gstr3b(start_dt, end_dt)
        
        return {
            "success": True,
            "gstr1_errors": gstr1_errors,
            "gstr3b_errors": gstr3b_errors,
            "gstr1_valid": len(gstr1_errors) == 0,
            "gstr3b_valid": len(gstr3b_errors) == 0,
            "period": f"{start_date} to {end_date}"
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": "Failed to validate GST data"
        }


# Inventory Reports Pydantic Models
class InventorySummaryItem(BaseModel):
    product_id: int
    product_name: str
    sku: str | None
    category: str | None
    unit: str | None
    current_stock: float
    purchase_price: float | None
    sales_price: float | None
    stock_value: float
    last_movement_date: str | None
    minimum_stock: float | None

class InventorySummaryReport(BaseModel):
    total_products: int
    total_stock_value: float
    low_stock_items: int
    out_of_stock_items: int
    items: list[InventorySummaryItem]
    generated_at: str
    filters_applied: dict | None

# Stock Ledger Report Models
class StockLedgerItem(BaseModel):
    transaction_id: int
    transaction_date: str
    product_id: int
    product_name: str
    sku: str | None
    entry_type: str  # 'in', 'out', 'adjust'
    quantity: float
    unit_price: float | None
    total_value: float | None
    running_balance: float
    reference_type: str | None  # 'invoice', 'purchase', 'adjustment'
    reference_id: int | None
    reference_number: str | None
    notes: str | None

class StockLedgerReport(BaseModel):
    total_transactions: int
    total_incoming: float
    total_outgoing: float
    total_adjustments: float
    opening_balance: float
    closing_balance: float
    transactions: list[StockLedgerItem]
    generated_at: str
    filters_applied: dict | None

# Inventory Valuation Report Models
class InventoryValuationItem(BaseModel):
    product_id: int
    product_name: str
    sku: str | None
    category: str | None
    current_stock: float
    unit_cost: float | None
    total_cost_value: float
    unit_market_price: float | None
    total_market_value: float
    valuation_difference: float
    last_updated: str | None

class InventoryValuationReport(BaseModel):
    total_products: int
    total_cost_value: float
    total_market_value: float
    total_valuation_difference: float
    items: list[InventoryValuationItem]
    generated_at: str
    filters_applied: dict | None

# Inventory Dashboard Models
class InventoryDashboardMetrics(BaseModel):
    total_products: int
    total_stock_value: float
    low_stock_items: int
    out_of_stock_items: int
    recent_movements: int
    average_stock_level: float
    top_moving_products: list[dict]
    low_stock_alerts: list[dict]
    generated_at: str = None


# Financial Reporting Models
class CashflowReportItem(BaseModel):
    transaction_id: int
    transaction_type: str  # 'income', 'expense', 'purchase_payment'
    transaction_date: str
    amount: float
    description: str
    category: str | None
    payment_method: str | None
    reference_type: str | None  # 'invoice', 'purchase', 'expense'
    reference_id: int | None
    reference_number: str | None
    party_name: str | None

class CashflowReport(BaseModel):
    period: dict
    summary: dict
    transactions: list[CashflowReportItem]
    trends: dict
    generated_at: str
    filters_applied: dict | None

class IncomeReportItem(BaseModel):
    invoice_id: int
    invoice_no: str
    invoice_date: str
    customer_name: str
    taxable_value: float
    total_tax: float
    grand_total: float
    payment_status: str
    payment_amount: float
    outstanding_amount: float

class IncomeReport(BaseModel):
    period: dict
    summary: dict
    transactions: list[IncomeReportItem]
    customer_breakdown: list[dict]
    product_breakdown: list[dict]
    trends: dict
    generated_at: str
    filters_applied: dict | None

class ExpenseReportItem(BaseModel):
    expense_id: int
    expense_date: str
    category: str
    description: str
    amount: float
    vendor_name: str | None
    payment_method: str | None
    reference_number: str | None

class ExpenseReport(BaseModel):
    period: dict
    summary: dict
    transactions: list[ExpenseReportItem]
    category_breakdown: list[dict]
    vendor_breakdown: list[dict]
    trends: dict
    generated_at: str
    filters_applied: dict | None

class PurchaseReportItem(BaseModel):
    purchase_id: int
    purchase_no: str
    purchase_date: str
    vendor_name: str
    taxable_value: float
    total_tax: float
    grand_total: float
    payment_status: str
    payment_amount: float
    outstanding_amount: float

class PurchaseReport(BaseModel):
    period: dict
    summary: dict
    transactions: list[PurchaseReportItem]
    vendor_breakdown: list[dict]
    product_breakdown: list[dict]
    trends: dict
    generated_at: str
    filters_applied: dict | None

class PaymentReportItem(BaseModel):
    payment_id: int
    payment_date: str
    payment_type: str  # 'invoice_payment', 'purchase_payment'
    payment_method: str
    amount: float
    reference_type: str
    reference_id: int
    reference_number: str
    party_name: str
    status: str

class PaymentReport(BaseModel):
    period: dict
    summary: dict
    transactions: list[PaymentReportItem]
    method_breakdown: list[dict]
    party_breakdown: list[dict]
    trends: dict
    generated_at: str
    filters_applied: dict | None

class FinancialReport(BaseModel):
    report_type: str
    period: dict
    data: dict
    generated_at: str
    filters_applied: dict | None


@api.get('/reports/inventory-summary', response_model=InventorySummaryReport)
def get_inventory_summary_report(
    category: str | None = Query(None, description="Filter by product category"),
    low_stock_only: bool = Query(False, description="Show only low stock items"),
    out_of_stock_only: bool = Query(False, description="Show only out of stock items"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive inventory summary report"""
    
    # Build base query
    query = db.query(Product).filter(Product.is_active == True)
    
    # Apply filters
    if category:
        query = query.filter(Product.category == category)
    
    products = query.all()
    
    # Calculate current stock levels and values
    summary_items = []
    total_stock_value = 0
    low_stock_count = 0
    out_of_stock_count = 0
    
    for product in products:
        # Calculate current stock from ledger entries
        stock_entries = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id
        ).all()
        
        current_stock = sum(
            entry.qty for entry in stock_entries 
            if entry.entry_type in ['in', 'adjust']
        ) - sum(
            entry.qty for entry in stock_entries 
            if entry.entry_type == 'out'
        )
        current_stock = max(0, current_stock)
        
        # Calculate stock value
        unit_price = product.purchase_price or product.sales_price or 0
        stock_value = current_stock * float(unit_price) if unit_price else 0
        
        # Get last movement date
        last_movement = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id
        ).order_by(StockLedgerEntry.created_at.desc()).first()
        
        last_movement_date = last_movement.created_at.isoformat() if last_movement else None
        
        # Check stock status (using default minimum stock of 10 since field doesn't exist)
        default_minimum_stock = 10
        is_low_stock = current_stock <= default_minimum_stock
        is_out_of_stock = current_stock == 0
        
        if is_low_stock:
            low_stock_count += 1
        if is_out_of_stock:
            out_of_stock_count += 1
        
        # Apply additional filters
        if low_stock_only and not is_low_stock:
            continue
        if out_of_stock_only and not is_out_of_stock:
            continue
        
        summary_items.append(InventorySummaryItem(
            product_id=product.id,
            product_name=product.name,
            sku=product.sku,
            category=product.category,
            unit=product.unit,
            current_stock=current_stock,
            purchase_price=float(product.purchase_price) if product.purchase_price else None,
            sales_price=float(product.sales_price) if product.sales_price else None,
            stock_value=stock_value,
            last_movement_date=last_movement_date,
            minimum_stock=default_minimum_stock
        ))
        
        total_stock_value += stock_value
    
    # Build filters applied
    filters_applied = {}
    if category:
        filters_applied["category"] = category
    if low_stock_only:
        filters_applied["low_stock_only"] = True
    if out_of_stock_only:
        filters_applied["out_of_stock_only"] = True
    
    return InventorySummaryReport(
        total_products=len(summary_items),
        total_stock_value=total_stock_value,
        low_stock_items=low_stock_count,
        out_of_stock_items=out_of_stock_count,
        items=summary_items,
        generated_at=datetime.now().isoformat(),
        filters_applied=filters_applied if filters_applied else None
    )


@api.get('/reports/stock-ledger', response_model=StockLedgerReport)
def get_stock_ledger_report(
    product_id: int | None = Query(None, description="Filter by product ID"),
    from_date: str | None = Query(None, description="Filter from date (YYYY-MM-DD)"),
    to_date: str | None = Query(None, description="Filter to date (YYYY-MM-DD)"),
    entry_type: str | None = Query(None, description="Filter by entry type (in/out/adjust)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate detailed stock ledger report with running balances"""
    try:
        # Build query for stock ledger entries
        query = db.query(StockLedgerEntry).join(Product)
        
        # Apply filters
        if product_id:
            query = query.filter(StockLedgerEntry.product_id == product_id)
        
        if from_date:
            query = query.filter(StockLedgerEntry.created_at >= from_date)
        
        if to_date:
            query = query.filter(StockLedgerEntry.created_at <= to_date + " 23:59:59")
        
        if entry_type:
            query = query.filter(StockLedgerEntry.entry_type == entry_type)
        
        # Order by date and product
        entries = query.order_by(StockLedgerEntry.created_at, StockLedgerEntry.product_id).all()
        
        # Group entries by product and calculate running balances
        ledger_items = []
        total_incoming = 0
        total_outgoing = 0
        total_adjustments = 0
        
        # Group by product for running balance calculation
        product_entries = {}
        for entry in entries:
            if entry.product_id not in product_entries:
                product_entries[entry.product_id] = []
            product_entries[entry.product_id].append(entry)
        
        # Calculate running balances for each product
        for product_id, product_entry_list in product_entries.items():
            running_balance = 0
            
            for entry in product_entry_list:
                # Calculate running balance
                if entry.entry_type == 'in':
                    running_balance += entry.qty
                    total_incoming += entry.qty
                elif entry.entry_type == 'out':
                    running_balance -= entry.qty
                    total_outgoing += entry.qty
                elif entry.entry_type == 'adjust':
                    running_balance += entry.qty
                    total_adjustments += abs(entry.qty)
                
                # Get product details
                product = db.query(Product).filter(Product.id == entry.product_id).first()
                
                # Determine reference information
                reference_type = entry.ref_type if entry.ref_type else 'adjustment'
                reference_id = entry.ref_id if entry.ref_id else entry.id
                reference_number = None
                
                if entry.ref_type == 'invoice' and entry.ref_id:
                    invoice = db.query(Invoice).filter(Invoice.id == entry.ref_id).first()
                    reference_number = invoice.invoice_no if invoice else None
                elif entry.ref_type == 'purchase' and entry.ref_id:
                    purchase = db.query(Purchase).filter(Purchase.id == entry.ref_id).first()
                    reference_number = purchase.purchase_no if purchase else None
                else:
                    reference_number = f"ADJ-{entry.id}"
                
                # Get unit price from product if available
                unit_price = None
                if product and product.purchase_price:
                    unit_price = float(product.purchase_price)
                
                ledger_items.append(StockLedgerItem(
                    transaction_id=entry.id,
                    transaction_date=entry.created_at.isoformat(),
                    product_id=entry.product_id,
                    product_name=product.name if product else "Unknown Product",
                    sku=product.sku if product else None,
                    entry_type=entry.entry_type,
                    quantity=entry.qty,
                    unit_price=unit_price,
                    total_value=entry.qty * (unit_price or 0),
                    running_balance=running_balance,
                    reference_type=reference_type,
                    reference_id=reference_id,
                    reference_number=reference_number,
                    notes=None  # StockLedgerEntry doesn't have notes field
                ))
        
        # Calculate opening and closing balances
        opening_balance = 0
        closing_balance = running_balance if ledger_items else 0
        
        return StockLedgerReport(
            total_transactions=len(ledger_items),
            total_incoming=total_incoming,
            total_outgoing=total_outgoing,
            total_adjustments=total_adjustments,
            opening_balance=opening_balance,
            closing_balance=closing_balance,
            transactions=ledger_items,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "product_id": product_id,
                "from_date": from_date,
                "to_date": to_date,
                "entry_type": entry_type
            } if any([product_id, from_date, to_date, entry_type]) else None
        )
        
    except Exception as e:
        logger.error(f"Error generating stock ledger report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate stock ledger report")


@api.get('/reports/inventory-valuation', response_model=InventoryValuationReport)
def get_inventory_valuation_report(
    category: str | None = Query(None, description="Filter by product category"),
    include_zero_stock: bool = Query(True, description="Include products with zero stock"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate inventory valuation report with cost and market value calculations"""
    try:
        # Build base query
        query = db.query(Product).filter(Product.is_active == True)
        
        # Apply filters
        if category:
            query = query.filter(Product.category == category)
        
        products = query.all()
        
        # Calculate valuations
        valuation_items = []
        total_cost_value = 0
        total_market_value = 0
        total_valuation_difference = 0
        
        for product in products:
            # Calculate current stock from ledger entries
            stock_entries = db.query(StockLedgerEntry).filter(
                StockLedgerEntry.product_id == product.id
            ).all()
            
            current_stock = sum(
                entry.qty for entry in stock_entries 
                if entry.entry_type in ['in', 'adjust']
            ) - sum(
                entry.qty for entry in stock_entries 
                if entry.entry_type == 'out'
            )
            current_stock = max(0, current_stock)
            
            # Skip zero stock items if not included
            if not include_zero_stock and current_stock == 0:
                continue
            
            # Calculate unit costs and values
            unit_cost = float(product.purchase_price) if product.purchase_price else 0
            unit_market_price = float(product.sales_price) if product.sales_price else unit_cost
            
            total_cost_value_product = current_stock * unit_cost
            total_market_value_product = current_stock * unit_market_price
            valuation_difference = total_market_value_product - total_cost_value_product
            
            # Get last updated date
            last_movement = db.query(StockLedgerEntry).filter(
                StockLedgerEntry.product_id == product.id
            ).order_by(StockLedgerEntry.created_at.desc()).first()
            
            last_updated = last_movement.created_at.isoformat() if last_movement else None
            
            valuation_items.append(InventoryValuationItem(
                product_id=product.id,
                product_name=product.name,
                sku=product.sku,
                category=product.category,
                current_stock=current_stock,
                unit_cost=unit_cost,
                total_cost_value=total_cost_value_product,
                unit_market_price=unit_market_price,
                total_market_value=total_market_value_product,
                valuation_difference=valuation_difference,
                last_updated=last_updated
            ))
            
            total_cost_value += total_cost_value_product
            total_market_value += total_market_value_product
            total_valuation_difference += valuation_difference
        
        # Build filters applied
        filters_applied = {}
        if category:
            filters_applied["category"] = category
        if not include_zero_stock:
            filters_applied["include_zero_stock"] = False
        
        return InventoryValuationReport(
            total_products=len(valuation_items),
            total_cost_value=total_cost_value,
            total_market_value=total_market_value,
            total_valuation_difference=total_valuation_difference,
            items=valuation_items,
            generated_at=datetime.now().isoformat(),
            filters_applied=filters_applied if filters_applied else None
        )
        
    except Exception as e:
        logger.error(f"Error generating inventory valuation report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate inventory valuation report")


@api.get('/reports/inventory-dashboard', response_model=InventoryDashboardMetrics)
def get_inventory_dashboard_metrics(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate real-time inventory dashboard metrics"""
    try:
        # Get all active products
        products = db.query(Product).filter(Product.is_active == True).all()
        
        total_products = len(products)
        total_stock_value = 0
        low_stock_items = 0
        out_of_stock_items = 0
        recent_movements = 0
        total_stock_quantity = 0
        
        # Calculate metrics for each product
        low_stock_alerts = []
        top_moving_products = []
        
        # Get recent movements (last 7 days)
        from datetime import timedelta
        recent_date = datetime.now() - timedelta(days=7)
        recent_entries = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.created_at >= recent_date
        ).all()
        recent_movements = len(recent_entries)
        
        for product in products:
            # Calculate current stock
            stock_entries = db.query(StockLedgerEntry).filter(
                StockLedgerEntry.product_id == product.id
            ).all()
            
            current_stock = sum(
                entry.qty for entry in stock_entries 
                if entry.entry_type in ['in', 'adjust']
            ) - sum(
                entry.qty for entry in stock_entries 
                if entry.entry_type == 'out'
            )
            current_stock = max(0, current_stock)
            
            total_stock_quantity += current_stock
            
            # Calculate stock value
            unit_price = product.purchase_price or product.sales_price or 0
            stock_value = current_stock * float(unit_price) if unit_price else 0
            total_stock_value += stock_value
            
            # Check stock status
            default_minimum_stock = 10
            if current_stock == 0:
                out_of_stock_items += 1
            elif current_stock <= default_minimum_stock:
                low_stock_items += 1
                low_stock_alerts.append({
                    "product_id": product.id,
                    "product_name": product.name,
                    "current_stock": current_stock,
                    "minimum_stock": default_minimum_stock,
                    "category": product.category
                })
            
            # Calculate movement frequency for top moving products
            movement_count = len(stock_entries)
            if movement_count > 0:
                top_moving_products.append({
                    "product_id": product.id,
                    "product_name": product.name,
                    "movement_count": movement_count,
                    "current_stock": current_stock,
                    "category": product.category
                })
        
        # Sort top moving products by movement count
        top_moving_products.sort(key=lambda x: x["movement_count"], reverse=True)
        top_moving_products = top_moving_products[:10]  # Top 10
        
        # Calculate average stock level
        average_stock_level = total_stock_quantity / total_products if total_products > 0 else 0
        
        return InventoryDashboardMetrics(
            total_products=total_products,
            total_stock_value=total_stock_value,
            low_stock_items=low_stock_items,
            out_of_stock_items=out_of_stock_items,
            recent_movements=recent_movements,
            average_stock_level=average_stock_level,
            top_moving_products=top_moving_products,
            low_stock_alerts=low_stock_alerts,
            generated_at=datetime.now().isoformat()
        )
        
    except Exception as e:
        logger.error(f"Error generating inventory dashboard metrics: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate inventory dashboard metrics")


# Cashflow Reports API Endpoints
@api.get('/reports/cashflow', response_model=CashflowReport)
def get_cashflow_report(
    start_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    transaction_type: str | None = Query(None, description="Filter by transaction type: income, expense, purchase_payment"),
    payment_method: str | None = Query(None, description="Filter by payment method"),
    category: str | None = Query(None, description="Filter by category"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive cashflow report with trends and analysis"""
    
    try:
        from datetime import datetime, timedelta
        
        # Set default date range if not provided
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Convert to datetime objects
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Initialize cashflow service
        from .profitpath_service import ProfitPathService
        profitpath_service = ProfitPathService(db)
        
        # Get cashflow summary
        summary = profitpath_service.get_profitpath_summary(start_dt, end_dt)
        
        # Get detailed transactions
        transactions = []
        
        # Income transactions (Invoice Payments)
        income_filter = []
        if start_date:
            income_filter.append(Payment.payment_date >= start_dt)
        if end_date:
            income_filter.append(Payment.payment_date <= end_dt)
        if payment_method:
            income_filter.append(Payment.payment_method == payment_method)
        
        income_payments = db.query(Payment).filter(and_(*income_filter)).all()
        for payment in income_payments:
            invoice = db.query(Invoice).filter(Invoice.id == payment.invoice_id).first()
            customer = db.query(Party).filter(Party.id == invoice.customer_id).first() if invoice else None
            
            transactions.append(CashflowReportItem(
                transaction_id=payment.id,
                transaction_type='income',
                transaction_date=payment.payment_date.isoformat(),
                amount=float(payment.payment_amount),
                description=f"Payment for Invoice {invoice.invoice_no}" if invoice else "Payment",
                category="Sales",
                payment_method=payment.payment_method,
                reference_type='invoice',
                reference_id=payment.invoice_id,
                reference_number=invoice.invoice_no if invoice else None,
                party_name=customer.name if customer else None
            ))
        
        # Expense transactions (Purchase Payments + Direct Expenses)
        purchase_payment_filter = []
        expense_filter = []
        if start_date:
            purchase_payment_filter.append(PurchasePayment.payment_date >= start_dt)
            expense_filter.append(Expense.expense_date >= start_dt)
        if end_date:
            purchase_payment_filter.append(PurchasePayment.payment_date <= end_dt)
            expense_filter.append(Expense.expense_date <= end_dt)
        if payment_method:
            purchase_payment_filter.append(PurchasePayment.payment_method == payment_method)
        
        # Purchase payments
        purchase_payments = db.query(PurchasePayment).filter(and_(*purchase_payment_filter)).all()
        for payment in purchase_payments:
            purchase = db.query(Purchase).filter(Purchase.id == payment.purchase_id).first()
            vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first() if purchase else None
            
            transactions.append(CashflowReportItem(
                transaction_id=payment.id,
                transaction_type='purchase_payment',
                transaction_date=payment.payment_date.isoformat(),
                amount=float(payment.payment_amount),
                description=f"Payment for Purchase {purchase.purchase_no}" if purchase else "Purchase Payment",
                category="Purchases",
                payment_method=payment.payment_method,
                reference_type='purchase',
                reference_id=payment.purchase_id,
                reference_number=purchase.purchase_no if purchase else None,
                party_name=vendor.name if vendor else None
            ))
        
        # Direct expenses
        expenses = db.query(Expense).filter(and_(*expense_filter)).all()
        for expense in expenses:
            # Get vendor name through relationship
            vendor = db.query(Party).filter(Party.id == expense.vendor_id).first() if expense.vendor_id else None
            
            transactions.append(CashflowReportItem(
                transaction_id=expense.id,
                transaction_type='expense',
                transaction_date=expense.expense_date.isoformat(),
                amount=float(expense.total_amount),
                description=expense.description,
                category=expense.category,
                payment_method=expense.payment_method,
                reference_type='expense',
                reference_id=expense.id,
                reference_number=expense.reference_number,
                party_name=vendor.name if vendor else None
            ))
        
        # Apply transaction type filter if specified
        if transaction_type:
            transactions = [t for t in transactions if t.transaction_type == transaction_type]
        
        # Apply category filter if specified
        if category:
            transactions = [t for t in transactions if t.category == category]
        
        # Sort transactions by date
        transactions.sort(key=lambda x: x.transaction_date, reverse=True)
        
        # Calculate trends (monthly breakdown)
        trends = {
            "monthly_breakdown": [],
            "cashflow_trend": "positive" if summary["net_cashflow"] > 0 else "negative"
        }
        
        # Generate monthly breakdown for the last 6 months
        for i in range(6):
            month_start = (datetime.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_summary = profitpath_service.get_profitpath_summary(month_start.date(), month_end.date())
            trends["monthly_breakdown"].append({
                "month": month_start.strftime('%Y-%m'),
                "income": month_summary["total_income"],
                "expenses": month_summary["total_outflow"],
                "net_cashflow": month_summary["net_cashflow"]
            })
        
        trends["monthly_breakdown"].reverse()
        
        return CashflowReport(
            period={
                "start_date": start_date,
                "end_date": end_date
            },
            summary=summary,
            transactions=transactions,
            trends=trends,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "transaction_type": transaction_type,
                "payment_method": payment_method,
                "category": category
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating cashflow report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate cashflow report")


# Income Reports API Endpoints
@api.get('/reports/income', response_model=IncomeReport)
def get_income_report(
    start_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    customer_id: int | None = Query(None, description="Filter by customer ID"),
    payment_status: str | None = Query(None, description="Filter by payment status"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive income report with customer and product breakdown"""
    
    try:
        from datetime import datetime, timedelta
        
        # Set default date range if not provided
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Validate date format
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
        
        # Validate date range
        if start_dt > end_dt:
            raise HTTPException(status_code=400, detail="Start date cannot be after end date")
        
        # Build base query
        query = db.query(Invoice).join(Party, Invoice.customer_id == Party.id)
        
        # Apply filters
        query = query.filter(Invoice.date >= start_dt, Invoice.date <= end_dt)
        if customer_id:
            query = query.filter(Invoice.customer_id == customer_id)
        if payment_status:
            # Map payment_status to status field
            status_mapping = {
                'paid': 'Paid',
                'unpaid': 'Sent',
                'partially_paid': 'Partially Paid',
                'overdue': 'Overdue'
            }
            mapped_status = status_mapping.get(payment_status.lower(), payment_status)
            query = query.filter(Invoice.status == mapped_status)
        
        invoices = query.all()
        
        # Calculate summary
        total_revenue = sum(float(inv.grand_total) for inv in invoices)
        total_tax = sum(float(inv.cgst + inv.sgst + inv.igst) for inv in invoices)
        total_taxable_value = sum(float(inv.taxable_value) for inv in invoices)
        
        # Calculate payment status breakdown
        paid_amount = 0
        outstanding_amount = 0
        for inv in invoices:
            if inv.status == 'Paid':
                paid_amount += float(inv.grand_total)
            else:
                outstanding_amount += float(inv.grand_total)
        
        summary = {
            "total_revenue": total_revenue,
            "total_tax": total_tax,
            "total_taxable_value": total_taxable_value,
            "paid_amount": paid_amount,
            "outstanding_amount": outstanding_amount,
            "invoice_count": len(invoices),
            "payment_status_breakdown": {
                "paid": paid_amount,
                "outstanding": outstanding_amount
            }
        }
        
        # Build transaction list
        transactions = []
        for inv in invoices:
            customer = db.query(Party).filter(Party.id == inv.customer_id).first()
            
            # Calculate payment amount
            payments = db.query(Payment).filter(Payment.invoice_id == inv.id).all()
            payment_amount = sum(float(p.payment_amount) for p in payments)
            outstanding = float(inv.grand_total) - payment_amount
            
            transactions.append(IncomeReportItem(
                invoice_id=inv.id,
                invoice_no=inv.invoice_no,
                invoice_date=inv.date.isoformat(),
                customer_name=customer.name if customer else "Unknown",
                taxable_value=float(inv.taxable_value),
                total_tax=float(inv.cgst + inv.sgst + inv.igst),
                grand_total=float(inv.grand_total),
                payment_status=inv.status,
                payment_amount=payment_amount,
                outstanding_amount=outstanding
            ))
        
        # Customer breakdown
        customer_breakdown = []
        customer_totals = {}
        for inv in invoices:
            customer = db.query(Party).filter(Party.id == inv.customer_id).first()
            customer_name = customer.name if customer else "Unknown"
            
            if customer_name not in customer_totals:
                customer_totals[customer_name] = {
                    "total_revenue": 0,
                    "invoice_count": 0,
                    "paid_amount": 0,
                    "outstanding_amount": 0
                }
            
            customer_totals[customer_name]["total_revenue"] += float(inv.grand_total)
            customer_totals[customer_name]["invoice_count"] += 1
            
            if inv.status == 'Paid':
                customer_totals[customer_name]["paid_amount"] += float(inv.grand_total)
            else:
                customer_totals[customer_name]["outstanding_amount"] += float(inv.grand_total)
        
        for customer_name, totals in customer_totals.items():
            customer_breakdown.append({
                "customer_name": customer_name,
                "total_revenue": totals["total_revenue"],
                "invoice_count": totals["invoice_count"],
                "paid_amount": totals["paid_amount"],
                "outstanding_amount": totals["outstanding_amount"]
            })
        
        # Product breakdown
        product_breakdown = []
        product_totals = {}
        
        for inv in invoices:
            items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == inv.id).all()
            for item in items:
                product = db.query(Product).filter(Product.id == item.product_id).first()
                product_name = product.name if product else "Unknown"
                
                if product_name not in product_totals:
                    product_totals[product_name] = 0
                product_totals[product_name] += float(item.amount)
        
        for product_name, total in product_totals.items():
            product_breakdown.append({
                "product_name": product_name,
                "total_revenue": total,
                "quantity_sold": sum(
                    float(item.qty) for inv in invoices 
                    for item in db.query(InvoiceItem).filter(InvoiceItem.invoice_id == inv.id).all()
                    if db.query(Product).filter(Product.id == item.product_id).first().name == product_name
                )
            })
        
        # Calculate trends
        trends = {
            "monthly_revenue": [],
            "growth_rate": 0
        }
        
        # Generate monthly revenue for the last 6 months
        for i in range(6):
            month_start = (datetime.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_invoices = db.query(Invoice).filter(
                Invoice.date >= month_start.date(),
                Invoice.date <= month_end.date()
            ).all()
            
            month_revenue = sum(float(inv.grand_total) for inv in month_invoices)
            trends["monthly_revenue"].append({
                "month": month_start.strftime('%Y-%m'),
                "revenue": month_revenue
            })
        
        trends["monthly_revenue"].reverse()
        
        # Calculate growth rate (comparing current month to previous month)
        if len(trends["monthly_revenue"]) >= 2:
            current_month = trends["monthly_revenue"][-1]["revenue"]
            previous_month = trends["monthly_revenue"][-2]["revenue"]
            if previous_month > 0:
                trends["growth_rate"] = ((current_month - previous_month) / previous_month) * 100
        
        return IncomeReport(
            period={
                "start_date": start_date,
                "end_date": end_date
            },
            summary=summary,
            transactions=transactions,
            customer_breakdown=customer_breakdown,
            product_breakdown=product_breakdown,
            trends=trends,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "customer_id": customer_id,
                "payment_status": payment_status
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating income report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate income report")


# Expense Reports API Endpoints
@api.get('/reports/expenses', response_model=ExpenseReport)
def get_expense_report(
    start_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    category: str | None = Query(None, description="Filter by expense category"),
    vendor_name: str | None = Query(None, description="Filter by vendor name"),
    payment_method: str | None = Query(None, description="Filter by payment method"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive expense report with category and vendor breakdown"""
    
    try:
        from datetime import datetime, timedelta
        
        # Set default date range if not provided
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Convert to datetime objects
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Build base query
        query = db.query(Expense)
        
        # Apply filters
        query = query.filter(Expense.expense_date >= start_dt, Expense.expense_date <= end_dt)
        if category:
            query = query.filter(Expense.category == category)
        if vendor_name:
            # Find vendor by name and filter by vendor_id
            vendor = db.query(Party).filter(Party.name.ilike(f"%{vendor_name}%")).first()
            if vendor:
                query = query.filter(Expense.vendor_id == vendor.id)
        if payment_method:
            query = query.filter(Expense.payment_method == payment_method)
        
        expenses = query.all()
        
        # Calculate summary
        total_expenses = sum(float(exp.total_amount) for exp in expenses)
        expense_count = len(expenses)
        
        summary = {
            "total_expenses": total_expenses,
            "expense_count": expense_count,
            "average_expense": total_expenses / expense_count if expense_count > 0 else 0
        }
        
        # Build transaction list
        transactions = []
        for exp in expenses:
            # Get vendor name through relationship
            vendor = db.query(Party).filter(Party.id == exp.vendor_id).first() if exp.vendor_id else None
            
            transactions.append(ExpenseReportItem(
                expense_id=exp.id,
                expense_date=exp.expense_date.isoformat(),
                category=exp.category,
                description=exp.description,
                amount=float(exp.total_amount),
                vendor_name=vendor.name if vendor else None,
                payment_method=exp.payment_method,
                reference_number=exp.reference_number
            ))
        
        # Category breakdown
        category_breakdown = []
        category_totals = {}
        for exp in expenses:
            if exp.category not in category_totals:
                category_totals[exp.category] = 0
            category_totals[exp.category] += float(exp.total_amount)
        
        for category_name, total in category_totals.items():
            category_breakdown.append({
                "category": category_name,
                "total_amount": total,
                "expense_count": len([exp for exp in expenses if exp.category == category_name]),
                "percentage": (total / total_expenses * 100) if total_expenses > 0 else 0
            })
        
        # Vendor breakdown
        vendor_breakdown = []
        vendor_totals = {}
        vendor_counts = {}
        
        for exp in expenses:
            vendor = db.query(Party).filter(Party.id == exp.vendor_id).first() if exp.vendor_id else None
            vendor_name = vendor.name if vendor else "Unknown"
            
            if vendor_name not in vendor_totals:
                vendor_totals[vendor_name] = 0
                vendor_counts[vendor_name] = 0
            
            vendor_totals[vendor_name] += float(exp.total_amount)
            vendor_counts[vendor_name] += 1
        
        for vendor_name, total in vendor_totals.items():
            vendor_breakdown.append({
                "vendor_name": vendor_name,
                "total_amount": total,
                "expense_count": vendor_counts[vendor_name],
                "percentage": (total / total_expenses * 100) if total_expenses > 0 else 0
            })
        
        # Calculate trends
        trends = {
            "monthly_expenses": [],
            "expense_trend": "increasing" if expense_count > 0 else "stable"
        }
        
        # Generate monthly expenses for the last 6 months
        for i in range(6):
            month_start = (datetime.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_expenses = db.query(Expense).filter(
                Expense.expense_date >= month_start.date(),
                Expense.expense_date <= month_end.date()
            ).all()
            
            month_total = sum(float(exp.total_amount) for exp in month_expenses)
            trends["monthly_expenses"].append({
                "month": month_start.strftime('%Y-%m'),
                "expenses": month_total
            })
        
        trends["monthly_expenses"].reverse()
        
        return ExpenseReport(
            period={
                "start_date": start_date,
                "end_date": end_date
            },
            summary=summary,
            transactions=transactions,
            category_breakdown=category_breakdown,
            vendor_breakdown=vendor_breakdown,
            trends=trends,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "category": category,
                "vendor_name": vendor_name,
                "payment_method": payment_method
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating expense report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate expense report")


# Purchase Reports API Endpoints
@api.get('/reports/purchases', response_model=PurchaseReport)
def get_purchase_report(
    start_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    vendor_id: int | None = Query(None, description="Filter by vendor ID"),
    payment_status: str | None = Query(None, description="Filter by payment status"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive purchase report with vendor and product breakdown"""
    
    try:
        from datetime import datetime, timedelta
        
        # Set default date range if not provided
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Convert to datetime objects
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Build base query
        query = db.query(Purchase).join(Party, Purchase.vendor_id == Party.id)
        
        # Apply filters
        query = query.filter(Purchase.date >= start_dt, Purchase.date <= end_dt)
        if vendor_id:
            query = query.filter(Purchase.vendor_id == vendor_id)
        
        purchases = query.all()
        
        # Filter by payment status dynamically if specified
        if payment_status:
            filtered_purchases = []
            for pur in purchases:
                # Calculate payment amount for this purchase
                payments = db.query(PurchasePayment).filter(PurchasePayment.purchase_id == pur.id).all()
                payment_amount = sum(float(p.payment_amount) for p in payments)
                outstanding = float(pur.grand_total) - payment_amount
                
                # Determine payment status
                if payment_amount == 0:
                    calc_status = 'pending'
                elif outstanding <= 0:
                    calc_status = 'paid'
                else:
                    calc_status = 'partial'
                
                # Filter based on requested payment_status
                if calc_status == payment_status:
                    filtered_purchases.append(pur)
            
            purchases = filtered_purchases
        
        # Calculate summary
        total_purchases = sum(float(pur.grand_total) for pur in purchases)
        total_tax = sum(float(pur.cgst + pur.sgst + pur.igst) for pur in purchases)
        total_taxable_value = sum(float(pur.taxable_value) for pur in purchases)
        
        # Calculate payment status breakdown
        paid_amount = 0
        outstanding_amount = 0
        for pur in purchases:
            # Calculate actual payment amount
            payments = db.query(PurchasePayment).filter(PurchasePayment.purchase_id == pur.id).all()
            payment_amount = sum(float(p.payment_amount) for p in payments)
            purchase_outstanding = float(pur.grand_total) - payment_amount
            
            paid_amount += payment_amount
            outstanding_amount += max(0, purchase_outstanding)
        
        summary = {
            "total_purchases": total_purchases,
            "total_tax": total_tax,
            "total_taxable_value": total_taxable_value,
            "paid_amount": paid_amount,
            "outstanding_amount": outstanding_amount,
            "purchase_count": len(purchases),
            "payment_status_breakdown": {
                "paid": paid_amount,
                "outstanding": outstanding_amount
            }
        }
        
        # Build transaction list
        transactions = []
        for pur in purchases:
            vendor = db.query(Party).filter(Party.id == pur.vendor_id).first()
            
            # Calculate payment amount and status
            payments = db.query(PurchasePayment).filter(PurchasePayment.purchase_id == pur.id).all()
            payment_amount = sum(float(p.payment_amount) for p in payments)
            outstanding = float(pur.grand_total) - payment_amount
            
            # Determine payment status dynamically
            if payment_amount == 0:
                calc_payment_status = 'pending'
            elif outstanding <= 0:
                calc_payment_status = 'paid'
            else:
                calc_payment_status = 'partial'
            
            transactions.append(PurchaseReportItem(
                purchase_id=pur.id,
                purchase_no=pur.purchase_no,
                purchase_date=pur.date.isoformat(),
                vendor_name=vendor.name if vendor else "Unknown",
                taxable_value=float(pur.taxable_value),
                total_tax=float(pur.cgst + pur.sgst + pur.igst),
                grand_total=float(pur.grand_total),
                payment_status=calc_payment_status,
                payment_amount=payment_amount,
                outstanding_amount=max(0, outstanding)
            ))
        
        # Vendor breakdown
        vendor_breakdown = []
        vendor_totals = {}
        for pur in purchases:
            vendor = db.query(Party).filter(Party.id == pur.vendor_id).first()
            vendor_name = vendor.name if vendor else "Unknown"
            
            if vendor_name not in vendor_totals:
                vendor_totals[vendor_name] = 0
            vendor_totals[vendor_name] += float(pur.grand_total)
        
        for vendor_name, total in vendor_totals.items():
            vendor_breakdown.append({
                "vendor_name": vendor_name,
                "total_purchases": total,
                "purchase_count": len([pur for pur in purchases if db.query(Party).filter(Party.id == pur.vendor_id).first().name == vendor_name])
            })
        
        # Product breakdown
        product_breakdown = []
        product_totals = {}
        
        for pur in purchases:
            items = db.query(PurchaseItem).filter(PurchaseItem.purchase_id == pur.id).all()
            for item in items:
                product = db.query(Product).filter(Product.id == item.product_id).first()
                product_name = product.name if product else "Unknown"
                
                if product_name not in product_totals:
                    product_totals[product_name] = 0
                product_totals[product_name] += float(item.amount)
        
        for product_name, total in product_totals.items():
            product_breakdown.append({
                "product_name": product_name,
                "total_purchases": total,
                "quantity_purchased": sum(
                    float(item.qty) for pur in purchases 
                    for item in db.query(PurchaseItem).filter(PurchaseItem.purchase_id == pur.id).all()
                    if db.query(Product).filter(Product.id == item.product_id).first().name == product_name
                )
            })
        
        # Calculate trends
        trends = {
            "monthly_purchases": [],
            "purchase_trend": "increasing" if len(purchases) > 0 else "stable"
        }
        
        # Generate monthly purchases for the last 6 months
        for i in range(6):
            month_start = (datetime.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_purchases = db.query(Purchase).filter(
                Purchase.date >= month_start.date(),
                Purchase.date <= month_end.date()
            ).all()
            
            month_total = sum(float(pur.grand_total) for pur in month_purchases)
            trends["monthly_purchases"].append({
                "month": month_start.strftime('%Y-%m'),
                "purchases": month_total
            })
        
        trends["monthly_purchases"].reverse()
        
        return PurchaseReport(
            period={
                "start_date": start_date,
                "end_date": end_date
            },
            summary=summary,
            transactions=transactions,
            vendor_breakdown=vendor_breakdown,
            product_breakdown=product_breakdown,
            trends=trends,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "vendor_id": vendor_id,
                "payment_status": payment_status
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating purchase report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate purchase report")


# Payment Reports API Endpoints
@api.get('/reports/payments', response_model=PaymentReport)
def get_payment_report(
    start_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    payment_type: str | None = Query(None, description="Filter by payment type: invoice_payment, purchase_payment"),
    payment_method: str | None = Query(None, description="Filter by payment method"),
    party_id: int | None = Query(None, description="Filter by party ID"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive payment report with method and party breakdown"""
    
    try:
        from datetime import datetime, timedelta
        
        # Set default date range if not provided
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Convert to datetime objects
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        transactions = []
        
        # Invoice payments
        if not payment_type or payment_type == 'invoice_payment':
            invoice_payment_filter = []
            if start_date:
                invoice_payment_filter.append(Payment.payment_date >= start_dt)
            if end_date:
                invoice_payment_filter.append(Payment.payment_date <= end_dt)
            if payment_method:
                invoice_payment_filter.append(Payment.payment_method == payment_method)
            if party_id:
                # Get invoices for the party
                party_invoices = db.query(Invoice).filter(Invoice.customer_id == party_id).all()
                invoice_ids = [inv.id for inv in party_invoices]
                invoice_payment_filter.append(Payment.invoice_id.in_(invoice_ids))
            
            invoice_payments = db.query(Payment).filter(and_(*invoice_payment_filter)).all()
            for payment in invoice_payments:
                invoice = db.query(Invoice).filter(Invoice.id == payment.invoice_id).first()
                customer = db.query(Party).filter(Party.id == invoice.customer_id).first() if invoice else None
                
                transactions.append(PaymentReportItem(
                    payment_id=payment.id,
                    payment_date=payment.payment_date.isoformat(),
                    payment_type='invoice_payment',
                    payment_method=payment.payment_method,
                    amount=float(payment.payment_amount),
                    reference_type='invoice',
                    reference_id=payment.invoice_id,
                    reference_number=invoice.invoice_no if invoice else None,
                    party_name=customer.name if customer else None,
                    status='completed'
                ))
        
        # Purchase payments
        if not payment_type or payment_type == 'purchase_payment':
            purchase_payment_filter = []
            if start_date:
                purchase_payment_filter.append(PurchasePayment.payment_date >= start_dt)
            if end_date:
                purchase_payment_filter.append(PurchasePayment.payment_date <= end_dt)
            if payment_method:
                purchase_payment_filter.append(PurchasePayment.payment_method == payment_method)
            if party_id:
                # Get purchases for the party
                party_purchases = db.query(Purchase).filter(Purchase.vendor_id == party_id).all()
                purchase_ids = [pur.id for pur in party_purchases]
                purchase_payment_filter.append(PurchasePayment.purchase_id.in_(purchase_ids))
            
            purchase_payments = db.query(PurchasePayment).filter(and_(*purchase_payment_filter)).all()
            for payment in purchase_payments:
                purchase = db.query(Purchase).filter(Purchase.id == payment.purchase_id).first()
                vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first() if purchase else None
                
                transactions.append(PaymentReportItem(
                    payment_id=payment.id,
                    payment_date=payment.payment_date.isoformat(),
                    payment_type='purchase_payment',
                    payment_method=payment.payment_method,
                    amount=float(payment.payment_amount),
                    reference_type='purchase',
                    reference_id=payment.purchase_id,
                    reference_number=purchase.purchase_no if purchase else None,
                    party_name=vendor.name if vendor else None,
                    status='completed'
                ))
        
        # Calculate summary
        total_payments = sum(float(t.amount) for t in transactions)
        payment_count = len(transactions)
        
        summary = {
            "total_payments": total_payments,
            "payment_count": payment_count,
            "average_payment": total_payments / payment_count if payment_count > 0 else 0
        }
        
        # Method breakdown
        method_breakdown = []
        method_totals = {}
        for t in transactions:
            if t.payment_method not in method_totals:
                method_totals[t.payment_method] = 0
            method_totals[t.payment_method] += float(t.amount)
        
        for method, total in method_totals.items():
            method_breakdown.append({
                "payment_method": method,
                "total_amount": total,
                "payment_count": len([t for t in transactions if t.payment_method == method]),
                "percentage": (total / total_payments * 100) if total_payments > 0 else 0
            })
        
        # Party breakdown
        party_breakdown = []
        party_totals = {}
        for t in transactions:
            party_name = t.party_name or "Unknown"
            if party_name not in party_totals:
                party_totals[party_name] = 0
            party_totals[party_name] += float(t.amount)
        
        for party_name, total in party_totals.items():
            party_breakdown.append({
                "party_name": party_name,
                "total_amount": total,
                "payment_count": len([t for t in transactions if (t.party_name or "Unknown") == party_name]),
                "percentage": (total / total_payments * 100) if total_payments > 0 else 0
            })
        
        # Calculate trends
        trends = {
            "monthly_payments": [],
            "payment_trend": "increasing" if payment_count > 0 else "stable"
        }
        
        # Generate monthly payments for the last 6 months
        for i in range(6):
            month_start = (datetime.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_invoice_payments = db.query(Payment).filter(
                Payment.payment_date >= month_start.date(),
                Payment.payment_date <= month_end.date()
            ).all()
            
            month_purchase_payments = db.query(PurchasePayment).filter(
                PurchasePayment.payment_date >= month_start.date(),
                PurchasePayment.payment_date <= month_end.date()
            ).all()
            
            month_total = sum(float(p.payment_amount) for p in month_invoice_payments) + sum(float(p.payment_amount) for p in month_purchase_payments)
            trends["monthly_payments"].append({
                "month": month_start.strftime('%Y-%m'),
                "payments": month_total
            })
        
        trends["monthly_payments"].reverse()
        
        return PaymentReport(
            period={
                "start_date": start_date,
                "end_date": end_date
            },
            summary=summary,
            transactions=transactions,
            method_breakdown=method_breakdown,
            party_breakdown=party_breakdown,
            trends=trends,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "payment_type": payment_type,
                "payment_method": payment_method,
                "party_id": party_id
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating payment report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate payment report")


# Financial Reports API Endpoints
@api.get('/reports/financial', response_model=FinancialReport)
def get_financial_report(
    report_type: str = Query(..., description="Report type: profit_loss, balance_sheet, cash_flow"),
    start_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    as_of_date: str | None = Query(None, description="As of date for balance sheet (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate comprehensive financial reports (P&L, Balance Sheet, Cash Flow)"""
    
    try:
        from datetime import datetime
        from .financial_reports import FinancialReports
        
        # Initialize financial reports service
        financial_service = FinancialReports(db)
        
        if report_type == "profit_loss":
            # Convert dates for P&L
            start_dt = None
            end_dt = None
            if start_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            data = financial_service.generate_profit_loss_statement(start_dt, end_dt)
            
        elif report_type == "balance_sheet":
            # Convert date for balance sheet
            as_of_dt = None
            if as_of_date:
                as_of_dt = datetime.strptime(as_of_date, '%Y-%m-%d').date()
            
            data = financial_service.generate_balance_sheet(as_of_dt)
            
        elif report_type == "cash_flow":
            # Convert dates for cash flow
            start_dt = None
            end_dt = None
            if start_date:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            data = financial_service.generate_cash_flow_statement(start_dt, end_dt)
            
        else:
            raise HTTPException(status_code=400, detail="Invalid report type. Use: profit_loss, balance_sheet, cash_flow")
        
        return FinancialReport(
            report_type=report_type,
            period={
                "start_date": start_date,
                "end_date": end_date,
                "as_of_date": as_of_date
            },
            data=data,
            generated_at=datetime.now().isoformat(),
            filters_applied={
                "report_type": report_type
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating financial report: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate financial report")


def _calculate_period_dates(period_type: str, period_value: str) -> tuple[str, str]:
    """Calculate start and end dates for the given period"""
    if period_type == "month":
        # period_value format: YYYY-MM
        year, month = period_value.split("-")
        start_date = f"{year}-{month}-01"
        # Get last day of month
        last_day = calendar.monthrange(int(year), int(month))[1]
        end_date = f"{year}-{month}-{last_day}"
    elif period_type == "quarter":
        # period_value format: YYYY-Q1/Q2/Q3/Q4
        year, quarter = period_value.split("-")
        year = int(year)
        quarter = int(quarter[1])  # Q1 -> 1, Q2 -> 2, etc.
        
        start_month = (quarter - 1) * 3 + 1
        end_month = quarter * 3
        
        start_date = f"{year}-{start_month:02d}-01"
        last_day = calendar.monthrange(year, end_month)[1]
        end_date = f"{year}-{end_month:02d}-{last_day}"
    elif period_type == "year":
        # period_value format: YYYY
        year = int(period_value)
        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"
    else:
        raise HTTPException(status_code=400, detail="Invalid period type. Use month, quarter, or year")
    
    return start_date, end_date


def _generate_gstr1_report(start_date: str, end_date: str, format: str, db: Session):
    """Generate GSTR-1 (Outward Supplies) report"""
    from sqlalchemy import func, select, cast, Date
    from .models import Invoice, InvoiceItem, Product, Party
    from datetime import datetime
    
    # Convert string dates to proper date objects
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get all invoices in the period
    invoices = db.query(Invoice).filter(
        cast(Invoice.date, Date) >= start_date_obj,
        cast(Invoice.date, Date) <= end_date_obj
    ).all()
    
    gstr1_data = {
        "period": f"{start_date} to {end_date}",
        "report_type": "GSTR-1",
        "generated_on": datetime.now().isoformat(),
        "sections": {
            "b2b": [],
            "b2c": [],
            "nil_rated": [],
            "exempted": [],
            "rate_wise_summary": []
        }
    }
    
    # Process each invoice
    for invoice in invoices:
        customer = db.query(Party).filter(Party.id == invoice.customer_id).first()
        
        # Determine if B2B or B2C based on customer GSTIN
        is_b2b = customer and customer.gstin and len(customer.gstin) == 15
        
        invoice_data = {
            "invoice_no": invoice.invoice_no,
            "invoice_date": invoice.date.isoformat(),
            "customer_name": customer.name if customer else "Unknown",
            "customer_gstin": customer.gstin if customer else "",
            "place_of_supply": invoice.place_of_supply,
            "reverse_charge": invoice.reverse_charge,
            "items": [],
            "total_taxable_value": float(invoice.taxable_value),
            "total_cgst": float(invoice.cgst),
            "total_sgst": float(invoice.sgst),
            "total_igst": float(invoice.igst),
            "grand_total": float(invoice.grand_total)
        }
        
        # Get invoice items
        items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice.id).all()
        for item in items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            item_data = {
                "description": item.description,
                "hsn_code": product.hsn if product else "",
                "qty": float(item.qty),
                "rate": float(item.rate),
                "taxable_value": float(item.taxable_value),
                "gst_rate": float(item.gst_rate),
                "cgst": float(item.cgst),
                "sgst": float(item.sgst),
                "igst": float(item.igst),
                "amount": float(item.amount)
            }
            invoice_data["items"].append(item_data)
        
        # Categorize invoice
        if is_b2b:
            gstr1_data["sections"]["b2b"].append(invoice_data)
        else:
            gstr1_data["sections"]["b2c"].append(invoice_data)
    
    # Generate rate-wise summary
    rate_summary = db.execute(
        select(Product.gst_rate, func.sum(InvoiceItem.taxable_value), func.sum(InvoiceItem.cgst), func.sum(InvoiceItem.sgst), func.sum(InvoiceItem.igst))
        .join(Product, Product.id == InvoiceItem.product_id)
        .join(Invoice, Invoice.id == InvoiceItem.invoice_id)
        .filter(
            cast(Invoice.date, Date) >= start_date_obj,
            cast(Invoice.date, Date) <= end_date_obj
        )
        .group_by(Product.gst_rate)
    ).all()
    
    for rate, taxable, cgst, sgst, igst in rate_summary:
        gstr1_data["sections"]["rate_wise_summary"].append({
            "gst_rate": float(rate),
            "taxable_value": float(taxable),
            "cgst": float(cgst),
            "sgst": float(sgst),
            "igst": float(igst)
        })
    
    return _format_report(gstr1_data, format, "gstr1")


def _generate_gstr2_report(start_date: str, end_date: str, format: str, db: Session):
    """Generate GSTR-2 (Inward Supplies) report"""
    from sqlalchemy import func, select, cast, Date
    from .models import Purchase, PurchaseItem, Product, Party
    from datetime import datetime
    
    # Convert string dates to proper date objects
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get all purchases in the period
    purchases = db.query(Purchase).filter(
        cast(Purchase.date, Date) >= start_date_obj,
        cast(Purchase.date, Date) <= end_date_obj
    ).all()
    
    gstr2_data = {
        "period": f"{start_date} to {end_date}",
        "report_type": "GSTR-2",
        "generated_on": datetime.now().isoformat(),
        "sections": {
            "b2b": [],
            "imports": [],
            "rate_wise_summary": []
        }
    }
    
    # Process each purchase
    for purchase in purchases:
        vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first()
        
        purchase_data = {
            "purchase_no": purchase.purchase_no,
            "purchase_date": purchase.date.isoformat(),
            "vendor_name": vendor.name if vendor else "Unknown",
            "vendor_gstin": vendor.gstin if vendor else "",
            "place_of_supply": purchase.place_of_supply,
            "reverse_charge": purchase.reverse_charge,
            "items": [],
            "total_taxable_value": float(purchase.taxable_value),
            "total_cgst": float(purchase.cgst),
            "total_sgst": float(purchase.sgst),
            "total_igst": float(purchase.igst),
            "grand_total": float(purchase.grand_total)
        }
        
        # Get purchase items
        items = db.query(PurchaseItem).filter(PurchaseItem.purchase_id == purchase.id).all()
        for item in items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            item_data = {
                "description": item.description,
                "hsn_code": product.hsn if product else "",
                "qty": float(item.qty),
                "rate": float(item.rate),
                "taxable_value": float(item.taxable_value),
                "gst_rate": float(item.gst_rate),
                "cgst": float(item.cgst),
                "sgst": float(item.sgst),
                "igst": float(item.igst),
                "amount": float(item.amount)
            }
            purchase_data["items"].append(item_data)
        
        gstr2_data["sections"]["b2b"].append(purchase_data)
    
    # Generate rate-wise summary
    rate_summary = db.execute(
        select(Product.gst_rate, func.sum(PurchaseItem.taxable_value), func.sum(PurchaseItem.cgst), func.sum(PurchaseItem.sgst), func.sum(PurchaseItem.igst))
        .join(Product, Product.id == PurchaseItem.product_id)
        .join(Purchase, Purchase.id == PurchaseItem.purchase_id)
        .filter(func.date(Purchase.date) >= func.date(start_date), func.date(Purchase.date) <= func.date(end_date))
        .group_by(Product.gst_rate)
    ).all()
    
    for rate, taxable, cgst, sgst, igst in rate_summary:
        gstr2_data["sections"]["rate_wise_summary"].append({
            "gst_rate": float(rate),
            "taxable_value": float(taxable),
            "cgst": float(cgst),
            "sgst": float(sgst),
            "igst": float(igst)
        })
    
    return _format_report(gstr2_data, format, "gstr2")


def _generate_gstr3b_report(start_date: str, end_date: str, format: str, db: Session):
    """Generate GSTR-3B (Summary) report"""
    from sqlalchemy import func, select, cast, Date
    from .models import Invoice, Purchase, InvoiceItem, PurchaseItem, Product
    from datetime import datetime
    
    # Convert string dates to proper date objects
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    gstr3b_data = {
        "period": f"{start_date} to {end_date}",
        "report_type": "GSTR-3B",
        "generated_on": datetime.now().isoformat(),
        "sections": {
            "outward_supplies": {},
            "inward_supplies": {},
            "summary": {}
        }
    }
    
    # Outward supplies (GSTR-1 data)
    invoices = db.query(Invoice).filter(
        cast(Invoice.date, Date) >= start_date_obj,
        cast(Invoice.date, Date) <= end_date_obj
    ).all()
    
    outward_taxable = sum([float(i.taxable_value) for i in invoices], 0.0)
    outward_cgst = sum([float(i.cgst) for i in invoices], 0.0)
    outward_sgst = sum([float(i.sgst) for i in invoices], 0.0)
    outward_igst = sum([float(i.igst) for i in invoices], 0.0)
    
    gstr3b_data["sections"]["outward_supplies"] = {
        "total_taxable_value": outward_taxable,
        "total_cgst": outward_cgst,
        "total_sgst": outward_sgst,
        "total_igst": outward_igst,
        "total_tax": outward_cgst + outward_sgst + outward_igst
    }
    
    # Inward supplies (GSTR-2 data)
    purchases = db.query(Purchase).filter(
        cast(Purchase.date, Date) >= start_date_obj,
        cast(Purchase.date, Date) <= end_date_obj
    ).all()
    
    inward_taxable = sum([float(p.taxable_value) for p in purchases], 0.0)
    inward_cgst = sum([float(p.cgst) for p in purchases], 0.0)
    inward_sgst = sum([float(p.sgst) for p in purchases], 0.0)
    inward_igst = sum([float(p.igst) for p in purchases], 0.0)
    
    gstr3b_data["sections"]["inward_supplies"] = {
        "total_taxable_value": inward_taxable,
        "total_cgst": inward_cgst,
        "total_sgst": inward_sgst,
        "total_igst": inward_igst,
        "total_tax": inward_cgst + inward_sgst + inward_igst
    }
    
    # Summary calculations
    net_cgst = outward_cgst - inward_cgst
    net_sgst = outward_sgst - inward_sgst
    net_igst = outward_igst - inward_igst
    
    gstr3b_data["sections"]["summary"] = {
        "net_cgst": net_cgst,
        "net_sgst": net_sgst,
        "net_igst": net_igst,
        "total_net_tax": net_cgst + net_sgst + net_igst,
        "total_outward_taxable": outward_taxable,
        "total_inward_taxable": inward_taxable
    }
    
    return _format_report(gstr3b_data, format, "gstr3b")


def _format_report(data: dict, format: str, report_type: str):
    """Format report in requested format"""
    if format == "json":
        return data
    elif format == "csv":
        return _convert_to_csv(data, report_type)
    elif format == "excel":
        return _convert_to_excel(data, report_type)
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use json, csv, or excel")


def _convert_to_csv(data: dict, report_type: str):
    """Convert report data to CSV format"""
    import csv
    from io import StringIO
    
    buf = StringIO()
    w = csv.writer(buf)
    
    # Write header
    w.writerow([f"GST {report_type.upper()} Report"])
    w.writerow([f"Period: {data['period']}"])
    w.writerow([f"Generated On: {data['generated_on']}"])
    w.writerow([])
    
    if report_type == "gstr1":
        # B2B Section
        w.writerow(["B2B Invoices"])
        w.writerow(["Invoice No", "Date", "Customer", "GSTIN", "Taxable Value", "CGST", "SGST", "IGST", "Total"])
        for invoice in data["sections"]["b2b"]:
            w.writerow([
                invoice["invoice_no"],
                invoice["invoice_date"],
                invoice["customer_name"],
                invoice["customer_gstin"],
                invoice["total_taxable_value"],
                invoice["total_cgst"],
                invoice["total_sgst"],
                invoice["total_igst"],
                invoice["grand_total"]
            ])
        
        w.writerow([])
        w.writerow(["Rate-wise Summary"])
        w.writerow(["GST Rate", "Taxable Value", "CGST", "SGST", "IGST"])
        for rate in data["sections"]["rate_wise_summary"]:
            w.writerow([
                rate["gst_rate"],
                rate["taxable_value"],
                rate["cgst"],
                rate["sgst"],
                rate["igst"]
            ])
    
    elif report_type == "gstr2":
        # B2B Section
        w.writerow(["B2B Purchases"])
        w.writerow(["Purchase No", "Date", "Vendor", "GSTIN", "Taxable Value", "CGST", "SGST", "IGST", "Total"])
        for purchase in data["sections"]["b2b"]:
            w.writerow([
                purchase["purchase_no"],
                purchase["purchase_date"],
                purchase["vendor_name"],
                purchase["vendor_gstin"],
                purchase["total_taxable_value"],
                purchase["total_cgst"],
                purchase["total_sgst"],
                purchase["total_igst"],
                purchase["grand_total"]
            ])
    
    elif report_type == "gstr3b":
        w.writerow(["GSTR-3B Summary"])
        w.writerow(["Section", "Taxable Value", "CGST", "SGST", "IGST", "Total Tax"])
        
        outward = data["sections"]["outward_supplies"]
        w.writerow(["Outward Supplies", outward["total_taxable_value"], outward["total_cgst"], outward["total_sgst"], outward["total_igst"], outward["total_tax"]])
        
        inward = data["sections"]["inward_supplies"]
        w.writerow(["Inward Supplies", inward["total_taxable_value"], inward["total_cgst"], inward["total_sgst"], inward["total_igst"], inward["total_tax"]])
        
        summary = data["sections"]["summary"]
        w.writerow(["Net Tax", "", summary["net_cgst"], summary["net_sgst"], summary["net_igst"], summary["total_net_tax"]])
    
    return Response(content=buf.getvalue(), media_type='text/csv')


def _convert_to_excel(data: dict, report_type: str):
    """Convert report data to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl package")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"GST {report_type.upper()}"
    
    # Header styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write header
    ws['A1'] = f"GST {report_type.upper()} Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Period: {data['period']}"
    ws['A3'] = f"Generated On: {data['generated_on']}"
    
    if report_type == "gstr1":
        # B2B Section
        row = 5
        ws[f'A{row}'] = "B2B Invoices"
        ws[f'A{row}'].font = header_font
        
        row += 1
        headers = ["Invoice No", "Date", "Customer", "GSTIN", "Taxable Value", "CGST", "SGST", "IGST", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        for invoice in data["sections"]["b2b"]:
            ws.cell(row=row, column=1, value=invoice["invoice_no"])
            ws.cell(row=row, column=2, value=invoice["invoice_date"])
            ws.cell(row=row, column=3, value=invoice["customer_name"])
            ws.cell(row=row, column=4, value=invoice["customer_gstin"])
            ws.cell(row=row, column=5, value=invoice["total_taxable_value"])
            ws.cell(row=row, column=6, value=invoice["total_cgst"])
            ws.cell(row=row, column=7, value=invoice["total_sgst"])
            ws.cell(row=row, column=8, value=invoice["total_igst"])
            ws.cell(row=row, column=9, value=invoice["grand_total"])
            row += 1
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="gst_{report_type}_{data["period"].replace(" ", "_")}.xlsx"'}
    )


class StockRow(BaseModel):
    product_id: int
    sku: str
    name: str
    onhand: int
    item_type: str
    unit: str


@api.get('/stock/summary', response_model=list[StockRow])
def stock_summary(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows: list[StockRow] = []
    products = db.query(Product).order_by(Product.id).all()
    for p in products:
        # Use the product.stock field directly as it's updated by all operations
        onhand = int(p.stock)
        rows.append(StockRow(product_id=p.id, sku=p.sku or '', name=p.name, onhand=onhand, item_type=p.item_type, unit=p.unit))
    return rows


class PurchaseItemIn(BaseModel):
    product_id: int
    qty: float
    rate: float
    description: str | None = None
    hsn_code: str | None = None
    discount: float = 0
    discount_type: str = "Percentage"  # Percentage, Fixed
    gst_rate: float = 0


class PurchaseCreate(BaseModel):
    vendor_id: int
    date: str  # ISO date string
    due_date: str  # ISO date string
    terms: str = "Due on Receipt"
    place_of_supply: str
    place_of_supply_state_code: str
    eway_bill_number: str | None = None
    reverse_charge: bool = False
    export_supply: bool = False
    bill_from_address: str
    ship_from_address: str
    total_discount: float = 0
    notes: str | None = None
    items: list[PurchaseItemIn]


class PurchaseOut(BaseModel):
    id: int
    purchase_no: str
    vendor_id: int
    vendor_name: str
    date: str
    due_date: str
    terms: str
    place_of_supply: str
    place_of_supply_state_code: str
    eway_bill_number: str | None
    reverse_charge: bool
    export_supply: bool
    bill_from_address: str
    ship_from_address: str
    taxable_value: float
    total_discount: float
    cgst: float
    sgst: float
    igst: float
    grand_total: float
    paid_amount: float
    balance_amount: float
    notes: str | None
    status: str
    created_at: str
    updated_at: str


def _next_purchase_no(db: Session) -> str:
    """Generate next purchase number"""
    settings = db.query(CompanySettings).first()
    if not settings:
        raise HTTPException(status_code=500, detail="Company settings not found")
    
    prefix = settings.invoice_series.replace("INV", "PUR")  # Use similar series for purchases
    last_purchase = db.query(Purchase).filter(Purchase.purchase_no.like(f"{prefix}%")).order_by(Purchase.purchase_no.desc()).first()
    
    if last_purchase:
        # Extract sequence number from last purchase number
        last_seq = int(last_purchase.purchase_no.replace(prefix, ""))
        seq = last_seq + 1
    else:
        seq = 1
    
    # Ensure the total length doesn't exceed 16 characters
    max_seq_length = 16 - len(prefix)
    if max_seq_length < 1:
        prefix = prefix[:15]
        max_seq_length = 16 - len(prefix)
    
    # Format sequence number with appropriate padding
    if max_seq_length >= 5:
        seq_str = f"{seq:05d}"
    elif max_seq_length >= 4:
        seq_str = f"{seq:04d}"
    elif max_seq_length >= 3:
        seq_str = f"{seq:03d}"
    elif max_seq_length >= 2:
        seq_str = f"{seq:02d}"
    else:
        seq_str = str(seq)
    
    # Ensure final length doesn't exceed 16
    result = f"{prefix}{seq_str}"
    if len(result) > 16:
        result = result[:16]
    
    return result


@api.post('/purchases', status_code=201)
def create_purchase(payload: PurchaseCreate, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    vendor = db.query(Party).filter(Party.id == payload.vendor_id, Party.type == 'vendor').first()
    if not vendor:
        raise HTTPException(status_code=400, detail='Invalid vendor')
    
    # Generate purchase number
    purchase_no = _next_purchase_no(db)
    
    # Calculate GST and totals
    taxable_value = Decimal('0.00')
    total_cgst = Decimal('0.00')
    total_sgst = Decimal('0.00')
    total_igst = Decimal('0.00')
    
    # Create purchase
    pur = Purchase(
        vendor_id=vendor.id,
        purchase_no=purchase_no,
        date=datetime.fromisoformat(payload.date),
        due_date=datetime.fromisoformat(payload.due_date),
        terms=payload.terms,
        place_of_supply=payload.place_of_supply,
        place_of_supply_state_code=payload.place_of_supply_state_code,
        eway_bill_number=payload.eway_bill_number,
        reverse_charge=payload.reverse_charge,
        export_supply=payload.export_supply,
        bill_from_address=payload.bill_from_address,
        ship_from_address=payload.ship_from_address,
        total_discount=money(payload.total_discount),
        notes=payload.notes,
        taxable_value=Decimal('0.00'),
        cgst=Decimal('0.00'),
        sgst=Decimal('0.00'),
        igst=Decimal('0.00'),
        grand_total=Decimal('0.00'),
        paid_amount=Decimal('0.00'),
        balance_amount=Decimal('0.00'),
        status="Draft"
    )
    db.add(pur)
    db.flush()
    
    # Process items
    for it in payload.items:
        prod = db.query(Product).filter(Product.id == it.product_id).first()
        if not prod:
            raise HTTPException(status_code=400, detail='Invalid product')
        
        # Calculate item amounts
        base_amount = Decimal(it.qty) * Decimal(it.rate)
        discount_amount = Decimal(it.discount) if it.discount_type == "Fixed" else (base_amount * Decimal(it.discount) / 100)
        taxable_amount = base_amount - discount_amount
        
        # Calculate GST
        gst_rate = Decimal(it.gst_rate)
        if payload.place_of_supply_state_code == "29":  # Intra-state
            cgst = taxable_amount * gst_rate / 200  # Half of GST rate
            sgst = taxable_amount * gst_rate / 200
            igst = Decimal('0.00')
        else:  # Inter-state
            cgst = Decimal('0.00')
            sgst = Decimal('0.00')
            igst = taxable_amount * gst_rate / 100
        
        item_amount = taxable_amount + cgst + sgst + igst
        
        # Create purchase item
        purchase_item = PurchaseItem(
            purchase_id=pur.id,
            product_id=prod.id,
            description=it.description or prod.name,
            hsn_code=it.hsn_code or prod.hsn,
            qty=it.qty,
            rate=money(it.rate),
            discount=money(it.discount),
            discount_type=it.discount_type,
            taxable_value=money(taxable_amount),
            gst_rate=it.gst_rate,
            cgst=money(cgst),
            sgst=money(sgst),
            igst=money(igst),
            amount=money(item_amount)
        )
        db.add(purchase_item)
        
        # Update totals
        taxable_value += taxable_amount
        total_cgst += cgst
        total_sgst += sgst
        total_igst += igst
        
        # Add stock ledger entry
        db.add(StockLedgerEntry(product_id=prod.id, qty=it.qty, entry_type='in', ref_type='purchase', ref_id=pur.id))
    
    # Update purchase totals
    pur.taxable_value = money(taxable_value)
    pur.cgst = money(total_cgst)
    pur.sgst = money(total_sgst)
    pur.igst = money(total_igst)
    pur.grand_total = money(taxable_value + total_cgst + total_sgst + total_igst - Decimal(payload.total_discount))
    pur.balance_amount = pur.grand_total
    
    db.commit()
    return {"id": pur.id, "purchase_no": purchase_no}


@api.get('/purchases', response_model=list[PurchaseOut])
def list_purchases(
    search: str | None = None,
    status: str | None = None,
    vendor_id: int | None = None,
    amount_min: float | None = None,
    amount_max: float | None = None,
    payment_status: str | None = None,
    place_of_supply: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    query = db.query(Purchase).join(Party).filter(Party.type == 'vendor')
    
    if search:
        search_filter = or_(
            Purchase.purchase_no.ilike(f'%{search}%'),
            Party.name.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    if status:
        query = query.filter(Purchase.status == status)
    
    if vendor_id:
        query = query.filter(Purchase.vendor_id == vendor_id)
    
    if amount_min is not None:
        query = query.filter(Purchase.grand_total >= amount_min)
    
    if amount_max is not None:
        query = query.filter(Purchase.grand_total <= amount_max)
    
    if payment_status:
        if payment_status == 'paid':
            query = query.filter(Purchase.paid_amount >= Purchase.grand_total)
        elif payment_status == 'partially_paid':
            query = query.filter(
                Purchase.paid_amount > 0,
                Purchase.paid_amount < Purchase.grand_total
            )
        elif payment_status == 'unpaid':
            query = query.filter(Purchase.paid_amount == 0)
        elif payment_status == 'overdue':
            query = query.filter(
                Purchase.due_date < func.date(func.now()),
                Purchase.paid_amount < Purchase.grand_total
            )
    
    if place_of_supply:
        query = query.filter(Purchase.place_of_supply.ilike(f"%{place_of_supply}%"))
    
    if date_from:
        query = query.filter(Purchase.date >= datetime.fromisoformat(date_from))
    
    if date_to:
        query = query.filter(Purchase.date <= datetime.fromisoformat(date_to))
    
    purchases = query.order_by(Purchase.date.desc()).all()
    
    result = []
    for pur in purchases:
        vendor = db.query(Party).filter(Party.id == pur.vendor_id).first()
        result.append(PurchaseOut(
            id=pur.id,
            purchase_no=pur.purchase_no,
            vendor_id=pur.vendor_id,
            vendor_name=vendor.name if vendor else "Unknown",
            date=pur.date.isoformat(),
            due_date=pur.due_date.isoformat(),
            terms=pur.terms,
            place_of_supply=pur.place_of_supply,
            place_of_supply_state_code=pur.place_of_supply_state_code,
            eway_bill_number=pur.eway_bill_number,
            reverse_charge=pur.reverse_charge,
            export_supply=pur.export_supply,
            bill_from_address=pur.bill_from_address,
            ship_from_address=pur.ship_from_address,
            taxable_value=float(pur.taxable_value),
            total_discount=float(pur.total_discount),
            cgst=float(pur.cgst),
            sgst=float(pur.sgst),
            igst=float(pur.igst),
            grand_total=float(pur.grand_total),
            paid_amount=float(pur.paid_amount),
            balance_amount=float(pur.balance_amount),
            notes=pur.notes,
            status=pur.status,
            created_at=pur.created_at.isoformat(),
            updated_at=pur.updated_at.isoformat()
        ))
    
    return result


@api.get('/purchases/{purchase_id}', response_model=PurchaseOut)
def get_purchase(purchase_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first()
    return PurchaseOut(
        id=purchase.id,
        purchase_no=purchase.purchase_no,
        vendor_id=purchase.vendor_id,
        vendor_name=vendor.name if vendor else "Unknown",
        date=purchase.date.isoformat(),
        due_date=purchase.due_date.isoformat(),
        terms=purchase.terms,
        place_of_supply=purchase.place_of_supply,
        place_of_supply_state_code=purchase.place_of_supply_state_code,
        eway_bill_number=purchase.eway_bill_number,
        reverse_charge=purchase.reverse_charge,
        export_supply=purchase.export_supply,
        bill_from_address=purchase.bill_from_address,
        ship_from_address=purchase.ship_from_address,
        taxable_value=float(purchase.taxable_value),
        total_discount=float(purchase.total_discount),
        cgst=float(purchase.cgst),
        sgst=float(purchase.sgst),
        igst=float(purchase.igst),
        grand_total=float(purchase.grand_total),
        paid_amount=float(purchase.paid_amount),
        balance_amount=float(purchase.balance_amount),
        notes=purchase.notes,
        status=purchase.status,
        created_at=purchase.created_at.isoformat(),
        updated_at=purchase.updated_at.isoformat()
    )


@api.delete('/purchases/{purchase_id}')
def delete_purchase(purchase_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    if purchase.status in ["Paid", "Partially Paid"]:
        raise HTTPException(status_code=400, detail='Cannot delete purchase with payments')
    
    # Delete related records
    db.query(PurchaseItem).filter(PurchaseItem.purchase_id == purchase_id).delete()
    db.query(StockLedgerEntry).filter(
        StockLedgerEntry.ref_type == 'purchase', 
        StockLedgerEntry.ref_id == purchase_id
    ).delete()
    
    db.delete(purchase)
    db.commit()
    return {"message": "Purchase deleted successfully"}


# Expense Management for Cashflow
class ExpenseIn(BaseModel):
    expense_date: str  # ISO date string
    expense_type: str
    category: str
    subcategory: str | None = None
    description: str
    amount: float
    payment_method: str
    account_head: str
    reference_number: str | None = None
    vendor_id: int | None = None
    gst_rate: float = 0
    notes: str | None = None


class ExpenseOut(BaseModel):
    id: int
    expense_date: str
    expense_type: str
    category: str
    subcategory: str | None
    description: str
    amount: float
    payment_method: str
    account_head: str
    reference_number: str | None
    vendor_id: int | None
    vendor_name: str | None
    gst_amount: float
    gst_rate: float
    total_amount: float
    notes: str | None
    created_at: str
    updated_at: str


@api.post('/expenses', status_code=201)
def create_expense(payload: ExpenseIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Validate vendor if provided
    vendor = None
    if payload.vendor_id:
        vendor = db.query(Party).filter(Party.id == payload.vendor_id, Party.type == 'vendor').first()
        if not vendor:
            raise HTTPException(status_code=400, detail='Invalid vendor')
    
    # Calculate GST and total
    gst_amount = Decimal(payload.amount) * Decimal(payload.gst_rate) / 100
    total_amount = Decimal(payload.amount) + gst_amount
    
    expense = Expense(
        expense_date=datetime.fromisoformat(payload.expense_date),
        expense_type=payload.expense_type,
        category=payload.category,
        subcategory=payload.subcategory,
        description=payload.description,
        amount=money(payload.amount),
        payment_method=payload.payment_method,
        account_head=payload.account_head,
        reference_number=payload.reference_number,
        vendor_id=payload.vendor_id,
        gst_amount=money(gst_amount),
        gst_rate=payload.gst_rate,
        total_amount=money(total_amount),
        notes=payload.notes
    )
    
    db.add(expense)
    db.commit()
    return {"id": expense.id}


@api.get('/expenses', response_model=list[ExpenseOut])
def list_expenses(
    search: str | None = None,
    category: str | None = None,
    expense_type: str | None = None,
    payment_method: str | None = None,
    amount_min: float | None = None,
    amount_max: float | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    query = db.query(Expense)
    
    if search:
        search_filter = or_(
            Expense.description.ilike(f'%{search}%'),
            Expense.expense_type.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    if category:
        query = query.filter(Expense.category == category)
    
    if expense_type:
        query = query.filter(Expense.expense_type == expense_type)
    
    if payment_method:
        query = query.filter(Expense.payment_method == payment_method)
    
    if amount_min is not None:
        query = query.filter(Expense.total_amount >= amount_min)
    
    if amount_max is not None:
        query = query.filter(Expense.total_amount <= amount_max)
    
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.fromisoformat(start_date))
    
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.fromisoformat(end_date))
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    result = []
    for exp in expenses:
        vendor_name = None
        if exp.vendor_id:
            vendor = db.query(Party).filter(Party.id == exp.vendor_id).first()
            vendor_name = vendor.name if vendor else "Unknown"
        
        result.append(ExpenseOut(
            id=exp.id,
            expense_date=exp.expense_date.isoformat(),
            expense_type=exp.expense_type,
            category=exp.category,
            subcategory=exp.subcategory,
            description=exp.description,
            amount=float(exp.amount),
            payment_method=exp.payment_method,
            account_head=exp.account_head,
            reference_number=exp.reference_number,
            vendor_id=exp.vendor_id,
            vendor_name=vendor_name,
            gst_amount=float(exp.gst_amount),
            gst_rate=exp.gst_rate,
            total_amount=float(exp.total_amount),
            notes=exp.notes,
            created_at=exp.created_at.isoformat(),
            updated_at=exp.updated_at.isoformat()
        ))
    
    return result


@api.get('/expenses/{expense_id}', response_model=ExpenseOut)
def get_expense(expense_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail='Expense not found')
    
    vendor_name = None
    if expense.vendor_id:
        vendor = db.query(Party).filter(Party.id == expense.vendor_id).first()
        vendor_name = vendor.name if vendor else "Unknown"
    
    return ExpenseOut(
        id=expense.id,
        expense_date=expense.expense_date.isoformat(),
        expense_type=expense.expense_type,
        category=expense.category,
        subcategory=expense.subcategory,
        description=expense.description,
        amount=float(expense.amount),
        payment_method=expense.payment_method,
        account_head=expense.account_head,
        reference_number=expense.reference_number,
        vendor_id=expense.vendor_id,
        vendor_name=vendor_name,
        gst_amount=float(expense.gst_amount),
        gst_rate=expense.gst_rate,
        total_amount=float(expense.total_amount),
        notes=expense.notes,
        created_at=expense.created_at.isoformat(),
        updated_at=expense.updated_at.isoformat()
    )


@api.put('/expenses/{expense_id}')
def update_expense(expense_id: int, payload: ExpenseIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail='Expense not found')
    
    # Validate vendor if provided
    if payload.vendor_id:
        vendor = db.query(Party).filter(Party.id == payload.vendor_id, Party.type == 'vendor').first()
        if not vendor:
            raise HTTPException(status_code=400, detail='Invalid vendor')
    
    # Calculate GST and total
    gst_amount = Decimal(payload.amount) * Decimal(payload.gst_rate) / 100
    total_amount = Decimal(payload.amount) + gst_amount
    
    # Update expense
    expense.expense_date = datetime.fromisoformat(payload.expense_date)
    expense.expense_type = payload.expense_type
    expense.category = payload.category
    expense.subcategory = payload.subcategory
    expense.description = payload.description
    expense.amount = money(payload.amount)
    expense.payment_method = payload.payment_method
    expense.account_head = payload.account_head
    expense.reference_number = payload.reference_number
    expense.vendor_id = payload.vendor_id
    expense.gst_amount = money(gst_amount)
    expense.gst_rate = payload.gst_rate
    expense.total_amount = money(total_amount)
    expense.notes = payload.notes
    
    db.commit()
    return {"message": "Expense updated successfully"}


@api.delete('/expenses/{expense_id}')
def delete_expense(expense_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail='Expense not found')
    
    db.delete(expense)
    db.commit()
    return {"message": "Expense deleted successfully"}


# Cashflow Summary - Replaced by consolidated service


class PaymentIn(BaseModel):
    amount: float
    method: str
    account_head: str
    reference_number: str | None = None
    notes: str | None = None


class PurchasePaymentIn(BaseModel):
    amount: float
    method: str
    account_head: str
    reference_number: str | None = None
    notes: str | None = None





@api.get('/invoices/{invoice_id}/payments')
def list_payments(invoice_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    
    pays = db.query(Payment).filter(Payment.invoice_id == invoice_id).all()
    total_paid = float(sum([p.payment_amount for p in pays], 0))
    outstanding = float(inv.grand_total) - total_paid
    
    return {
        "payments": [{
            "id": p.id, 
            "payment_date": p.payment_date.isoformat(),
            "amount": float(p.payment_amount), 
            "method": p.payment_method, 
            "account_head": p.account_head,
            "reference_number": p.reference_number,
            "notes": p.notes
        } for p in pays], 
        "total_paid": total_paid, 
        "outstanding": outstanding
    }


@api.post('/purchases/{purchase_id}/payments', status_code=201)
def add_purchase_payment(purchase_id: int, payload: PurchasePaymentIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    # Validate payment amount
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be greater than 0")
    
    outstanding_amount = float(purchase.grand_total - purchase.paid_amount)
    if payload.amount > outstanding_amount:
        raise HTTPException(status_code=400, detail=f"Payment amount cannot exceed outstanding amount of ₹{outstanding_amount:.2f}")
    
    try:
        pay = PurchasePayment(
            purchase_id=purchase_id, 
            payment_amount=money(payload.amount), 
            payment_method=payload.method, 
            account_head=payload.account_head,
            reference_number=payload.reference_number,
            notes=payload.notes
        )
        db.add(pay)
        
        # Update purchase paid amount
        purchase.paid_amount += money(payload.amount)
        purchase.balance_amount = purchase.grand_total - purchase.paid_amount
        
        # Update purchase status
        if purchase.balance_amount == 0:
            purchase.status = "Paid"
        elif purchase.paid_amount > 0:
            purchase.status = "Partially Paid"
        
        # Cashflow transaction is now handled by the source table (PurchasePayment) directly
        
        db.commit()
        return {"id": pay.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to process payment: {str(e)}")


@api.get('/purchase-payments', response_model=list[PurchasePaymentOut])
def list_all_purchase_payments(
    search: str | None = None,
    payment_status: str | None = None,
    payment_method: str | None = None,
    vendor_id: int | None = None,
    amount_min: float | None = None,
    amount_max: float | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Get all purchase payments with filtering"""
    query = db.query(PurchasePayment).join(Purchase, PurchasePayment.purchase_id == Purchase.id).join(Party, Purchase.vendor_id == Party.id)
    
    if search:
        search_filter = (
            PurchasePayment.reference_number.ilike(f"%{search}%") |
            PurchasePayment.notes.ilike(f"%{search}%") |
            Party.name.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    if payment_status:
        if payment_status == 'paid':
            query = query.filter(PurchasePayment.payment_amount > 0)
        elif payment_status == 'pending':
            query = query.filter(PurchasePayment.payment_amount == 0)
    
    if payment_method:
        query = query.filter(PurchasePayment.payment_method == payment_method)
    
    if vendor_id:
        query = query.filter(Purchase.vendor_id == vendor_id)
    
    if amount_min is not None:
        query = query.filter(PurchasePayment.payment_amount >= amount_min)
    
    if amount_max is not None:
        query = query.filter(PurchasePayment.payment_amount <= amount_max)
    
    if date_from:
        query = query.filter(PurchasePayment.payment_date >= datetime.fromisoformat(date_from))
    
    if date_to:
        query = query.filter(PurchasePayment.payment_date <= datetime.fromisoformat(date_to))
    
    payments = query.order_by(PurchasePayment.payment_date.desc()).all()
    
    result = []
    for payment in payments:
        purchase = db.query(Purchase).filter(Purchase.id == payment.purchase_id).first()
        vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first() if purchase else None
        
        result.append(PurchasePaymentOut(
            id=payment.id,
            purchase_id=payment.purchase_id,
            payment_amount=float(payment.payment_amount),
            payment_method=payment.payment_method,
            account_head=payment.account_head,
            reference_number=payment.reference_number,
            payment_date=payment.payment_date.isoformat(),
            notes=payment.notes,
            vendor_name=vendor.name if vendor else "Unknown",
            purchase_number=purchase.purchase_no if purchase else "Unknown"
        ))
    
    return result

@api.get('/purchases/{purchase_id}/payments')
def list_purchase_payments(purchase_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    pays = db.query(PurchasePayment).filter(PurchasePayment.purchase_id == purchase_id).all()
    total_paid = float(sum([p.payment_amount for p in pays], 0))
    outstanding = float(purchase.grand_total) - total_paid
    
    return {
        "payments": [{
            "id": p.id, 
            "payment_date": p.payment_date.isoformat(),
            "amount": float(p.payment_amount), 
            "method": p.payment_method, 
            "account_head": p.account_head,
            "reference_number": p.reference_number,
            "notes": p.notes
        } for p in pays], 
        "total_paid": total_paid, 
        "outstanding": outstanding
    }


@api.patch('/purchases/{purchase_id}/status')
def update_purchase_status(purchase_id: int, status: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    valid_statuses = ["Draft", "Partially Paid", "Paid", "Cancelled"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f'Invalid status. Must be one of: {", ".join(valid_statuses)}')
    
    purchase.status = status
    db.commit()
    db.refresh(purchase)
    
    return {
        "id": purchase.id,
        "purchase_no": purchase.purchase_no,
        "status": purchase.status
    }


class StockAdjustmentIn(BaseModel):
    product_id: int
    quantity: int
    adjustment_type: str  # "add" or "reduce"
    date_of_adjustment: str  # ISO date string
    reference_bill_number: str | None = None
    supplier: str | None = None
    category: str | None = None
    notes: str | None = None


@api.post('/stock/adjust', status_code=201)
def stock_adjust(payload: StockAdjustmentIn, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        # Validation
        if payload.quantity < 0 or payload.quantity > 999999:
            raise HTTPException(status_code=400, detail="Quantity must be between 0 and 999999")
        
        if payload.adjustment_type not in ["add", "reduce", "adjust"]:
            raise HTTPException(status_code=400, detail="Adjustment type must be 'add', 'reduce', or 'adjust'")
        
        if payload.reference_bill_number and len(payload.reference_bill_number) > 10:
            raise HTTPException(status_code=400, detail="Reference bill number must be 10 characters or less")
        
        if payload.supplier and len(payload.supplier) > 50:
            raise HTTPException(status_code=400, detail="Supplier must be 50 characters or less")
        
        if payload.category and len(payload.category) > 50:
            raise HTTPException(status_code=400, detail="Category must be 50 characters or less")
        
        if payload.notes and len(payload.notes) > 200:
            raise HTTPException(status_code=400, detail="Notes must be 200 characters or less")
        
        prod = db.query(Product).filter(Product.id == payload.product_id).first()
        if not prod:
            raise HTTPException(status_code=404, detail='Product not found')
        
        # Calculate the delta based on adjustment type
        if payload.adjustment_type == "add":
            delta = payload.quantity
        elif payload.adjustment_type == "reduce":
            delta = -payload.quantity
            # Check if reducing stock would result in negative stock
            if (prod.stock - payload.quantity) < 0:
                raise HTTPException(status_code=400, detail="Cannot reduce stock below 0")
        elif payload.adjustment_type == "adjust":
            # For adjustments, quantity can be positive or negative
            delta = payload.quantity
            # Check if adjustment would result in negative stock
            if (prod.stock + payload.quantity) < 0:
                raise HTTPException(status_code=400, detail="Cannot adjust stock below 0")
        
        # Update product stock
        prod.stock += delta
        
        # Add stock ledger entry with correct entry type
        if payload.adjustment_type == "add":
            entry_type = 'in'
        elif payload.adjustment_type == "reduce":
            entry_type = 'out'
        else:  # adjust
            entry_type = 'adjust'
        
        db.add(StockLedgerEntry(
            product_id=prod.id, 
            qty=delta, 
            entry_type=entry_type, 
            ref_type=payload.adjustment_type, 
            ref_id=0
        ))
        
        db.commit()
        return {"ok": True, "new_stock": prod.stock}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to adjust stock")


class PartyOut(BaseModel):
    id: int
    type: str
    name: str
    contact_person: str | None
    contact_number: str | None
    email: str | None
    gstin: str | None
    gst_registration_status: str
    gst_enabled: bool  # New field for GST toggle
    billing_address_line1: str
    billing_address_line2: str | None
    billing_city: str
    billing_state: str
    billing_country: str
    billing_pincode: str | None
    shipping_address_line1: str | None
    shipping_address_line2: str | None
    shipping_city: str | None
    shipping_state: str | None
    shipping_country: str | None
    shipping_pincode: str | None
    notes: str | None
    is_active: bool

    class Config:
        from_attributes = True


class PartyCreate(BaseModel):
    type: str  # customer|vendor
    name: str
    contact_person: str | None = None
    contact_number: str | None = None
    email: str | None = None
    gstin: str | None = None
    gst_registration_status: str = "GST not registered"
    gst_enabled: bool = True  # New field for GST toggle
    billing_address_line1: str
    billing_address_line2: str | None = None
    billing_city: str
    billing_state: str
    billing_country: str = "India"
    billing_pincode: str | None = None
    shipping_address_line1: str | None = None
    shipping_address_line2: str | None = None
    shipping_city: str | None = None
    shipping_state: str | None = None
    shipping_country: str | None = None
    shipping_pincode: str | None = None
    notes: str | None = None
    
    @validator('gstin')
    def validate_gstin(cls, v, values):
        if v and values.get('gst_enabled', True):
            from .gst import validate_gstin
            if not validate_gstin(v):
                raise ValueError('Invalid GSTIN format')
        return v


class PartyUpdate(BaseModel):
    name: str | None = None
    contact_person: str | None = None
    contact_number: str | None = None
    email: str | None = None
    gstin: str | None = None
    gst_registration_status: str | None = None
    gst_enabled: bool | None = None  # New field for GST toggle
    billing_address_line1: str | None = None
    billing_address_line2: str | None = None
    billing_city: str | None = None
    billing_state: str | None = None
    billing_country: str | None = None
    billing_pincode: str | None = None
    shipping_address_line1: str | None = None
    shipping_address_line2: str | None = None
    shipping_city: str | None = None
    shipping_state: str | None = None
    shipping_country: str | None = None
    shipping_pincode: str | None = None
    notes: str | None = None
    
    @validator('gstin')
    def validate_gstin(cls, v, values):
        if v and values.get('gst_enabled', True):
            from .gst import validate_gstin
            if not validate_gstin(v):
                raise ValueError('Invalid GSTIN format')
        return v


@api.get('/parties', response_model=list[PartyOut])
def list_parties(
    type: str | None = None,
    search: str | None = None,
    include_inactive: bool = False,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    query = db.query(Party)
    
    if type:
        query = query.filter(Party.type == type)
    
    if search:
        search_filter = (
            Party.name.ilike(f"%{search}%") |
            Party.contact_person.ilike(f"%{search}%") |
            Party.contact_number.ilike(f"%{search}%") |
            Party.email.ilike(f"%{search}%") |
            Party.gstin.ilike(f"%{search}%") |
            Party.gst_registration_status.ilike(f"%{search}%") |
            Party.billing_address_line1.ilike(f"%{search}%") |
            Party.shipping_address_line1.ilike(f"%{search}%") |
            Party.notes.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Include inactive parties if requested
    if not include_inactive:
        query = query.filter(Party.is_active == True)
    
    return query.order_by(Party.name).all()


@api.get('/parties/customers', response_model=list[PartyOut])
def list_customers(
    search: str | None = None,
    include_inactive: bool = False,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    query = db.query(Party).filter(Party.type == "customer")
    
    if search:
        search_filter = (
            Party.name.ilike(f"%{search}%") |
            Party.contact_person.ilike(f"%{search}%") |
            Party.contact_number.ilike(f"%{search}%") |
            Party.email.ilike(f"%{search}%") |
            Party.gstin.ilike(f"%{search}%") |
            Party.gst_registration_status.ilike(f"%{search}%") |
            Party.billing_address_line1.ilike(f"%{search}%") |
            Party.shipping_address_line1.ilike(f"%{search}%") |
            Party.notes.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Include inactive customers if requested
    if not include_inactive:
        query = query.filter(Party.is_active == True)
    
    return query.order_by(Party.name).all()


@api.get('/parties/vendors', response_model=list[PartyOut])
def list_vendors(
    search: str | None = None,
    include_inactive: bool = False,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    query = db.query(Party).filter(Party.type == "vendor")
    
    if search:
        search_filter = (
            Party.name.ilike(f"%{search}%") |
            Party.contact_person.ilike(f"%{search}%") |
            Party.contact_number.ilike(f"%{search}%") |
            Party.email.ilike(f"%{search}%") |
            Party.gstin.ilike(f"%{search}%") |
            Party.gst_registration_status.ilike(f"%{search}%") |
            Party.billing_address_line1.ilike(f"%{search}%") |
            Party.shipping_address_line1.ilike(f"%{search}%") |
            Party.notes.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    # Include inactive vendors if requested
    if not include_inactive:
        query = query.filter(Party.is_active == True)
    
    return query.order_by(Party.name).all()


@api.post('/parties', response_model=PartyOut, status_code=status.HTTP_201_CREATED)
def create_party(payload: PartyCreate, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        # Validate GSTIN if GST is enabled
        if payload.gst_enabled and payload.gstin:
            from .gst import validate_gstin
            if not validate_gstin(payload.gstin):
                raise HTTPException(status_code=400, detail="Invalid GSTIN format")
        
        party = Party(**payload.model_dump())
        db.add(party)
        db.commit()
        db.refresh(party)
        return party
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Database error: {e}")
        raise HTTPException(status_code=500, detail="Failed to create party")


@api.put('/parties/{party_id}', response_model=PartyOut)
def update_party(party_id: int, payload: PartyUpdate, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    party = db.query(Party).filter(Party.id == party_id).first()
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    try:
        # Validate GSTIN if GST is enabled
        if payload.gst_enabled is not None and payload.gst_enabled and payload.gstin:
            from .gst import validate_gstin
            if not validate_gstin(payload.gstin):
                raise HTTPException(status_code=400, detail="Invalid GSTIN format")
        
        for field, value in payload.model_dump(exclude_unset=True).items():
            setattr(party, field, value)
        
        db.commit()
        db.refresh(party)
        return party
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Database error: {e}")
        raise HTTPException(status_code=500, detail="Failed to update party")


@api.patch('/parties/{party_id}/toggle', response_model=PartyOut)
def toggle_party(party_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    party = db.query(Party).filter(Party.id == party_id).first()
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    party.is_active = not party.is_active
    db.commit()
    db.refresh(party)
    return party


# Cashflow Management - Using consolidated service
@api.get('/cashflow/summary')
def get_cashflow_summary(
    start_date: str | None = None,
    end_date: str | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get cashflow summary for dashboard widgets"""
    service = ProfitPathService(db)
    
    start_dt = datetime.fromisoformat(start_date).date() if start_date else None
    end_dt = datetime.fromisoformat(end_date).date() if end_date else None
    
    return service.get_profitpath_summary(start_dt, end_dt)


@api.get('/cashflow/transactions')
def get_cashflow_transactions(
    search: str | None = None,
    type_filter: str | None = None,
    transaction_type: str | None = None,
    payment_method: str | None = None,
    account_head: str | None = None,
    amount_min: float | None = None,
    amount_max: float | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    page: int = 1,
    limit: int = 25,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get consolidated cashflow transactions from all source tables"""
    service = ProfitPathService(db)
    
    start_dt = datetime.fromisoformat(start_date).date() if start_date else None
    end_dt = datetime.fromisoformat(end_date).date() if end_date else None
    
    return service.get_cashflow_transactions(
        search=search,
        type_filter=type_filter,
        transaction_type=transaction_type,
        payment_method=payment_method,
        account_head=account_head,
        amount_min=amount_min,
        amount_max=amount_max,
        start_date=start_dt,
        end_date=end_dt,
        page=page,
        limit=limit
    )


@api.get('/cashflow/pending-payments')
def get_pending_payments(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get pending payments for invoices and purchases"""
    service = ProfitPathService(db)
    return service.get_pending_payments()


@api.get('/cashflow/financial-year/{financial_year}')
def get_financial_year_summary(
    financial_year: str,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get cashflow summary for a specific financial year (e.g., '2024-25')"""
    service = ProfitPathService(db)
    return service.get_financial_year_summary(financial_year)


@api.get('/cashflow/expenses/{financial_year}')
def get_expense_history_by_financial_year(
    financial_year: str,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get expense history for a specific financial year"""
    service = ProfitPathService(db)
    return service.get_expense_history_by_financial_year(financial_year)





@api.get('/purchases/{purchase_id}/pdf')
def purchase_pdf(purchase_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    # Get related data
    company = db.query(CompanySettings).first()
    vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first()
    items = db.query(PurchaseItem).filter(PurchaseItem.purchase_id == purchase.id).all()
    
    # Create PDF buffer
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=darkblue,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=darkblue,
        spaceAfter=6
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=3
    )
    
    # Build PDF content
    story = []
    
    # Header - Company Details
    if company:
        story.append(Paragraph(f"<b>{company.name}</b>", title_style))
        story.append(Paragraph(f"GSTIN: {company.gstin}", normal_style))
        story.append(Paragraph(f"State: {company.state} - {company.state_code}", normal_style))
    else:
        story.append(Paragraph("<b>ProfitPath</b>", title_style))
        story.append(Paragraph("Track Your Success, Step by Step", normal_style))
    
    story.append(Spacer(1, 20))
    
    # Purchase Header
    story.append(Paragraph(f"<b>PURCHASE ORDER</b>", heading_style))
    
    # Purchase Details Table
    purchase_data = [
        ['Purchase No:', purchase.purchase_no, 'Date:', purchase.date.strftime('%d/%m/%Y')],
        ['Due Date:', purchase.due_date.strftime('%d/%m/%Y'), 'Terms:', purchase.terms],
        ['Place of Supply:', purchase.place_of_supply, 'State Code:', purchase.place_of_supply_state_code]
    ]
    
    if purchase.eway_bill_number:
        purchase_data.append(['E-way Bill No:', purchase.eway_bill_number, '', ''])
    
    purchase_table = Table(purchase_data, colWidths=[2*cm, 6*cm, 2*cm, 6*cm])
    purchase_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(purchase_table)
    story.append(Spacer(1, 15))
    
    # Vendor Details
    if vendor:
        vendor_data = [
            ['Vendor:', vendor.name],
            ['', f"GSTIN: {vendor.gstin}" if vendor.gstin else "GSTIN: Not Available"],
            ['', vendor.address],
            ['', f"Email: {vendor.email}" if vendor.email else ""],
            ['', f"Phone: {vendor.phone}" if vendor.phone else ""]
        ]
        
        vendor_table = Table(vendor_data, colWidths=[2*cm, 14*cm])
        vendor_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        story.append(vendor_table)
        story.append(Spacer(1, 15))
    
    # Items Table
    story.append(Paragraph("<b>Item Details</b>", heading_style))
    
    # Table headers
    headers = ['S.No', 'Description', 'HSN', 'Qty', 'Rate', 'Amount', 'GST %', 'CGST', 'SGST', 'Total']
    table_data = [headers]
    
    # Add items
    for i, item in enumerate(items, 1):
        product = db.query(Product).filter(Product.id == item.product_id).first()
        description = product.name if product else item.description
        
        row = [
            str(i),
            description,
            item.hsn_code or '',
            str(item.qty),
            format_currency_for_pdf(float(item.rate), purchase.currency),
            format_currency_for_pdf(float(item.taxable_value), purchase.currency),
            f"{item.gst_rate}%",
            format_currency_for_pdf(float(item.cgst), purchase.currency),
            format_currency_for_pdf(float(item.sgst), purchase.currency),
            format_currency_for_pdf(float(item.amount), purchase.currency)
        ]
        table_data.append(row)
    
    # Create items table
    items_table = Table(table_data, colWidths=[0.8*cm, 4*cm, 1.5*cm, 1*cm, 1.5*cm, 1.5*cm, 1*cm, 1.2*cm, 1.2*cm, 1.5*cm])
    items_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Description left-aligned
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header row
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 15))
    
    # Totals Table
    totals_data = [
        ['Subtotal:', format_currency_for_pdf(float(purchase.taxable_value), purchase.currency)],
        ['CGST:', format_currency_for_pdf(float(purchase.cgst), purchase.currency)],
        ['SGST:', format_currency_for_pdf(float(purchase.sgst), purchase.currency)],
        ['IGST:', format_currency_for_pdf(float(purchase.igst), purchase.currency)],
        ['Total:', format_currency_for_pdf(float(purchase.grand_total), purchase.currency)]
    ]
    
    totals_table = Table(totals_data, colWidths=[4*cm, 2*cm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Total row bold
        ('FONTSIZE', (0, -1), (-1, -1), 12),  # Total row larger
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 20))
    
    # Notes
    if purchase.notes:
        story.append(Paragraph("<b>Notes:</b>", heading_style))
        story.append(Paragraph(purchase.notes, normal_style))
        story.append(Spacer(1, 15))
    
    # Footer
    story.append(Paragraph("Thank you for your service!", normal_style))
    story.append(Paragraph("This is a computer generated purchase order", normal_style))
    
    # Build PDF
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    
    return Response(content=pdf, media_type='application/pdf')


@api.post('/purchases/{purchase_id}/email', status_code=202)
def email_purchase(purchase_id: int, payload: EmailRequest, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    purchase = db.query(Purchase).filter(Purchase.id == purchase_id).first()
    if not purchase:
        raise HTTPException(status_code=404, detail='Purchase not found')
    
    # Get vendor details
    vendor = db.query(Party).filter(Party.id == purchase.vendor_id).first()
    vendor_name = vendor.name if vendor else "Valued Vendor"
    
    # Get company details
    company = db.query(CompanySettings).first()
    company_name = company.name if company else "ProfitPath"
    
    # Generate PDF
    try:
        # Create PDF buffer
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=darkblue,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=darkblue,
            spaceAfter=6
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=3
        )
        
        # Build PDF content (same as purchase_pdf function)
        story = []
        
        # Header - Company Details
        if company:
            story.append(Paragraph(f"<b>{company.name}</b>", title_style))
            story.append(Paragraph(f"GSTIN: {company.gstin}", normal_style))
            story.append(Paragraph(f"State: {company.state} - {company.state_code}", normal_style))
        else:
            story.append(Paragraph("<b>CASHFLOW</b>", title_style))
            story.append(Paragraph("Financial Management System", normal_style))
        
        story.append(Spacer(1, 20))
        
        # Purchase Header
        story.append(Paragraph(f"<b>PURCHASE ORDER</b>", heading_style))
        
        # Purchase Details Table
        purchase_data = [
            ['Purchase No:', purchase.purchase_no, 'Date:', purchase.date.strftime('%d/%m/%Y')],
            ['Due Date:', purchase.due_date.strftime('%d/%m/%Y'), 'Terms:', purchase.terms],
            ['Place of Supply:', purchase.place_of_supply, 'State Code:', purchase.place_of_supply_state_code]
        ]
        
        if purchase.eway_bill_number:
            purchase_data.append(['E-way Bill No:', purchase.eway_bill_number, '', ''])
        
        purchase_table = Table(purchase_data, colWidths=[2*cm, 6*cm, 2*cm, 6*cm])
        purchase_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(purchase_table)
        story.append(Spacer(1, 15))
        
        # Vendor Details
        if vendor:
            vendor_address = f"{vendor.billing_address_line1}"
            if vendor.billing_address_line2:
                vendor_address += f", {vendor.billing_address_line2}"
            vendor_address += f", {vendor.billing_city}, {vendor.billing_state} - {vendor.billing_pincode or ''}"
            
            vendor_data = [
                ['Vendor:', vendor.name],
                ['', f"GSTIN: {vendor.gstin}" if vendor.gstin else "GSTIN: Not Available"],
                ['', vendor_address],
                ['', f"Email: {vendor.email}" if vendor.email else ""],
                ['', f"Phone: {vendor.contact_number}" if vendor.contact_number else ""]
            ]
            
            vendor_table = Table(vendor_data, colWidths=[2*cm, 14*cm])
            vendor_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            story.append(vendor_table)
            story.append(Spacer(1, 15))
        
        # Items Table
        story.append(Paragraph("<b>Item Details</b>", heading_style))
        
        # Table headers
        headers = ['S.No', 'Description', 'HSN', 'Qty', 'Rate', 'Amount', 'GST %', 'CGST', 'SGST', 'Total']
        table_data = [headers]
        
        # Add items
        items = db.query(PurchaseItem).filter(PurchaseItem.purchase_id == purchase.id).all()
        for i, item in enumerate(items, 1):
            product = db.query(Product).filter(Product.id == item.product_id).first()
            description = product.name if product else item.description
            
            row = [
                str(i),
                description,
                item.hsn_code or '',
                str(item.qty),
                f"₹{float(item.rate):.2f}",
                f"₹{float(item.taxable_value):.2f}",
                f"{item.gst_rate}%",
                f"₹{float(item.cgst):.2f}",
                f"₹{float(item.sgst):.2f}",
                f"₹{float(item.amount):.2f}"
            ]
            table_data.append(row)
        
        # Create items table
        items_table = Table(table_data, colWidths=[0.8*cm, 4*cm, 1.5*cm, 1*cm, 1.5*cm, 1.5*cm, 1*cm, 1.2*cm, 1.2*cm, 1.5*cm])
        items_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Description left-aligned
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header row
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 15))
        
        # Totals Table
        totals_data = [
            ['Subtotal:', f"₹{float(purchase.taxable_value):.2f}"],
            ['CGST:', f"₹{float(purchase.cgst):.2f}"],
            ['SGST:', f"₹{float(purchase.sgst):.2f}"],
            ['IGST:', f"₹{float(purchase.igst):.2f}"],
            ['Total:', f"₹{float(purchase.grand_total):.2f}"]
        ]
        
        totals_table = Table(totals_data, colWidths=[4*cm, 2*cm])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Total row bold
            ('FONTSIZE', (0, -1), (-1, -1), 12),  # Total row larger
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 20))
        
        # Notes
        if purchase.notes:
            story.append(Paragraph("<b>Notes:</b>", heading_style))
            story.append(Paragraph(purchase.notes, normal_style))
            story.append(Spacer(1, 15))
        
        # Footer
        story.append(Paragraph("Thank you for your service!", normal_style))
        story.append(Paragraph("This is a computer generated purchase order", normal_style))
        
        # Build PDF
        doc.build(story)
        pdf_content = buf.getvalue()
        buf.close()
        
        # Create email template
        text_body, html_body = create_purchase_email_template(
            purchase_no=purchase.purchase_no,
            vendor_name=vendor_name,
            amount=float(purchase.grand_total),
            due_date=purchase.due_date.strftime('%d/%m/%Y'),
            company_name=company_name
        )
        
        # Send email with PDF attachment
        subject = f"Purchase Order {purchase.purchase_no} - {company_name}"
        filename = f"PurchaseOrder_{purchase.purchase_no}.pdf"
        
        success = send_email(
            to=payload.to,
            subject=subject,
            body=html_body,  # Use HTML version
            pdf_attachment=pdf_content,
            filename=filename
        )
        
        return {"status": "sent" if success else "failed"}
        
    except Exception as e:
        return {"status": "failed", "error": str(e)}


# Company Settings
@api.get('/company/settings')
def get_company_settings(_: User = Depends(get_current_user), db: Session = Depends(get_db)):
    settings = db.query(CompanySettings).first()
    if not settings:
        raise HTTPException(status_code=404, detail='Company settings not found')
    return settings

@api.put('/company/settings')
def update_company_settings(settings_data: dict, _: User = Depends(get_current_user), db: Session = Depends(get_db)):
    settings = db.query(CompanySettings).first()
    if not settings:
        raise HTTPException(status_code=404, detail='Company settings not found')
    
    for key, value in settings_data.items():
        if hasattr(settings, key):
            setattr(settings, key, value)
    
    db.commit()
    db.refresh(settings)
    return settings


# Stock Movement History Endpoints
class StockMovementOut(BaseModel):
    product_id: int
    product_name: str
    financial_year: str
    opening_stock: float
    incoming_stock: float
    outgoing_stock: float
    closing_stock: float

    class Config:
        from_attributes = True


class StockTransactionOut(BaseModel):
    id: int
    product_id: int
    product_name: str
    transaction_date: datetime
    entry_type: str  # 'in', 'out', 'adjust'
    quantity: float
    unit_price: float | None
    total_value: float | None
    ref_type: str | None  # 'invoice', 'purchase', 'adjustment', etc.
    ref_id: int | None
    reference_number: str | None
    notes: str | None
    financial_year: str
    running_balance: float

    class Config:
        from_attributes = True


class StockMovementSummaryOut(BaseModel):
    product_id: int
    product_name: str
    financial_year: str
    opening_stock: float
    opening_value: float
    total_incoming: float
    total_incoming_value: float
    total_outgoing: float
    total_outgoing_value: float
    closing_stock: float
    closing_value: float
    transactions: list[StockTransactionOut]

    class Config:
        from_attributes = True


# Advanced Invoice Features - Pydantic Models
class RecurringInvoiceTemplateCreate(BaseModel):
    name: str
    customer_id: int
    supplier_id: int
    recurrence_type: str  # weekly, monthly, yearly
    recurrence_interval: int = 1
    start_date: str
    end_date: str | None = None
    currency: str = "INR"
    exchange_rate: Decimal = Decimal('1.0')
    terms: str = "Due on Receipt"
    place_of_supply: str
    place_of_supply_state_code: str
    bill_to_address: str
    ship_to_address: str
    notes: str | None = None

    @validator('recurrence_type')
    def validate_recurrence_type(cls, v):
        if v not in ['weekly', 'monthly', 'yearly']:
            raise ValueError('Recurrence type must be weekly, monthly, or yearly')
        return v

    @validator('currency')
    def validate_currency(cls, v):
        supported_currencies = get_supported_currencies()
        if v.upper() not in supported_currencies:
            raise ValueError(f'Currency {v} is not supported')
        return v.upper()


class RecurringInvoiceTemplateItemCreate(BaseModel):
    product_id: int
    description: str
    hsn_code: str | None = None
    qty: float
    rate: Decimal
    discount: Decimal = Decimal('0')
    discount_type: str = "Percentage"
    gst_rate: float

    @validator('discount_type')
    def validate_discount_type(cls, v):
        if v not in ['Percentage', 'Fixed']:
            raise ValueError('Discount type must be Percentage or Fixed')
        return v


class RecurringInvoiceTemplateOut(BaseModel):
    id: int
    name: str
    customer_id: int
    supplier_id: int
    recurrence_type: str
    recurrence_interval: int
    start_date: datetime
    end_date: datetime | None
    next_generation_date: datetime
    currency: str
    exchange_rate: Decimal
    terms: str
    place_of_supply: str
    place_of_supply_state_code: str
    bill_to_address: str
    ship_to_address: str
    notes: str | None
    is_active: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class RecurringInvoiceTemplateItemOut(BaseModel):
    id: int
    template_id: int
    product_id: int
    description: str
    hsn_code: str | None
    qty: float
    rate: Decimal
    discount: Decimal
    discount_type: str
    gst_rate: float
    created_at: datetime

    class Config:
        from_attributes = True


class RecurringInvoiceOut(BaseModel):
    id: int
    template_id: int
    invoice_id: int
    generation_date: datetime
    due_date: datetime
    status: str
    created_at: datetime

    class Config:
        from_attributes = True


class CurrencyInfo(BaseModel):
    code: str
    symbol: str
    name: str


class ExchangeRateResponse(BaseModel):
    from_currency: str
    to_currency: str
    rate: float
    last_updated: datetime


# Purchase Order Management - Pydantic Models
class PurchaseOrderItemCreate(BaseModel):
    product_id: int
    description: str
    hsn_code: str | None = None
    qty: float
    expected_rate: Decimal
    discount: Decimal = Decimal('0')
    discount_type: str = "Percentage"
    gst_rate: float

    @validator('discount_type')
    def validate_discount_type(cls, v):
        if v not in ['Percentage', 'Fixed']:
            raise ValueError('Discount type must be Percentage or Fixed')
        return v


class PurchaseOrderCreate(BaseModel):
    vendor_id: int
    po_number: str | None = None
    date: str
    expected_delivery_date: str
    currency: str = "INR"
    exchange_rate: Decimal = Decimal('1.0')
    terms: str = "Net 30"
    place_of_supply: str
    place_of_supply_state_code: str
    reverse_charge: bool = False
    ship_from_address: str
    ship_to_address: str
    notes: str | None = None
    items: list[PurchaseOrderItemCreate]

    @validator('currency')
    def validate_currency(cls, v):
        supported_currencies = get_supported_currencies()
        if v.upper() not in supported_currencies:
            raise ValueError(f'Currency {v} is not supported')
        return v.upper()


class PurchaseOrderItemOut(BaseModel):
    id: int
    purchase_order_id: int
    product_id: int
    description: str
    hsn_code: str | None
    qty: float
    expected_rate: Decimal
    discount: Decimal
    discount_type: str
    gst_rate: float
    amount: Decimal
    created_at: datetime

    class Config:
        from_attributes = True


class PurchaseOrderOut(BaseModel):
    id: int
    vendor_id: int
    po_number: str
    date: datetime
    expected_delivery_date: datetime
    status: str
    currency: str
    exchange_rate: Decimal
    terms: str
    place_of_supply: str
    place_of_supply_state_code: str
    reverse_charge: bool
    ship_from_address: str
    ship_to_address: str
    subtotal: Decimal
    total_discount: Decimal
    cgst: Decimal
    sgst: Decimal
    igst: Decimal
    utgst: Decimal
    cess: Decimal
    round_off: Decimal
    grand_total: Decimal
    approved_by: int | None
    approved_at: datetime | None
    sent_at: datetime | None
    received_at: datetime | None
    closed_at: datetime | None
    notes: str | None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


@api.get('/stock/history', response_model=list[StockLedgerEntryOut])
def get_stock_history(
    search: str | None = None,
    product_id: int | None = None,
    entry_type: str | None = None,
    reference_number: str | None = None,
    quantity_min: float | None = None,
    quantity_max: float | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    _: User = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Get stock history with filtering"""
    query = db.query(StockLedgerEntry).join(Product, StockLedgerEntry.product_id == Product.id)
    
    if search:
        search_filter = (
            StockLedgerEntry.reference_bill_number.ilike(f"%{search}%") |
            StockLedgerEntry.notes.ilike(f"%{search}%") |
            Product.name.ilike(f"%{search}%")
        )
        query = query.filter(search_filter)
    
    if product_id:
        query = query.filter(StockLedgerEntry.product_id == product_id)
    
    if entry_type:
        query = query.filter(StockLedgerEntry.entry_type == entry_type)
    
    if reference_number:
        query = query.filter(StockLedgerEntry.reference_bill_number.ilike(f"%{reference_number}%"))
    
    if quantity_min is not None:
        query = query.filter(StockLedgerEntry.qty >= quantity_min)
    
    if quantity_max is not None:
        query = query.filter(StockLedgerEntry.qty <= quantity_max)
    
    if date_from:
        query = query.filter(StockLedgerEntry.created_at >= datetime.fromisoformat(date_from))
    
    if date_to:
        query = query.filter(StockLedgerEntry.created_at <= datetime.fromisoformat(date_to))
    
    entries = query.order_by(StockLedgerEntry.created_at.desc()).all()
    
    result = []
    for entry in entries:
        product = db.query(Product).filter(Product.id == entry.product_id).first()
        result.append(StockLedgerEntryOut(
            id=entry.id,
            product_id=entry.product_id,
            entry_type=entry.entry_type,
            qty=float(entry.qty),
            reference_bill_number=entry.reference_bill_number,
            notes=entry.notes,
            created_at=entry.created_at.isoformat(),
            product_name=product.name if product else "Unknown"
        ))
    
    return result

@api.get('/stock/movement-history', response_model=list[StockMovementSummaryOut])
def get_stock_movement_history(
    financial_year: str | None = None,
    product_id: int | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get stock movement history with individual transactions for all products or specific product by financial year"""
    
    # If no financial year specified, use current year
    if not financial_year:
        current_year = datetime.now().year
        financial_year = f"{current_year}-{current_year + 1}"
    
    # Parse financial year (format: "2024-2025")
    try:
        start_year = int(financial_year.split('-')[0])
        end_year = int(financial_year.split('-')[1])
        fy_start_date = datetime(start_year, 4, 1)  # April 1st
        fy_end_date = datetime(end_year, 3, 31)     # March 31st
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid financial year format. Use YYYY-YYYY")
    
    # Get products (all or specific)
    if product_id:
        products = db.query(Product).filter(Product.id == product_id).all()
        if not products:
            raise HTTPException(status_code=404, detail="Product not found")
    else:
        products = db.query(Product).all()
    
    movements = []
    
    for product in products:
        # Get all stock transactions for this product in the financial year
        stock_transactions = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id,
            StockLedgerEntry.created_at >= fy_start_date,
            StockLedgerEntry.created_at <= fy_end_date
        ).order_by(StockLedgerEntry.created_at).all()
        
        # Calculate opening stock (stock at the beginning of FY)
        # Get all transactions before the financial year
        pre_fy_transactions = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id,
            StockLedgerEntry.created_at < fy_start_date
        ).all()
        
        opening_stock = sum(
            adj.qty for adj in pre_fy_transactions 
            if adj.entry_type == 'in'
        ) - sum(
            adj.qty for adj in pre_fy_transactions 
            if adj.entry_type == 'out'
        )
        opening_stock = max(0, opening_stock)
        
        # Calculate opening value (using average cost method)
        unit_price = product.purchase_price or product.sales_price or 0
        unit_price_float = float(unit_price) if unit_price else 0.0
        opening_value = opening_stock * unit_price_float
        
        # Process individual transactions
        transactions = []
        running_balance = opening_stock
        
        for transaction in stock_transactions:
            # Calculate unit price and total value
            unit_price = product.purchase_price or product.sales_price or 0
            # Convert Decimal to float for calculation
            unit_price_float = float(unit_price) if unit_price else 0.0
            total_value = transaction.qty * unit_price_float
            
            # Determine reference information
            ref_type = None
            ref_id = None
            reference_number = None
            notes = None
            
            if transaction.ref_type and transaction.ref_id:
                ref_type = transaction.ref_type
                ref_id = transaction.ref_id
                # Try to get reference number based on ref_type
                if ref_type == 'invoice':
                    invoice = db.query(Invoice).filter(Invoice.id == ref_id).first()
                    reference_number = invoice.invoice_no if invoice else None
                elif ref_type == 'purchase':
                    purchase = db.query(Purchase).filter(Purchase.id == ref_id).first()
                    reference_number = purchase.purchase_no if purchase else None
                elif ref_type == 'adjustment':
                    reference_number = f"ADJ-{ref_id}"
            
            # Update running balance
            if transaction.entry_type == 'in':
                running_balance += transaction.qty
            elif transaction.entry_type == 'out':
                running_balance -= transaction.qty
            elif transaction.entry_type == 'adjust':
                # For adjustments: positive quantity increases stock, negative decreases
                running_balance += transaction.qty
            
            transactions.append(StockTransactionOut(
                id=transaction.id,
                product_id=product.id,
                product_name=product.name,
                transaction_date=transaction.created_at,
                entry_type=transaction.entry_type,
                quantity=transaction.qty,
                unit_price=unit_price,
                total_value=total_value,
                ref_type=ref_type,
                ref_id=ref_id,
                reference_number=reference_number,
                notes=notes,
                financial_year=financial_year,
                running_balance=running_balance
            ))
        
        # Calculate summary totals
        # For adjustments: positive quantity = incoming, negative quantity = outgoing
        total_incoming = sum(t.quantity for t in transactions if t.entry_type == 'in') + \
                        sum(t.quantity for t in transactions if t.entry_type == 'adjust' and t.quantity > 0)
        total_incoming_value = sum(t.total_value or 0 for t in transactions if t.entry_type == 'in') + \
                              sum(t.total_value or 0 for t in transactions if t.entry_type == 'adjust' and t.quantity > 0)
        total_outgoing = sum(t.quantity for t in transactions if t.entry_type == 'out') + \
                        sum(abs(t.quantity) for t in transactions if t.entry_type == 'adjust' and t.quantity < 0)
        total_outgoing_value = sum(t.total_value or 0 for t in transactions if t.entry_type == 'out') + \
                              sum(t.total_value or 0 for t in transactions if t.entry_type == 'adjust' and t.quantity < 0)
        closing_stock = running_balance
        closing_value = closing_stock * unit_price_float
        
        movements.append(StockMovementSummaryOut(
            product_id=product.id,
            product_name=product.name,
            financial_year=financial_year,
            opening_stock=opening_stock,
            opening_value=opening_value,
            total_incoming=total_incoming,
            total_incoming_value=total_incoming_value,
            total_outgoing=total_outgoing,
            total_outgoing_value=total_outgoing_value,
            closing_stock=closing_stock,
            closing_value=closing_value,
            transactions=transactions
        ))
    
    return movements


@api.get('/stock/movement-history/pdf')
def get_stock_movement_history_pdf(
    financial_year: str | None = None,
    product_id: int | None = None,
    product_filter: str | None = None,
    entry_type_filter: str | None = None,
    reference_type_filter: str | None = None,
    reference_search: str | None = None,
    stock_level_filter: str | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate PDF report for stock movement history with filters applied"""
    
    # If no financial year specified, use current year
    if not financial_year:
        current_year = datetime.now().year
        financial_year = f"{current_year}-{current_year + 1}"
    
    # Parse financial year (format: "2024-2025")
    try:
        start_year = int(financial_year.split('-')[0])
        end_year = int(financial_year.split('-')[1])
        fy_start_date = datetime(start_year, 4, 1)  # April 1st
        fy_end_date = datetime(end_year, 3, 31)     # March 31st
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid financial year format. Use YYYY-YYYY")
    
    # Get products (all or specific)
    if product_id:
        products = db.query(Product).filter(Product.id == product_id).all()
        if not products:
            raise HTTPException(status_code=404, detail="Product not found")
    elif product_filter and product_filter != 'all':
        # Filter by product name
        products = db.query(Product).filter(Product.name.ilike(f"%{product_filter}%")).all()
    else:
        products = db.query(Product).all()
    
    # Create PDF buffer
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=darkblue,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=darkblue,
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=3
    )
    
    # Build PDF content
    story = []
    
    # Header
    story.append(Paragraph("Stock Movement History Report", title_style))
    story.append(Paragraph(f"Financial Year: {financial_year}", normal_style))
    if product_id:
        product = products[0] if products else None
        if product:
            story.append(Paragraph(f"Product: {product.name}", normal_style))
    else:
        story.append(Paragraph(f"All Products ({len(products)} items)", normal_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 20))
    
    # Process each product with enhanced sections
    for i, product in enumerate(products, 1):
        # Add page break for each product (except the first one)
        if i > 1:
            story.append(PageBreak())
        
        # Enhanced Product Section Header
        story.append(Paragraph(f"PRODUCT {i}: {product.name.upper()}", heading_style))
        story.append(Spacer(1, 5))
        
        # Product Details Box
        product_details = [
            ['Product Name:', product.name],
            ['SKU:', product.sku or 'N/A'],
            ['Category:', product.category or 'N/A'],
            ['Unit:', product.unit or 'N/A'],
            ['Purchase Price:', format_currency_for_pdf(float(product.purchase_price or 0), 'INR')],
            ['Sales Price:', format_currency_for_pdf(float(product.sales_price or 0), 'INR')]
        ]
        
        product_table = Table(product_details, colWidths=[3*cm, 8*cm])
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(product_table)
        story.append(Spacer(1, 15))
        
        # Get all stock transactions for this product in the financial year
        stock_transactions = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id,
            StockLedgerEntry.created_at >= fy_start_date,
            StockLedgerEntry.created_at <= fy_end_date
        ).order_by(StockLedgerEntry.created_at).all()
        
        # Apply filters to transactions
        filtered_transactions = []
        for transaction in stock_transactions:
            # Entry type filter
            if entry_type_filter and entry_type_filter != 'all':
                if entry_type_filter == 'incoming' and transaction.entry_type not in ['in', 'adjust']:
                    continue
                elif entry_type_filter == 'outgoing' and transaction.entry_type not in ['out', 'adjust']:
                    continue
                elif entry_type_filter == 'adjustment' and transaction.entry_type != 'adjust':
                    continue
            
            # Reference type filter
            if reference_type_filter and reference_type_filter != 'all':
                if reference_type_filter != transaction.ref_type:
                    continue
            
            # Reference search filter
            if reference_search:
                ref_type = transaction.ref_type
                ref_id = transaction.ref_id
                reference_number = None
                
                if ref_type and ref_id:
                    if ref_type == 'invoice':
                        invoice = db.query(Invoice).filter(Invoice.id == ref_id).first()
                        reference_number = invoice.invoice_no if invoice else None
                    elif ref_type == 'purchase':
                        purchase = db.query(Purchase).filter(Purchase.id == ref_id).first()
                        reference_number = purchase.purchase_no if purchase else None
                    elif ref_type == 'adjustment':
                        reference_number = f"ADJ-{ref_id}"
                
                if not reference_number or reference_search.lower() not in reference_number.lower():
                    continue
            
            filtered_transactions.append(transaction)
        
        # Calculate opening stock
        pre_fy_transactions = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id,
            StockLedgerEntry.created_at < fy_start_date
        ).all()
        
        opening_stock = sum(
            adj.qty for adj in pre_fy_transactions 
            if adj.entry_type == 'in'
        ) - sum(
            adj.qty for adj in pre_fy_transactions 
            if adj.entry_type == 'out'
        )
        opening_stock = max(0, opening_stock)
        
        # Calculate opening value
        unit_price = product.purchase_price or product.sales_price or 0
        unit_price_float = float(unit_price) if unit_price else 0.0
        opening_value = opening_stock * unit_price_float
        
        # Process filtered transactions
        transactions = []
        running_balance = opening_stock
        
        for transaction in filtered_transactions:
            # Calculate unit price and total value
            unit_price = product.purchase_price or product.sales_price or 0
            unit_price_float = float(unit_price) if unit_price else 0.0
            total_value = transaction.qty * unit_price_float
            
            # Determine reference information
            ref_type = None
            ref_id = None
            reference_number = None
            
            if transaction.ref_type and transaction.ref_id:
                ref_type = transaction.ref_type
                ref_id = transaction.ref_id
                if ref_type == 'invoice':
                    invoice = db.query(Invoice).filter(Invoice.id == ref_id).first()
                    reference_number = invoice.invoice_no if invoice else None
                elif ref_type == 'purchase':
                    purchase = db.query(Purchase).filter(Purchase.id == ref_id).first()
                    reference_number = purchase.purchase_no if purchase else None
                elif ref_type == 'adjustment':
                    reference_number = f"ADJ-{ref_id}"
            
            # Update running balance
            if transaction.entry_type == 'in':
                running_balance += transaction.qty
            elif transaction.entry_type == 'out':
                running_balance -= transaction.qty
            elif transaction.entry_type == 'adjust':
                running_balance += transaction.qty
            
            transactions.append({
                'date': transaction.created_at.strftime('%d/%m/%Y'),
                'type': transaction.entry_type.upper(),
                'quantity': f"{transaction.qty:.2f}",
                'unit_price': f"₹{unit_price_float:.2f}" if unit_price_float > 0 else '-',
                'total_value': f"₹{total_value:.2f}" if total_value > 0 else '-',
                'running_balance': f"{running_balance:.2f}",
                'reference': reference_number or '-'
            })
        
        # Calculate summary totals from filtered transactions
        total_incoming = sum(t.qty for t in filtered_transactions if t.entry_type == 'in') + \
                        sum(t.qty for t in filtered_transactions if t.entry_type == 'adjust' and t.qty > 0)
        total_incoming_value = sum(t.qty * unit_price_float for t in filtered_transactions if t.entry_type == 'in') + \
                              sum(t.qty * unit_price_float for t in filtered_transactions if t.entry_type == 'adjust' and t.qty > 0)
        total_outgoing = sum(t.qty for t in filtered_transactions if t.entry_type == 'out') + \
                        sum(abs(t.qty) for t in filtered_transactions if t.entry_type == 'adjust' and t.qty < 0)
        total_outgoing_value = sum(t.qty * unit_price_float for t in filtered_transactions if t.entry_type == 'out') + \
                              sum(abs(t.qty) * unit_price_float for t in filtered_transactions if t.entry_type == 'adjust' and t.qty < 0)
        closing_stock = running_balance
        closing_value = closing_stock * unit_price_float
        
        # Apply stock level filter
        if stock_level_filter and stock_level_filter != 'all':
            if stock_level_filter == 'out_of_stock' and closing_stock > 0:
                continue
            elif stock_level_filter == 'in_stock' and closing_stock <= 0:
                continue
            elif stock_level_filter == 'low_stock' and (closing_stock <= 0 or closing_stock >= 10):
                continue
        
        # Enhanced Summary Section
        story.append(Paragraph("STOCK SUMMARY", heading_style))
        story.append(Spacer(1, 5))
        
        summary_data = [
            ['Opening Stock', f"{opening_stock:.2f}", f"₹{opening_value:.2f}"],
            ['Total Incoming', f"{total_incoming:.2f}", f"₹{total_incoming_value:.2f}"],
            ['Total Outgoing', f"{total_outgoing:.2f}", f"₹{total_outgoing_value:.2f}"],
            ['Closing Stock', f"{closing_stock:.2f}", f"₹{closing_value:.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*cm, 3*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        # Enhanced Transactions Section
        if transactions:
            story.append(Paragraph("TRANSACTION DETAILS", heading_style))
            story.append(Spacer(1, 5))
            
            # Add opening balance row
            transactions.insert(0, {
                'date': f"{start_year}-04-01",
                'type': 'OPENING',
                'quantity': f"{opening_stock:.2f}",
                'unit_price': f"₹{unit_price_float:.2f}" if unit_price_float > 0 else '-',
                'total_value': f"₹{opening_value:.2f}" if opening_value > 0 else '-',
                'running_balance': f"{opening_stock:.2f}",
                'reference': '-'
            })
            
            # Create table data
            table_data = [['Date', 'Type', 'Quantity', 'Unit Price', 'Total Value', 'Running Balance', 'Reference']]
            for t in transactions:
                table_data.append([
                    t['date'], t['type'], t['quantity'], t['unit_price'], 
                    t['total_value'], t['running_balance'], t['reference']
                ])
            
            # Create table with enhanced styling
            transaction_table = Table(table_data, colWidths=[2*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm])
            transaction_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            story.append(transaction_table)
        else:
            story.append(Paragraph("No transactions found for this period.", normal_style))
        
        story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    
    # Generate filename
    filename = f"stock_movement_history_{financial_year}"
    if product_id:
        product = products[0] if products else None
        if product:
            filename += f"_{product.name.replace(' ', '_')}"
    filename += ".pdf"
    
    return Response(
        content=pdf, 
        media_type='application/pdf',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api.get('/stock/movement-history/pdf-preview')
def get_stock_movement_history_pdf_preview(
    financial_year: str | None = None,
    product_id: int | None = None,
    product_filter: str | None = None,
    entry_type_filter: str | None = None,
    reference_type_filter: str | None = None,
    reference_search: str | None = None,
    stock_level_filter: str | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate PDF preview for stock movement history (inline display)"""
    
    # If no financial year specified, use current year
    if not financial_year:
        current_year = datetime.now().year
        financial_year = f"{current_year}-{current_year + 1}"
    
    # Parse financial year (format: "2024-2025")
    try:
        start_year = int(financial_year.split('-')[0])
        end_year = int(financial_year.split('-')[1])
        fy_start_date = datetime(start_year, 4, 1)  # April 1st
        fy_end_date = datetime(end_year, 3, 31)     # March 31st
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid financial year format. Use YYYY-YYYY")
    
    # Get products (all or specific)
    if product_id:
        products = db.query(Product).filter(Product.id == product_id).all()
        if not products:
            raise HTTPException(status_code=404, detail="Product not found")
    elif product_filter and product_filter != 'all':
        # Filter by product name
        products = db.query(Product).filter(Product.name.ilike(f"%{product_filter}%")).all()
    else:
        products = db.query(Product).all()
    
    # Create PDF buffer
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=darkblue,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=darkblue,
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=3
    )
    
    # Build PDF content
    story = []
    
    # Header
    story.append(Paragraph("Stock Movement History Report", title_style))
    story.append(Paragraph(f"Financial Year: {financial_year}", normal_style))
    if product_id:
        product = products[0] if products else None
        if product:
            story.append(Paragraph(f"Product: {product.name}", normal_style))
    else:
        story.append(Paragraph(f"All Products ({len(products)} items)", normal_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
    
    # Add filter information
    active_filters = []
    if product_filter and product_filter != 'all':
        active_filters.append(f"Product: {product_filter}")
    if entry_type_filter and entry_type_filter != 'all':
        active_filters.append(f"Entry Type: {entry_type_filter}")
    if reference_type_filter and reference_type_filter != 'all':
        active_filters.append(f"Reference Type: {reference_type_filter}")
    if reference_search:
        active_filters.append(f"Reference Search: {reference_search}")
    if stock_level_filter and stock_level_filter != 'all':
        active_filters.append(f"Stock Level: {stock_level_filter}")
    
    if active_filters:
        story.append(Paragraph("Active Filters:", normal_style))
        for filter_info in active_filters:
            story.append(Paragraph(f"• {filter_info}", normal_style))
    
    story.append(Spacer(1, 20))
    
    # Process each product with enhanced sections (same logic as download endpoint)
    for i, product in enumerate(products, 1):
        # Add page break for each product (except the first one)
        if i > 1:
            story.append(PageBreak())
        
        # Enhanced Product Section Header
        story.append(Paragraph(f"PRODUCT {i}: {product.name.upper()}", heading_style))
        story.append(Spacer(1, 5))
        
        # Product Details Box
        product_details = [
            ['Product Name:', product.name],
            ['SKU:', product.sku or 'N/A'],
            ['Category:', product.category or 'N/A'],
            ['Unit:', product.unit or 'N/A'],
            ['Purchase Price:', format_currency_for_pdf(float(product.purchase_price or 0), 'INR')],
            ['Sales Price:', format_currency_for_pdf(float(product.sales_price or 0), 'INR')]
        ]
        
        product_table = Table(product_details, colWidths=[3*cm, 8*cm])
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(product_table)
        story.append(Spacer(1, 15))
        
        # Get all stock transactions for this product in the financial year
        stock_transactions = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id,
            StockLedgerEntry.created_at >= fy_start_date,
            StockLedgerEntry.created_at <= fy_end_date
        ).order_by(StockLedgerEntry.created_at).all()
        
        # Apply filters to transactions (same logic as download endpoint)
        filtered_transactions = []
        for transaction in stock_transactions:
            # Entry type filter
            if entry_type_filter and entry_type_filter != 'all':
                if entry_type_filter == 'incoming' and transaction.entry_type not in ['in', 'adjust']:
                    continue
                elif entry_type_filter == 'outgoing' and transaction.entry_type not in ['out', 'adjust']:
                    continue
                elif entry_type_filter == 'adjustment' and transaction.entry_type != 'adjust':
                    continue
            
            # Reference type filter
            if reference_type_filter and reference_type_filter != 'all':
                if reference_type_filter != transaction.ref_type:
                    continue
            
            # Reference search filter
            if reference_search:
                ref_type = transaction.ref_type
                ref_id = transaction.ref_id
                reference_number = None
                
                if ref_type and ref_id:
                    if ref_type == 'invoice':
                        invoice = db.query(Invoice).filter(Invoice.id == ref_id).first()
                        reference_number = invoice.invoice_no if invoice else None
                    elif ref_type == 'purchase':
                        purchase = db.query(Purchase).filter(Purchase.id == ref_id).first()
                        reference_number = purchase.purchase_no if purchase else None
                    elif ref_type == 'adjustment':
                        reference_number = f"ADJ-{ref_id}"
                
                if not reference_number or reference_search.lower() not in reference_number.lower():
                    continue
            
            filtered_transactions.append(transaction)
        
        # Calculate opening stock
        pre_fy_transactions = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product.id,
            StockLedgerEntry.created_at < fy_start_date
        ).all()
        
        opening_stock = sum(
            adj.qty for adj in pre_fy_transactions 
            if adj.entry_type == 'in'
        ) - sum(
            adj.qty for adj in pre_fy_transactions 
            if adj.entry_type == 'out'
        )
        opening_stock = max(0, opening_stock)
        
        # Calculate opening value
        unit_price = product.purchase_price or product.sales_price or 0
        unit_price_float = float(unit_price) if unit_price else 0.0
        opening_value = opening_stock * unit_price_float
        
        # Process filtered transactions
        transactions = []
        running_balance = opening_stock
        
        for transaction in filtered_transactions:
            # Calculate unit price and total value
            unit_price = product.purchase_price or product.sales_price or 0
            unit_price_float = float(unit_price) if unit_price else 0.0
            total_value = transaction.qty * unit_price_float
            
            # Determine reference information
            ref_type = None
            ref_id = None
            reference_number = None
            
            if transaction.ref_type and transaction.ref_id:
                ref_type = transaction.ref_type
                ref_id = transaction.ref_id
                if ref_type == 'invoice':
                    invoice = db.query(Invoice).filter(Invoice.id == ref_id).first()
                    reference_number = invoice.invoice_no if invoice else None
                elif ref_type == 'purchase':
                    purchase = db.query(Purchase).filter(Purchase.id == ref_id).first()
                    reference_number = purchase.purchase_no if purchase else None
                elif ref_type == 'adjustment':
                    reference_number = f"ADJ-{ref_id}"
            
            # Update running balance
            if transaction.entry_type == 'in':
                running_balance += transaction.qty
            elif transaction.entry_type == 'out':
                running_balance -= transaction.qty
            elif transaction.entry_type == 'adjust':
                running_balance += transaction.qty
            
            transactions.append({
                'date': transaction.created_at.strftime('%d/%m/%Y'),
                'type': transaction.entry_type.upper(),
                'quantity': f"{transaction.qty:.2f}",
                'unit_price': f"₹{unit_price_float:.2f}" if unit_price_float > 0 else '-',
                'total_value': f"₹{total_value:.2f}" if total_value > 0 else '-',
                'running_balance': f"{running_balance:.2f}",
                'reference': reference_number or '-'
            })
        
        # Calculate summary totals from filtered transactions
        total_incoming = sum(t.qty for t in filtered_transactions if t.entry_type == 'in') + \
                        sum(t.qty for t in filtered_transactions if t.entry_type == 'adjust' and t.qty > 0)
        total_incoming_value = sum(t.qty * unit_price_float for t in filtered_transactions if t.entry_type == 'in') + \
                              sum(t.qty * unit_price_float for t in filtered_transactions if t.entry_type == 'adjust' and t.qty > 0)
        total_outgoing = sum(t.qty for t in filtered_transactions if t.entry_type == 'out') + \
                        sum(abs(t.qty) for t in filtered_transactions if t.entry_type == 'adjust' and t.qty < 0)
        total_outgoing_value = sum(t.qty * unit_price_float for t in filtered_transactions if t.entry_type == 'out') + \
                              sum(abs(t.qty) * unit_price_float for t in filtered_transactions if t.entry_type == 'adjust' and t.qty < 0)
        closing_stock = running_balance
        closing_value = closing_stock * unit_price_float
        
        # Apply stock level filter
        if stock_level_filter and stock_level_filter != 'all':
            if stock_level_filter == 'out_of_stock' and closing_stock > 0:
                continue
            elif stock_level_filter == 'in_stock' and closing_stock <= 0:
                continue
            elif stock_level_filter == 'low_stock' and (closing_stock <= 0 or closing_stock >= 10):
                continue
        
        # Enhanced Summary Section
        story.append(Paragraph("STOCK SUMMARY", heading_style))
        story.append(Spacer(1, 5))
        
        summary_data = [
            ['Opening Stock', f"{opening_stock:.2f}", f"₹{opening_value:.2f}"],
            ['Total Incoming', f"{total_incoming:.2f}", f"₹{total_incoming_value:.2f}"],
            ['Total Outgoing', f"{total_outgoing:.2f}", f"₹{total_outgoing_value:.2f}"],
            ['Closing Stock', f"{closing_stock:.2f}", f"₹{closing_value:.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*cm, 3*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        # Enhanced Transactions Section
        if transactions:
            story.append(Paragraph("TRANSACTION DETAILS", heading_style))
            story.append(Spacer(1, 5))
            
            # Add opening balance row
            transactions.insert(0, {
                'date': f"{start_year}-04-01",
                'type': 'OPENING',
                'quantity': f"{opening_stock:.2f}",
                'unit_price': f"₹{unit_price_float:.2f}" if unit_price_float > 0 else '-',
                'total_value': f"₹{opening_value:.2f}" if opening_value > 0 else '-',
                'running_balance': f"{opening_stock:.2f}",
                'reference': '-'
            })
            
            # Create table data
            table_data = [['Date', 'Type', 'Quantity', 'Unit Price', 'Total Value', 'Running Balance', 'Reference']]
            for t in transactions:
                table_data.append([
                    t['date'], t['type'], t['quantity'], t['unit_price'], 
                    t['total_value'], t['running_balance'], t['reference']
                ])
            
            # Create table with enhanced styling
            transaction_table = Table(table_data, colWidths=[2*cm, 1.5*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm])
            transaction_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            story.append(transaction_table)
        else:
            story.append(Paragraph("No transactions found for this period.", normal_style))
        
        story.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    
    return Response(
        content=pdf, 
        media_type='application/pdf',
        headers={"Content-Disposition": "inline; filename=stock_movement_history_preview.pdf"}
    )


@api.get('/stock/movement-history/{product_id}', response_model=list[StockMovementOut])
def get_product_stock_movement_history(
    product_id: int,
    from_year: int | None = None,
    to_year: int | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get stock movement history for a specific product across multiple financial years"""
    
    # Check if product exists
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Set default year range if not provided
    if not from_year:
        from_year = datetime.now().year - 2  # Last 3 years
    if not to_year:
        to_year = datetime.now().year
    
    movements = []
    
    for year in range(from_year, to_year + 1):
        financial_year = f"{year}-{year + 1}"
        fy_start_date = datetime(year, 4, 1)  # April 1st
        fy_end_date = datetime(year + 1, 3, 31)  # March 31st
        
        # Get stock adjustments for this product in the financial year
        stock_adjustments = db.query(StockLedgerEntry).filter(
            StockLedgerEntry.product_id == product_id,
            StockLedgerEntry.created_at >= fy_start_date,
            StockLedgerEntry.created_at <= fy_end_date
        ).all()
        
        # Calculate incoming and outgoing stock
        incoming_stock = sum(
            adj.qty for adj in stock_adjustments 
            if adj.entry_type == 'in'
        )
        outgoing_stock = sum(
            adj.qty for adj in stock_adjustments 
            if adj.entry_type == 'out'
        )
        
        # For historical years, we need to calculate opening stock differently
        # This is a simplified calculation - in a real system, you'd need to track opening balances
        opening_stock = 0  # Placeholder - would need historical opening balance tracking
        closing_stock = opening_stock + incoming_stock - outgoing_stock
        
        movements.append(StockMovementOut(
            product_id=product.id,
            product_name=product.name,
            financial_year=financial_year,
            opening_stock=opening_stock,
            incoming_stock=incoming_stock,
            outgoing_stock=outgoing_stock,
            closing_stock=closing_stock
        ))
    
    return movements


# Advanced Invoice Features - API Endpoints

@api.get('/currencies', response_model=list[CurrencyInfo])
def get_currencies(_: User = Depends(get_current_user)):
    """Get list of supported currencies"""
    supported_currencies = get_supported_currencies()
    currencies = []
    for code, info in supported_currencies.items():
        currencies.append(CurrencyInfo(
            code=code,
            symbol=info['symbol'],
            name=info['name']
        ))
    return currencies


@api.get('/exchange-rate/{from_currency}/{to_currency}', response_model=ExchangeRateResponse)
def get_exchange_rate_endpoint(
    from_currency: str,
    to_currency: str = "INR",
    _: User = Depends(get_current_user)
):
    """Get exchange rate between two currencies"""
    try:
        rate = get_exchange_rate(from_currency.upper(), to_currency.upper())
        return ExchangeRateResponse(
            from_currency=from_currency.upper(),
            to_currency=to_currency.upper(),
            rate=rate,
            last_updated=datetime.utcnow()
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@api.post('/recurring-invoice-templates', response_model=RecurringInvoiceTemplateOut, status_code=status.HTTP_201_CREATED)
def create_recurring_invoice_template(
    payload: RecurringInvoiceTemplateCreate,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new recurring invoice template"""
    # Validate customer and supplier exist
    customer = db.query(Party).filter(Party.id == payload.customer_id).first()
    supplier = db.query(Party).filter(Party.id == payload.supplier_id).first()
    
    if not customer:
        raise HTTPException(status_code=400, detail='Invalid customer')
    if not supplier:
        raise HTTPException(status_code=400, detail='Invalid supplier')
    
    # Calculate next generation date
    start_date = datetime.fromisoformat(payload.start_date.replace('Z', '+00:00'))
    next_generation_date = start_date
    
    template_data = payload.dict()
    template_data['start_date'] = start_date
    template_data['next_generation_date'] = next_generation_date
    
    if payload.end_date:
        template_data['end_date'] = datetime.fromisoformat(payload.end_date.replace('Z', '+00:00'))
    
    service = RecurringInvoiceService(db)
    template = service.create_template(template_data)
    
    return template


@api.post('/recurring-invoice-templates/{template_id}/items', response_model=RecurringInvoiceTemplateItemOut, status_code=status.HTTP_201_CREATED)
def add_recurring_invoice_template_item(
    template_id: int,
    payload: RecurringInvoiceTemplateItemCreate,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Add an item to a recurring invoice template"""
    # Validate template exists
    service = RecurringInvoiceService(db)
    template = service.get_template_by_id(template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Template not found')
    
    # Validate product exists
    product = db.query(Product).filter(Product.id == payload.product_id).first()
    if not product:
        raise HTTPException(status_code=400, detail='Invalid product')
    
    item = service.add_template_item(template_id, payload.dict())
    return item


@api.get('/recurring-invoice-templates', response_model=list[RecurringInvoiceTemplateOut])
def get_recurring_invoice_templates(
    active_only: bool = True,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all recurring invoice templates"""
    service = RecurringInvoiceService(db)
    if active_only:
        templates = service.get_active_templates()
    else:
        templates = db.query(RecurringInvoiceTemplate).all()
    
    return templates


@api.get('/recurring-invoice-templates/{template_id}', response_model=RecurringInvoiceTemplateOut)
def get_recurring_invoice_template(
    template_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific recurring invoice template"""
    service = RecurringInvoiceService(db)
    template = service.get_template_by_id(template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Template not found')
    
    return template


@api.get('/recurring-invoice-templates/{template_id}/items', response_model=list[RecurringInvoiceTemplateItemOut])
def get_recurring_invoice_template_items(
    template_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all items for a recurring invoice template"""
    service = RecurringInvoiceService(db)
    template = service.get_template_by_id(template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Template not found')
    
    items = service.get_template_items(template_id)
    return items


@api.put('/recurring-invoice-templates/{template_id}', response_model=RecurringInvoiceTemplateOut)
def update_recurring_invoice_template(
    template_id: int,
    payload: RecurringInvoiceTemplateCreate,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a recurring invoice template"""
    service = RecurringInvoiceService(db)
    template = service.get_template_by_id(template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Template not found')
    
    # Validate customer and supplier exist
    customer = db.query(Party).filter(Party.id == payload.customer_id).first()
    supplier = db.query(Party).filter(Party.id == payload.supplier_id).first()
    
    if not customer:
        raise HTTPException(status_code=400, detail='Invalid customer')
    if not supplier:
        raise HTTPException(status_code=400, detail='Invalid supplier')
    
    update_data = payload.dict()
    update_data['start_date'] = datetime.fromisoformat(payload.start_date.replace('Z', '+00:00'))
    
    if payload.end_date:
        update_data['end_date'] = datetime.fromisoformat(payload.end_date.replace('Z', '+00:00'))
    
    updated_template = service.update_template(template_id, update_data)
    return updated_template


@api.delete('/recurring-invoice-templates/{template_id}')
def deactivate_recurring_invoice_template(
    template_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Deactivate a recurring invoice template"""
    service = RecurringInvoiceService(db)
    success = service.deactivate_template(template_id)
    if not success:
        raise HTTPException(status_code=404, detail='Template not found')
    
    return {"message": "Template deactivated successfully"}


@api.post('/recurring-invoices/generate')
def generate_recurring_invoices_endpoint(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate invoices for all due recurring templates"""
    try:
        generated_invoices = generate_recurring_invoices(db)
        return {
            "message": f"Generated {len(generated_invoices)} invoices",
            "generated_invoices": [invoice.invoice_no for invoice in generated_invoices]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating invoices: {str(e)}")


@api.get('/recurring-invoices', response_model=list[RecurringInvoiceOut])
def get_recurring_invoices(
    template_id: int | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all recurring invoices, optionally filtered by template"""
    service = RecurringInvoiceService(db)
    recurring_invoices = service.get_recurring_invoices(template_id)
    return recurring_invoices


@api.put('/recurring-invoices/{recurring_invoice_id}/status')
def update_recurring_invoice_status(
    recurring_invoice_id: int,
    status: str,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update status of a recurring invoice"""
    service = RecurringInvoiceService(db)
    success = service.update_recurring_invoice_status(recurring_invoice_id, status)
    if not success:
        raise HTTPException(status_code=404, detail='Recurring invoice not found')
    
    return {"message": "Status updated successfully"}


# Purchase Order Management - API Endpoints

@api.post('/purchase-orders', response_model=PurchaseOrderOut, status_code=status.HTTP_201_CREATED)
def create_purchase_order(
    payload: PurchaseOrderCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new purchase order"""
    # Validate vendor exists
    vendor = db.query(Party).filter(Party.id == payload.vendor_id).first()
    if not vendor:
        raise HTTPException(status_code=400, detail='Invalid vendor')
    
    service = PurchaseOrderService(db)
    
    # Create PO data
    po_data = {
        'vendor_id': payload.vendor_id,
        'po_number': payload.po_number,
        'date': datetime.fromisoformat(payload.date.replace('Z', '+00:00')),
        'expected_delivery_date': datetime.fromisoformat(payload.expected_delivery_date.replace('Z', '+00:00')),
        'currency': payload.currency,
        'exchange_rate': payload.exchange_rate,
        'terms': payload.terms,
        'place_of_supply': payload.place_of_supply,
        'place_of_supply_state_code': payload.place_of_supply_state_code,
        'reverse_charge': payload.reverse_charge,
        'ship_from_address': payload.ship_from_address,
        'ship_to_address': payload.ship_to_address,
        'notes': payload.notes
    }
    
    # Create purchase order
    po = service.create_purchase_order(po_data)
    
    # Add items
    for item_data in payload.items:
        service.add_po_item(po.id, item_data.dict())
    
    # Refresh PO to get updated totals
    db.refresh(po)
    return po


@api.get('/purchase-orders', response_model=list[PurchaseOrderOut])
def get_purchase_orders(
    status: str | None = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get purchase orders, optionally filtered by status"""
    service = PurchaseOrderService(db)
    purchase_orders = service.get_purchase_orders(status)
    return purchase_orders


@api.get('/purchase-orders/{po_id}', response_model=PurchaseOrderOut)
def get_purchase_order(
    po_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific purchase order"""
    service = PurchaseOrderService(db)
    po = service.get_purchase_order(po_id)
    if not po:
        raise HTTPException(status_code=404, detail='Purchase order not found')
    
    return po


@api.get('/purchase-orders/{po_id}/items', response_model=list[PurchaseOrderItemOut])
def get_purchase_order_items(
    po_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all items for a purchase order"""
    service = PurchaseOrderService(db)
    po = service.get_purchase_order(po_id)
    if not po:
        raise HTTPException(status_code=404, detail='Purchase order not found')
    
    items = service.get_po_items(po_id)
    return items


@api.put('/purchase-orders/{po_id}/status')
def update_purchase_order_status(
    po_id: int,
    status: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update purchase order status"""
    service = PurchaseOrderService(db)
    success = service.update_po_status(po_id, status, current_user.id)
    if not success:
        raise HTTPException(status_code=400, detail='Invalid status transition or PO not found')
    
    return {"message": "Status updated successfully"}


@api.post('/purchase-orders/{po_id}/convert-to-purchase')
def convert_po_to_purchase_endpoint(
    po_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Convert a purchase order to a purchase invoice"""
    purchase = convert_po_to_purchase(db, po_id)
    if not purchase:
        raise HTTPException(status_code=400, detail='Cannot convert PO to purchase. PO must be approved.')
    
    return {
        "message": "Purchase order converted to purchase successfully",
        "purchase_id": purchase.id,
        "purchase_no": purchase.purchase_no
    }


# Payment Scheduler API Endpoints
@api.get('/payment-schedule')
def get_payment_schedule(
    payment_type: str = Query("all", description="Type of payments: invoice, purchase, or all"),
    status: str = Query(None, description="Payment status: pending, overdue, paid"),
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get payment schedule with filtering options"""
    scheduler = PaymentScheduler(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    # Parse status
    status_enum = None
    if status:
        try:
            status_enum = PaymentStatus(status)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid status value")
    
    schedule = scheduler.get_payment_schedule(
        payment_type=payment_type,
        status=status_enum,
        start_date=start_date_obj,
        end_date=end_date_obj
    )
    
    return schedule


@api.get('/payment-reminders')
def get_payment_reminders(
    reminder_type: str = Query(None, description="Reminder type: due_soon, overdue, critical"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get payment reminders based on due dates"""
    scheduler = PaymentScheduler(db)
    
    # Parse reminder type
    reminder_enum = None
    if reminder_type:
        try:
            reminder_enum = PaymentReminderType(reminder_type)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid reminder type")
    
    reminders = scheduler.get_payment_reminders(reminder_enum)
    return reminders


@api.get('/payment-analytics')
def get_payment_analytics(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get payment analytics and insights"""
    scheduler = PaymentScheduler(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    analytics = scheduler.get_payment_analytics(start_date_obj, end_date_obj)
    return analytics


# Inventory Management API Endpoints
@api.get('/inventory/summary')
def get_inventory_summary(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get comprehensive inventory summary"""
    manager = InventoryManager(db)
    return manager.get_inventory_summary()


@api.get('/inventory/low-stock-alerts')
def get_low_stock_alerts(
    threshold: int = Query(None, description="Custom threshold for low stock"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get low stock alerts"""
    manager = InventoryManager(db)
    return manager.get_low_stock_alerts(threshold)


@api.get('/inventory/stock-movements')
def get_stock_movements(
    product_id: int = Query(None, description="Filter by product ID"),
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    movement_type: str = Query(None, description="Movement type: in, out, adjust"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get stock movement history"""
    manager = InventoryManager(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    return manager.get_stock_movements(
        product_id=product_id,
        start_date=start_date_obj,
        end_date=end_date_obj,
        movement_type=movement_type
    )


@api.get('/inventory/analytics')
def get_inventory_analytics(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get inventory analytics"""
    manager = InventoryManager(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    return manager.get_inventory_analytics(start_date_obj, end_date_obj)


@api.get('/inventory/stock-value')
def get_stock_value(
    valuation_method: str = Query("fifo", description="Valuation method: fifo, lifo, average"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Calculate stock value using specified valuation method"""
    manager = InventoryManager(db)
    
    try:
        method = StockValuationMethod(valuation_method)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid valuation method")
    
    return manager.calculate_stock_value(method)


@api.get('/inventory/aging-report')
def get_stock_aging_report(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get stock aging report"""
    manager = InventoryManager(db)
    return manager.get_stock_aging_report()


# Financial Reports API Endpoints
@api.get('/financial-reports/profit-loss')
def get_profit_loss_statement(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate Profit & Loss Statement"""
    reports = FinancialReports(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    return reports.generate_profit_loss_statement(start_date_obj, end_date_obj)


@api.get('/financial-reports/balance-sheet')
def get_balance_sheet(
    as_of_date: str = Query(None, description="As of date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate Balance Sheet"""
    reports = FinancialReports(db)
    
    # Parse date
    as_of_date_obj = None
    if as_of_date:
        as_of_date_obj = datetime.strptime(as_of_date, "%Y-%m-%d").date()
    
    return reports.generate_balance_sheet(as_of_date_obj)


@api.get('/financial-reports/cash-flow')
def get_cash_flow_statement(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generate Cash Flow Statement"""
    reports = FinancialReports(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    return reports.generate_cash_flow_statement(start_date_obj, end_date_obj)


@api.get('/financial-reports/summary')
def get_financial_summary(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get comprehensive financial summary"""
    reports = FinancialReports(db)
    
    # Parse dates
    start_date_obj = None
    end_date_obj = None
    if start_date:
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    # Generate all reports
    pl_statement = reports.generate_profit_loss_statement(start_date_obj, end_date_obj)
    balance_sheet = reports.generate_balance_sheet(end_date_obj if end_date_obj else date.today())
    cash_flow = reports.generate_cash_flow_statement(start_date_obj, end_date_obj)
    
    return {
        "period": {
            "start_date": start_date_obj.isoformat() if start_date_obj else date.today().replace(day=1).isoformat(),
            "end_date": end_date_obj.isoformat() if end_date_obj else date.today().isoformat()
        },
        "profit_loss": pl_statement,
        "balance_sheet": balance_sheet,
        "cash_flow": cash_flow,
        "key_metrics": {
            "revenue": pl_statement["revenue"]["total_revenue"],
            "net_profit": pl_statement["net_profit_after_tax"],
            "total_assets": balance_sheet["assets"]["total_assets"],
            "total_liabilities": balance_sheet["liabilities"]["total_liabilities"],
            "net_cash_flow": cash_flow["net_cash_flow"],
            "cash_balance": cash_flow["cash_balances"]["closing_balance"]
        }
    }


# Invoice Template Management - API Endpoints

class InvoiceTemplateCreate(BaseModel):
    name: str
    description: str | None = None
    template_type: str = "professional"
    primary_color: str = "#2c3e50"
    secondary_color: str = "#3498db"
    accent_color: str = "#e74c3c"
    header_font: str = "Helvetica-Bold"
    body_font: str = "Helvetica"
    header_font_size: int = 18
    body_font_size: int = 10
    show_logo: bool = True
    logo_position: str = "top-left"
    show_company_details: bool = True
    show_customer_details: bool = True
    show_supplier_details: bool = True
    show_terms: bool = True
    show_notes: bool = True
    show_footer: bool = True
    header_text: str = "TAX INVOICE"
    footer_text: str = "Thank you for your business!"
    terms_text: str = "Payment is due within the terms specified above."


class InvoiceTemplateUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    template_type: str | None = None
    primary_color: str | None = None
    secondary_color: str | None = None
    accent_color: str | None = None
    header_font: str | None = None
    body_font: str | None = None
    header_font_size: int | None = None
    body_font_size: int | None = None
    show_logo: bool | None = None
    logo_position: str | None = None
    show_company_details: bool | None = None
    show_customer_details: bool | None = None
    show_supplier_details: bool | None = None
    show_terms: bool | None = None
    show_notes: bool | None = None
    show_footer: bool | None = None
    header_text: str | None = None
    footer_text: str | None = None
    terms_text: str | None = None


class InvoiceTemplateOut(BaseModel):
    id: int
    name: str
    description: str | None
    template_type: str
    primary_color: str
    secondary_color: str
    accent_color: str
    header_font: str
    body_font: str
    header_font_size: int
    body_font_size: int
    show_logo: bool
    logo_position: str
    show_company_details: bool
    show_customer_details: bool
    show_supplier_details: bool
    show_terms: bool
    show_notes: bool
    show_footer: bool
    header_text: str
    footer_text: str
    terms_text: str
    is_active: bool
    is_default: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


@api.post('/invoice-templates', response_model=InvoiceTemplateOut, status_code=status.HTTP_201_CREATED)
def create_invoice_template(
    payload: InvoiceTemplateCreate,
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Create a new invoice template"""
    # Validate template type
    valid_types = ["professional", "modern", "classic", "minimal"]
    if payload.template_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Template type must be one of: {', '.join(valid_types)}")
    
    # Validate logo position
    valid_positions = ["top-left", "top-right", "center"]
    if payload.logo_position not in valid_positions:
        raise HTTPException(status_code=400, detail=f"Logo position must be one of: {', '.join(valid_positions)}")
    
    # Validate colors (hex format)
    import re
    hex_pattern = r'^#[0-9A-Fa-f]{6}$'
    for color_field, color_value in [("primary_color", payload.primary_color), 
                                    ("secondary_color", payload.secondary_color), 
                                    ("accent_color", payload.accent_color)]:
        if not re.match(hex_pattern, color_value):
            raise HTTPException(status_code=400, detail=f"{color_field} must be a valid hex color (e.g., #2c3e50)")
    
    # If this is the first template, make it default
    existing_templates = db.query(InvoiceTemplate).filter(InvoiceTemplate.is_active == True).count()
    is_default = existing_templates == 0
    
    template = InvoiceTemplate(
        name=payload.name,
        description=payload.description,
        template_type=payload.template_type,
        primary_color=payload.primary_color,
        secondary_color=payload.secondary_color,
        accent_color=payload.accent_color,
        header_font=payload.header_font,
        body_font=payload.body_font,
        header_font_size=payload.header_font_size,
        body_font_size=payload.body_font_size,
        show_logo=payload.show_logo,
        logo_position=payload.logo_position,
        show_company_details=payload.show_company_details,
        show_customer_details=payload.show_customer_details,
        show_supplier_details=payload.show_supplier_details,
        show_terms=payload.show_terms,
        show_notes=payload.show_notes,
        show_footer=payload.show_footer,
        header_text=payload.header_text,
        footer_text=payload.footer_text,
        terms_text=payload.terms_text,
        is_default=is_default
    )
    
    db.add(template)
    db.commit()
    db.refresh(template)
    
    return template


@api.get('/invoice-templates/presets')
def get_preset_themes():
    """Get preset theme configurations"""
    presets = {
        "professional": {
            "name": "Professional Blue",
            "description": "Clean and professional blue theme",
            "primary_color": "#2c3e50",
            "secondary_color": "#3498db",
            "accent_color": "#e74c3c",
            "header_font": "Helvetica-Bold",
            "body_font": "Helvetica",
            "header_font_size": 18,
            "body_font_size": 10
        },
        "modern": {
            "name": "Modern Dark",
            "description": "Sleek modern dark theme",
            "primary_color": "#1a1a1a",
            "secondary_color": "#4a90e2",
            "accent_color": "#f39c12",
            "header_font": "Arial-Bold",
            "body_font": "Arial",
            "header_font_size": 20,
            "body_font_size": 11
        },
        "classic": {
            "name": "Classic Elegant",
            "description": "Timeless classic theme",
            "primary_color": "#2f4f4f",
            "secondary_color": "#708090",
            "accent_color": "#8b4513",
            "header_font": "Times-Bold",
            "body_font": "Times",
            "header_font_size": 16,
            "body_font_size": 10
        },
        "minimal": {
            "name": "Minimal Clean",
            "description": "Minimalist clean theme",
            "primary_color": "#333333",
            "secondary_color": "#666666",
            "accent_color": "#999999",
            "header_font": "Helvetica",
            "body_font": "Helvetica",
            "header_font_size": 14,
            "body_font_size": 9
        }
    }
    return presets


@api.get('/invoice-templates', response_model=list[InvoiceTemplateOut])
def get_invoice_templates(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all active invoice templates"""
    templates = db.query(InvoiceTemplate).filter(InvoiceTemplate.is_active == True).order_by(InvoiceTemplate.is_default.desc(), InvoiceTemplate.name).all()
    return templates


@api.get('/invoice-templates/default', response_model=InvoiceTemplateOut)
def get_default_invoice_template(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get the default invoice template"""
    template = db.query(InvoiceTemplate).filter(
        InvoiceTemplate.is_default == True,
        InvoiceTemplate.is_active == True
    ).first()
    
    if not template:
        # Create a default template if none exists
        template = InvoiceTemplate(
            name="Default Professional",
            description="Default professional invoice template",
            template_type="professional",
            is_default=True
        )
        db.add(template)
        db.commit()
        db.refresh(template)
    
    return template


@api.get('/invoice-templates/{template_id}', response_model=InvoiceTemplateOut)
def get_invoice_template(
    template_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific invoice template"""
    template = db.query(InvoiceTemplate).filter(InvoiceTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail='Invoice template not found')
    
    return template


@api.put('/invoice-templates/{template_id}', response_model=InvoiceTemplateOut)
def update_invoice_template(
    template_id: int,
    payload: InvoiceTemplateUpdate,
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Update an invoice template"""
    template = db.query(InvoiceTemplate).filter(InvoiceTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail='Invoice template not found')
    
    # Update fields
    update_data = payload.dict(exclude_unset=True)
    
    # Validate template type if provided
    if 'template_type' in update_data:
        valid_types = ["professional", "modern", "classic", "minimal"]
        if update_data['template_type'] not in valid_types:
            raise HTTPException(status_code=400, detail=f"Template type must be one of: {', '.join(valid_types)}")
    
    # Validate logo position if provided
    if 'logo_position' in update_data:
        valid_positions = ["top-left", "top-right", "center"]
        if update_data['logo_position'] not in valid_positions:
            raise HTTPException(status_code=400, detail=f"Logo position must be one of: {', '.join(valid_positions)}")
    
    # Validate colors if provided
    import re
    hex_pattern = r'^#[0-9A-Fa-f]{6}$'
    for color_field in ['primary_color', 'secondary_color', 'accent_color']:
        if color_field in update_data and not re.match(hex_pattern, update_data[color_field]):
            raise HTTPException(status_code=400, detail=f"{color_field} must be a valid hex color (e.g., #2c3e50)")
    
    for field, value in update_data.items():
        setattr(template, field, value)
    
    template.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(template)
    
    return template


@api.delete('/invoice-templates/{template_id}')
def delete_invoice_template(
    template_id: int,
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Delete an invoice template (soft delete)"""
    template = db.query(InvoiceTemplate).filter(InvoiceTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail='Invoice template not found')
    
    if template.is_default:
        raise HTTPException(status_code=400, detail='Cannot delete the default template')
    
    template.is_active = False
    template.updated_at = datetime.utcnow()
    db.commit()
    
    return {"message": "Template deleted successfully"}


@api.post('/invoice-templates/{template_id}/set-default')
def set_default_invoice_template(
    template_id: int,
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Set an invoice template as default"""
    template = db.query(InvoiceTemplate).filter(InvoiceTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail='Invoice template not found')
    
    # Remove default from all other templates
    db.query(InvoiceTemplate).update({InvoiceTemplate.is_default: False})
    
    # Set this template as default
    template.is_default = True
    template.updated_at = datetime.utcnow()
    db.commit()
    
    return {"message": "Default template updated successfully"}


@api.post('/upload-logo')
async def upload_logo(
    file: UploadFile = File(...),
    _: User = Depends(get_current_user)
):
    """Upload logo for invoice templates"""
    
    # Validate file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(
            status_code=400, 
            detail="Only image files are allowed"
        )
    
    # Validate file size (max 2MB)
    if file.size and file.size > 2 * 1024 * 1024:
        raise HTTPException(
            status_code=400, 
            detail="File size must be less than 2MB"
        )
    
    try:
        # Read file content
        content = await file.read()
        
        # Convert to base64 for storage
        base64_content = base64.b64encode(content).decode('utf-8')
        
        # Create logo data
        logo_data = {
            "filename": file.filename,
            "content_type": file.content_type,
            "size": len(content),
            "data": base64_content
        }
        
        # In a production environment, you would save to a file system or cloud storage
        # For now, we'll return the base64 data
        return {
            "success": True,
            "logo_url": f"data:{file.content_type};base64,{base64_content}",
            "filename": file.filename,
            "size": len(content)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to upload logo: {str(e)}"
        )


@api.post('/invoice-templates/import')
async def import_invoice_template(
    file: UploadFile = File(...),
    _: User = Depends(require_role("Admin")),
    db: Session = Depends(get_db)
):
    """Import invoice template from JSON file"""
    
    if not file.filename.endswith('.json'):
        raise HTTPException(
            status_code=400,
            detail="Only JSON files are allowed"
        )
    
    try:
        content = await file.read()
        template_data = json.loads(content.decode('utf-8'))
        
        # Validate required fields
        required_fields = ['name', 'description', 'template_type']
        for field in required_fields:
            if field not in template_data:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required field: {field}"
                )
        
        # Create new template
        template = InvoiceTemplate(**template_data)
        template.is_default = False  # Imported templates are not default
        template.is_active = True
        
        db.add(template)
        db.commit()
        db.refresh(template)
        
        return {
            "success": True,
            "message": "Template imported successfully",
            "template_id": template.id
        }
        
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=400,
            detail="Invalid JSON format"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to import template: {str(e)}"
        )


@api.get('/invoice-templates/{template_id}/export')
def export_invoice_template(
    template_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export invoice template as JSON file"""
    
    template = db.query(InvoiceTemplate).filter(InvoiceTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail='Invoice template not found')
    
    # Convert template to dict, excluding internal fields
    template_dict = {
        "name": template.name,
        "description": template.description,
        "template_type": template.template_type,
        "primary_color": template.primary_color,
        "secondary_color": template.secondary_color,
        "accent_color": template.accent_color,
        "header_font": template.header_font,
        "body_font": template.body_font,
        "header_font_size": template.header_font_size,
        "body_font_size": template.body_font_size,
        "show_logo": template.show_logo,
        "logo_position": template.logo_position,
        "show_company_details": template.show_company_details,
        "show_customer_details": template.show_customer_details,
        "show_supplier_details": template.show_supplier_details,
        "show_terms": template.show_terms,
        "show_notes": template.show_notes,
        "show_footer": template.show_footer,
        "header_text": template.header_text,
        "footer_text": template.footer_text,
        "terms_text": template.terms_text
    }
    
    return Response(
        content=json.dumps(template_dict, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=template_{template_id}.json"}
    )


@api.get('/dashboard')
def get_dashboard(
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get dashboard data with quick links"""
    
    quick_links = [
        {
            "text": "Add Product",
            "url": "/api/products",
            "icon": "product",
            "description": "Create a new product"
        },
        {
            "text": "Add Invoice", 
            "url": "/api/invoices",
            "icon": "invoice",
            "description": "Create a new invoice"
        },
        {
            "text": "Add Purchase",
            "url": "/api/purchases", 
            "icon": "purchase",
            "description": "Create a new purchase"
        },
        {
            "text": "Add Expense",
            "url": "/api/expenses",
            "icon": "expense", 
            "description": "Create a new expense"
        }
    ]
    
    return {
        "quick_links": quick_links,
        "dashboard_type": "main"
    }


# Include branding router
api.include_router(branding_router)

# Include dental router
api.include_router(dental_router)

# Include manufacturing router
api.include_router(manufacturing_router)




